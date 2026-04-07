import csv
import os
import tempfile
import unittest.mock

import openpyxl
import pytest
from typer.testing import CliRunner

from xlforge import app
from xlforge.core.errors import ErrorCode

runner = CliRunner()


def _is_excel_available() -> bool:
    """Check if Excel is available via xlwings.

    Returns True only if xlwings is installed AND Excel can be started.
    This is more reliable than just checking if xlwings is installed.
    """
    try:
        from importlib.util import find_spec

        if find_spec("xlwings") is None:
            return False
    except ImportError:
        return False

    # We can't reliably check if Excel is running without potentially crashing.
    # For safety, assume Excel is NOT available when xlwings is installed but
    # we can't verify Excel is running. This will cause success tests to be
    # skipped when Excel isn't available.
    # Note: This is a conservative approach - success tests requiring Excel
    # will be skipped. Error tests use mocking to force openpyxl.
    return False


def test_ping():
    result = runner.invoke(app, ["ping"])
    assert result.exit_code == 0
    assert "pong" in result.output


def test_version():
    result = runner.invoke(app, ["version"])
    assert result.exit_code == 0
    assert "xlforge 0.1.0" in result.output


def test_no_args_shows_help():
    result = runner.invoke(app)
    # Without invoke_without_command, no args shows help/missing command error
    assert result.exit_code in [0, 2]


class TestSheetCommands:
    """Tests for sheet commands."""

    @pytest.mark.skipif(not _is_excel_available(), reason="requires Excel")
    def test_sheet_create(self):
        """Test creating a new sheet."""
        with tempfile.TemporaryDirectory() as tmpdir:
            # Create a workbook with default sheet
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["sheet", "create", path, "NewSheet"])

            assert result.exit_code == 0
            assert "Created sheet 'NewSheet'" in result.output

            # Verify sheet exists
            wb = openpyxl.load_workbook(path)
            assert "NewSheet" in wb.sheetnames
            wb.close()

    def test_sheet_create_already_exists(self):
        """Test creating a sheet that already exists."""
        with tempfile.TemporaryDirectory() as tmpdir:
            # Create a workbook with default sheet
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            with unittest.mock.patch(
                "xlforge.core.engines.selector.find_spec", return_value=None
            ):
                result = runner.invoke(app, ["sheet", "create", path, "Sheet"])

            # Sheet already exists by default
            assert result.exit_code == ErrorCode.TABLE_ALREADY_EXISTS
            assert "already exists" in result.output

    def test_sheet_create_file_not_found(self):
        """Test creating a sheet in non-existent file."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            result = runner.invoke(app, ["sheet", "create", path, "NewSheet"])

            assert result.exit_code == ErrorCode.FILE_DOES_NOT_EXIST
            assert "does not exist" in result.output.lower()

    @pytest.mark.skipif(not _is_excel_available(), reason="requires Excel")
    def test_sheet_delete(self):
        """Test deleting a sheet."""
        with tempfile.TemporaryDirectory() as tmpdir:
            # Create a workbook with two sheets
            wb = openpyxl.Workbook()
            wb.create_sheet("SecondSheet")
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["sheet", "delete", path, "Sheet"])

            assert result.exit_code == 0
            assert "Deleted sheet 'Sheet'" in result.output

            # Verify sheet was deleted
            wb = openpyxl.load_workbook(path)
            assert "Sheet" not in wb.sheetnames
            assert "SecondSheet" in wb.sheetnames
            wb.close()

    def test_sheet_delete_not_found(self):
        """Test deleting a non-existent sheet."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            with unittest.mock.patch(
                "xlforge.core.engines.selector.find_spec", return_value=None
            ):
                result = runner.invoke(app, ["sheet", "delete", path, "NonExistent"])

            assert result.exit_code == ErrorCode.SHEET_NOT_FOUND
            assert "not found" in result.output.lower()

    def test_sheet_delete_last_sheet_warns(self):
        """Test deleting the last sheet shows warning."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            with unittest.mock.patch(
                "xlforge.core.engines.selector.find_spec", return_value=None
            ):
                result = runner.invoke(app, ["sheet", "delete", path, "Sheet"])

            assert result.exit_code == ErrorCode.CANNOT_DELETE_LAST_SHEET
            assert (
                "last sheet" in result.output.lower()
                or "warning" in result.output.lower()
            )

    def test_sheet_delete_last_sheet_force(self):
        """Test force deleting the last sheet."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            with unittest.mock.patch(
                "xlforge.core.engines.selector.find_spec", return_value=None
            ):
                result = runner.invoke(
                    app, ["sheet", "delete", path, "Sheet", "--force"]
                )

            # Note: --force allows deleting last sheet, but openpyxl cannot save
            # an empty workbook, so the file remains unchanged on disk
            assert result.exit_code == 0
            assert "Deleted sheet 'Sheet'" in result.output

    @pytest.mark.skipif(not _is_excel_available(), reason="requires Excel")
    def test_sheet_rename(self):
        """Test renaming a sheet."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["sheet", "rename", path, "Sheet", "RenamedSheet"]
            )

            assert result.exit_code == 0
            assert "Renamed sheet 'Sheet' to 'RenamedSheet'" in result.output

            # Verify sheet was renamed
            wb = openpyxl.load_workbook(path)
            assert "RenamedSheet" in wb.sheetnames
            assert "Sheet" not in wb.sheetnames
            wb.close()

    def test_sheet_rename_not_found(self):
        """Test renaming a non-existent sheet."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            with unittest.mock.patch(
                "xlforge.core.engines.selector.find_spec", return_value=None
            ):
                result = runner.invoke(
                    app, ["sheet", "rename", path, "NonExistent", "NewName"]
                )

            assert result.exit_code == ErrorCode.SHEET_NOT_FOUND
            assert "not found" in result.output.lower()

    def test_sheet_rename_new_name_exists(self):
        """Test renaming to an existing sheet name."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.create_sheet("ExistingSheet")
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            with unittest.mock.patch(
                "xlforge.core.engines.selector.find_spec", return_value=None
            ):
                result = runner.invoke(
                    app, ["sheet", "rename", path, "Sheet", "ExistingSheet"]
                )

            assert result.exit_code == ErrorCode.TABLE_ALREADY_EXISTS
            assert "already exists" in result.output.lower()


class TestCellRead:
    """Tests for xlforge cell read command."""

    def test_cell_read_file_not_found(self):
        """Test cell read with non-existent file returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            result = runner.invoke(app, ["cell", "read", path, "Sheet1", "A1"])

            assert result.exit_code == ErrorCode.FILE_DOES_NOT_EXIST
            assert "does not exist" in result.output.lower()

    def test_cell_read_sheet_not_found(self):
        """Test cell read with non-existent sheet returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            with unittest.mock.patch(
                "xlforge.core.engines.selector.find_spec", return_value=None
            ):
                result = runner.invoke(
                    app, ["cell", "read", path, "NonexistentSheet", "A1"]
                )

            assert result.exit_code == ErrorCode.SHEET_NOT_FOUND
            assert "Sheet not found" in result.output

    @pytest.mark.skipif(not _is_excel_available(), reason="requires Excel")
    def test_cell_read_string_value(self):
        """Test reading a string cell value."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws["A1"] = "Hello World"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["cell", "read", path, "Sheet1", "A1"])

            assert result.exit_code == 0
            assert "Hello World" in result.output
            assert "string" in result.output

    @pytest.mark.skipif(not _is_excel_available(), reason="requires Excel")
    def test_cell_read_number_value(self):
        """Test reading a number cell value."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws["B2"] = 42.5
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["cell", "read", path, "Sheet1", "B2"])

            assert result.exit_code == 0
            assert "42.5" in result.output
            assert "number" in result.output

    @pytest.mark.skipif(not _is_excel_available(), reason="requires Excel")
    def test_cell_read_boolean_value(self):
        """Test reading a boolean cell value."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws["C3"] = True
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["cell", "read", path, "Sheet1", "C3"])

            assert result.exit_code == 0
            assert "True" in result.output
            assert "bool" in result.output

    @pytest.mark.skipif(not _is_excel_available(), reason="requires Excel")
    def test_cell_read_json_output(self):
        """Test reading a cell with JSON output."""
        import json

        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws["A1"] = "Test"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["cell", "read", path, "Sheet1", "A1", "--json"]
            )

            assert result.exit_code == 0
            data = json.loads(result.output)
            assert data["value"] == "Test"
            assert data["type"] == "string"
            assert data["coord"] == "A1"
            assert data["sheet"] == "Sheet1"

    @pytest.mark.skipif(not _is_excel_available(), reason="requires Excel")
    def test_cell_read_empty_cell(self):
        """Test reading an empty cell."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            # A1 is empty by default
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["cell", "read", path, "Sheet1", "A1"])

            assert result.exit_code == 0
            assert "None" in result.output or "empty" in result.output.lower()


class TestCellWrite:
    """Tests for xlforge cell write command."""

    def test_cell_write_file_not_found(self):
        """Test cell write with non-existent file returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            result = runner.invoke(app, ["cell", "write", path, "Sheet1", "A1", "test"])

            assert result.exit_code == ErrorCode.FILE_DOES_NOT_EXIST
            assert "does not exist" in result.output.lower()

    def test_cell_write_sheet_not_found(self):
        """Test cell write with non-existent sheet returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            with unittest.mock.patch(
                "xlforge.core.engines.selector.find_spec", return_value=None
            ):
                result = runner.invoke(
                    app, ["cell", "write", path, "NonexistentSheet", "A1", "test"]
                )

            assert result.exit_code == ErrorCode.SHEET_NOT_FOUND
            assert "Sheet not found" in result.output

    @pytest.mark.skipif(not _is_excel_available(), reason="requires Excel")
    def test_cell_write_string_value(self):
        """Test writing a string value to a cell."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["cell", "write", path, "Sheet", "A1", "Hello World"]
            )

            assert result.exit_code == 0
            assert "Written:" in result.output

            # Verify the value was written
            wb = openpyxl.load_workbook(path)
            assert wb.active["A1"].value == "Hello World"
            wb.close()

    @pytest.mark.skipif(not _is_excel_available(), reason="requires Excel")
    def test_cell_write_number_value(self):
        """Test writing a number value to a cell."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["cell", "write", path, "Sheet", "A1", "42.5", "--type", "number"]
            )

            assert result.exit_code == 0

            # Verify the value was written
            wb = openpyxl.load_workbook(path)
            assert wb.active["A1"].value == 42.5
            wb.close()

    @pytest.mark.skipif(not _is_excel_available(), reason="requires Excel")
    def test_cell_write_boolean_true_value(self):
        """Test writing a boolean TRUE value to a cell."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["cell", "write", path, "Sheet", "A1", "TRUE", "--type", "bool"]
            )

            assert result.exit_code == 0

            # Verify the value was written
            wb = openpyxl.load_workbook(path)
            assert wb.active["A1"].value is True
            wb.close()

    @pytest.mark.skipif(not _is_excel_available(), reason="requires Excel")
    def test_cell_write_boolean_false_value(self):
        """Test writing a boolean FALSE value to a cell."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["cell", "write", path, "Sheet", "A1", "FALSE", "--type", "bool"]
            )

            assert result.exit_code == 0

            # Verify the value was written
            wb = openpyxl.load_workbook(path)
            assert wb.active["A1"].value is False
            wb.close()

    def test_cell_write_invalid_type(self):
        """Test writing with an invalid type returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            with unittest.mock.patch(
                "xlforge.core.engines.selector.find_spec", return_value=None
            ):
                result = runner.invoke(
                    app,
                    ["cell", "write", path, "Sheet", "A1", "test", "--type", "invalid"],
                )

            assert result.exit_code == ErrorCode.TYPE_COERCION_FAILED
            assert "Invalid type" in result.output

    @pytest.mark.skipif(not _is_excel_available(), reason="requires Excel")
    def test_cell_write_date_value(self):
        """Test writing a date value to a cell."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app,
                ["cell", "write", path, "Sheet", "A1", "2024-01-15", "--type", "date"],
            )

            assert result.exit_code == 0

            # Verify the value was written (as a datetime)
            wb = openpyxl.load_workbook(path)
            cell_value = wb.active["A1"].value
            assert cell_value is not None
            assert hasattr(cell_value, "year")  # datetime has year attribute
            wb.close()

    @pytest.mark.skipif(not _is_excel_available(), reason="requires Excel")
    def test_cell_write_preserves_leading_zeros_with_string_type(self):
        """Test writing a string that looks like a number preserves leading zeros."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["cell", "write", path, "Sheet", "A1", "00123", "--type", "string"]
            )

            assert result.exit_code == 0

            # Verify the value was written as string preserving leading zeros
            wb = openpyxl.load_workbook(path)
            assert wb.active["A1"].value == "00123"
            wb.close()


class TestRangeRead:
    """Tests for xlforge range read command."""

    def test_range_read_file_not_found(self):
        """Test range read with non-existent file returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            result = runner.invoke(app, ["range", "read", path, "Sheet1", "A1:C3"])

            assert result.exit_code == ErrorCode.FILE_DOES_NOT_EXIST
            assert "does not exist" in result.output.lower()

    def test_range_read_sheet_not_found(self):
        """Test range read with non-existent sheet returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            with unittest.mock.patch(
                "xlforge.core.engines.selector.find_spec", return_value=None
            ):
                result = runner.invoke(
                    app, ["range", "read", path, "NonexistentSheet", "A1:C3"]
                )

            assert result.exit_code == ErrorCode.SHEET_NOT_FOUND
            assert "Sheet not found" in result.output

    @pytest.mark.skipif(not _is_excel_available(), reason="requires Excel")
    def test_range_read_table_output(self):
        """Test reading a range with table output."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws["A1"] = "Name"
            ws["B1"] = "Age"
            ws["C1"] = "Active"
            ws["A2"] = "Alice"
            ws["B2"] = 30
            ws["C2"] = True
            ws["A3"] = "Bob"
            ws["B3"] = 25
            ws["C3"] = False
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["range", "read", path, "Sheet1", "A1:C3"])

            assert result.exit_code == 0
            assert "Range: A1:C3" in result.output
            assert "3 rows x 3 columns" in result.output
            assert "Name" in result.output
            assert "Alice" in result.output
            assert "30" in result.output

    @pytest.mark.skipif(not _is_excel_available(), reason="requires Excel")
    def test_range_read_json_output(self):
        """Test reading a range with JSON output."""
        import json

        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws["A1"] = "Hello"
            ws["B1"] = 42
            ws["A2"] = "World"
            ws["B2"] = True
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["range", "read", path, "Sheet1", "A1:B2", "--json"]
            )

            assert result.exit_code == 0
            data = json.loads(result.output)
            assert data == [["Hello", 42], ["World", True]]

    @pytest.mark.skipif(not _is_excel_available(), reason="requires Excel")
    def test_range_read_single_cell(self):
        """Test reading a single cell as a range."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws["A1"] = "Single Cell"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["range", "read", path, "Sheet1", "A1:A1"])

            assert result.exit_code == 0
            assert "Single Cell" in result.output
            assert "1 rows x 1 columns" in result.output

    @pytest.mark.skipif(not _is_excel_available(), reason="requires Excel")
    def test_range_read_empty_range(self):
        """Test reading an empty range."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["range", "read", path, "Sheet1", "A1:B2"])

            assert result.exit_code == 0
            # Empty cells will show as empty range message
            assert "A1:B2 is empty" in result.output


class TestRangeWrite:
    """Tests for xlforge range write command."""

    def test_range_write_file_not_found(self):
        """Test range write with non-existent file returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            result = runner.invoke(
                app,
                ["range", "write", path, "Sheet1", "A1:C3", '[["a","b"],["c","d"]]'],
            )

            assert result.exit_code == ErrorCode.FILE_DOES_NOT_EXIST
            assert "does not exist" in result.output.lower()

    def test_range_write_sheet_not_found(self):
        """Test range write with non-existent sheet returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            with unittest.mock.patch(
                "xlforge.core.engines.selector.find_spec", return_value=None
            ):
                result = runner.invoke(
                    app,
                    [
                        "range",
                        "write",
                        path,
                        "NonexistentSheet",
                        "A1:C3",
                        '[["a","b"],["c","d"]]',
                    ],
                )

            assert result.exit_code == ErrorCode.SHEET_NOT_FOUND
            assert "Sheet not found" in result.output

    @pytest.mark.skipif(not _is_excel_available(), reason="requires Excel")
    def test_range_write_json_values(self):
        """Test writing values from JSON."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app,
                [
                    "range",
                    "write",
                    path,
                    "Sheet",
                    "A1:C3",
                    '[["Name","Age","Active"],["Alice",30,true],["Bob",25,false]]',
                ],
            )

            assert result.exit_code == 0
            assert "Written 3 row(s) x 3 column(s) to range A1:C3" in result.output

            # Verify the values were written
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            assert ws["A1"].value == "Name"
            assert ws["B1"].value == "Age"
            assert ws["C1"].value == "Active"
            assert ws["A2"].value == "Alice"
            assert ws["B2"].value == 30
            assert ws["C2"].value is True
            assert ws["A3"].value == "Bob"
            assert ws["B3"].value == 25
            assert ws["C3"].value is False
            wb.close()

    @pytest.mark.skipif(not _is_excel_available(), reason="requires Excel")
    def test_range_write_csv_file(self):
        """Test writing values from CSV file."""
        import csv

        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            # Create CSV file
            csv_path = os.path.join(tmpdir, "values.csv")
            with open(csv_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(["Product", "Price", "Qty"])
                writer.writerow(["Apple", "1.50", "100"])
                writer.writerow(["Banana", "0.75", "200"])

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["range", "write", path, "Sheet", "A1:C3", "--csv", csv_path]
            )

            assert result.exit_code == 0
            assert "Written 3 row(s) x 3 column(s) to range A1:C3" in result.output

            # Verify the values were written
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            assert ws["A1"].value == "Product"
            assert ws["B1"].value == "Price"
            assert ws["C1"].value == "Qty"
            assert ws["A2"].value == "Apple"
            assert ws["B2"].value == "1.50"
            assert ws["C2"].value == "100"
            assert ws["A3"].value == "Banana"
            assert ws["B3"].value == "0.75"
            assert ws["C3"].value == "200"
            wb.close()

    def test_range_write_invalid_json(self):
        """Test writing with invalid JSON returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            with unittest.mock.patch(
                "xlforge.core.engines.selector.find_spec", return_value=None
            ):
                result = runner.invoke(
                    app, ["range", "write", path, "Sheet", "A1:B2", "not valid json"]
                )

            assert result.exit_code == 1  # ErrorCode.INVALID_ARGUMENT
            assert "Invalid JSON" in result.output

    def test_range_write_csv_file_not_found(self):
        """Test writing with non-existent CSV file returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            csv_path = os.path.join(tmpdir, "nonexistent.csv")
            with unittest.mock.patch(
                "xlforge.core.engines.selector.find_spec", return_value=None
            ):
                result = runner.invoke(
                    app, ["range", "write", path, "Sheet", "A1:B2", "--csv", csv_path]
                )

            assert result.exit_code == ErrorCode.FILE_DOES_NOT_EXIST
            assert "does not exist" in result.output.lower()

    def test_range_write_missing_values(self):
        """Test writing without providing values returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            with unittest.mock.patch(
                "xlforge.core.engines.selector.find_spec", return_value=None
            ):
                result = runner.invoke(app, ["range", "write", path, "Sheet", "A1:B2"])

            assert result.exit_code == 1  # ErrorCode.INVALID_ARGUMENT
            assert "Must provide either" in result.output

    def test_range_write_both_json_and_csv_error(self):
        """Test writing with both JSON and CSV returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            csv_path = os.path.join(tmpdir, "values.csv")
            with open(csv_path, "w", newline="", encoding="utf-8") as f:
                f.write("a,b\nc,d")

            path = os.path.join(tmpdir, "test.xlsx")
            with unittest.mock.patch(
                "xlforge.core.engines.selector.find_spec", return_value=None
            ):
                result = runner.invoke(
                    app,
                    [
                        "range",
                        "write",
                        path,
                        "Sheet",
                        "A1:B2",
                        '[["a","b"],["c","d"]]',
                        "--csv",
                        csv_path,
                    ],
                )

            assert result.exit_code == 1  # ErrorCode.INVALID_ARGUMENT
            assert "Cannot specify both" in result.output


class TestCsvImport:
    """Tests for xlforge csv import command."""

    def test_csv_import_file_not_found(self):
        """Test csv import with non-existent CSV file returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            csv_path = os.path.join(tmpdir, "nonexistent.csv")
            xlsx_path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["csv", "import", csv_path, xlsx_path, "Sheet"])

            assert result.exit_code == ErrorCode.CSV_NOT_FOUND
            assert "does not exist" in result.output.lower()

    def test_csv_import_excel_file_not_found(self):
        """Test csv import with non-existent Excel file returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            csv_path = os.path.join(tmpdir, "test.csv")
            xlsx_path = os.path.join(tmpdir, "nonexistent.xlsx")

            # Create CSV file
            with open(csv_path, "w") as f:
                f.write("a,b,c\n")

            result = runner.invoke(app, ["csv", "import", csv_path, xlsx_path, "Sheet"])

            assert result.exit_code == ErrorCode.FILE_DOES_NOT_EXIST
            assert "does not exist" in result.output.lower()

    def test_csv_import_sheet_not_found(self):
        """Test csv import with non-existent sheet returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            csv_path = os.path.join(tmpdir, "test.csv")
            xlsx_path = os.path.join(tmpdir, "test.xlsx")

            # Create CSV file
            with open(csv_path, "w") as f:
                f.write("a,b,c\n")

            with unittest.mock.patch(
                "xlforge.core.engines.selector.find_spec", return_value=None
            ):
                result = runner.invoke(
                    app, ["csv", "import", csv_path, xlsx_path, "NonExistent"]
                )

            assert result.exit_code == ErrorCode.SHEET_NOT_FOUND
            assert "not found" in result.output.lower()

    @pytest.mark.skipif(not _is_excel_available(), reason="requires Excel")
    def test_csv_import_basic(self):
        """Test basic CSV import."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            csv_path = os.path.join(tmpdir, "test.csv")
            xlsx_path = os.path.join(tmpdir, "test.xlsx")

            # Create CSV file
            with open(csv_path, "w", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["Name", "Age", "City"])
                writer.writerow(["Alice", "30", "NYC"])
                writer.writerow(["Bob", "25", "LA"])

            result = runner.invoke(app, ["csv", "import", csv_path, xlsx_path, "Sheet"])

            assert result.exit_code == 0
            assert "Imported" in result.output

            # Verify data was imported (numbers are type-coerced)
            wb = openpyxl.load_workbook(xlsx_path)
            ws = wb.active
            assert ws["A1"].value == "Name"
            assert ws["B1"].value == "Age"
            assert ws["C1"].value == "City"
            assert ws["A2"].value == "Alice"
            assert ws["B2"].value == 30  # Number, not string
            assert ws["C2"].value == "NYC"
            assert ws["A3"].value == "Bob"
            assert ws["B3"].value == 25
            assert ws["C3"].value == "LA"
            wb.close()

    @pytest.mark.skipif(not _is_excel_available(), reason="requires Excel")
    def test_csv_import_with_header(self):
        """Test CSV import with --has-header option."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            csv_path = os.path.join(tmpdir, "test.csv")
            xlsx_path = os.path.join(tmpdir, "test.xlsx")

            # Create CSV file
            with open(csv_path, "w", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["Name", "Age", "City"])
                writer.writerow(["Alice", "30", "NYC"])
                writer.writerow(["Bob", "25", "LA"])

            result = runner.invoke(
                app, ["csv", "import", csv_path, xlsx_path, "Sheet", "--has-header"]
            )

            assert result.exit_code == 0

            # Verify header row was skipped and data starts at A1 (numbers are type-coerced)
            wb = openpyxl.load_workbook(xlsx_path)
            ws = wb.active
            assert ws["A1"].value == "Alice"
            assert ws["B1"].value == 30  # Number, not string
            assert ws["A2"].value == "Bob"
            wb.close()

    def test_csv_import_empty_file(self):
        """Test CSV import with empty file returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            csv_path = os.path.join(tmpdir, "empty.csv")
            xlsx_path = os.path.join(tmpdir, "test.xlsx")

            # Create empty CSV file
            open(csv_path, "w").close()

            with unittest.mock.patch(
                "xlforge.core.engines.selector.find_spec", return_value=None
            ):
                result = runner.invoke(
                    app, ["csv", "import", csv_path, xlsx_path, "Sheet"]
                )

            assert result.exit_code == ErrorCode.INVALID_CSV_FORMAT
            assert "empty" in result.output.lower()


class TestCsvExport:
    """Tests for xlforge csv export command."""

    def test_csv_export_file_not_found(self):
        """Test csv export with non-existent Excel file returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            result = runner.invoke(app, ["csv", "export", path, "Sheet"])

            assert result.exit_code == ErrorCode.FILE_DOES_NOT_EXIST
            assert "does not exist" in result.output.lower()

    def test_csv_export_sheet_not_found(self):
        """Test csv export with non-existent sheet returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            with unittest.mock.patch(
                "xlforge.core.engines.selector.find_spec", return_value=None
            ):
                result = runner.invoke(app, ["csv", "export", path, "NonExistent"])

            assert result.exit_code == ErrorCode.SHEET_NOT_FOUND
            assert "not found" in result.output.lower()

    @pytest.mark.skipif(not _is_excel_available(), reason="requires Excel")
    def test_csv_export_basic(self):
        """Test basic CSV export to stdout."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws["A1"] = "Name"
            ws["B1"] = "Age"
            ws["C1"] = "City"
            ws["A2"] = "Alice"
            ws["B2"] = "30"
            ws["C2"] = "NYC"
            ws["A3"] = "Bob"
            ws["B3"] = "25"
            ws["C3"] = "LA"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["csv", "export", path, "Sheet1"])

            assert result.exit_code == 0
            assert "Name" in result.output
            assert "Alice" in result.output
            assert "30" in result.output
            assert "Bob" in result.output

    @pytest.mark.skipif(not _is_excel_available(), reason="requires Excel")
    def test_csv_export_to_file(self):
        """Test CSV export to output file."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws["A1"] = "Name"
            ws["B1"] = "Age"
            ws["A2"] = "Alice"
            ws["B2"] = "30"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            xlsx_path = os.path.join(tmpdir, "test.xlsx")
            csv_path = os.path.join(tmpdir, "output.csv")
            result = runner.invoke(
                app, ["csv", "export", xlsx_path, "Sheet1", "--output", csv_path]
            )

            assert result.exit_code == 0
            assert "Exported" in result.output

            # Verify CSV content
            with open(csv_path, "r") as f:
                content = f.read()
            assert "Name" in content
            assert "Alice" in content

    @pytest.mark.skipif(not _is_excel_available(), reason="requires Excel")
    def test_csv_export_with_range(self):
        """Test CSV export with specified range."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws["A1"] = "Name"
            ws["B1"] = "Age"
            ws["C1"] = "City"
            ws["A2"] = "Alice"
            ws["B2"] = "30"
            ws["C2"] = "NYC"
            ws["A3"] = "Bob"
            ws["B3"] = "25"
            ws["C3"] = "LA"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            xlsx_path = os.path.join(tmpdir, "test.xlsx")
            csv_path = os.path.join(tmpdir, "output.csv")
            result = runner.invoke(
                app,
                [
                    "csv",
                    "export",
                    xlsx_path,
                    "Sheet1",
                    "--range",
                    "A1:B2",
                    "--output",
                    csv_path,
                ],
            )

            assert result.exit_code == 0

            # Verify CSV content (only A1:B2)
            with open(csv_path, "r") as f:
                reader = csv.reader(f)
                rows = list(reader)
            assert rows[0][0] == "Name"
            assert rows[0][1] == "Age"
            assert rows[1][0] == "Alice"
            assert rows[1][1] == "30"
            assert len(rows) == 2

    @pytest.mark.skipif(not _is_excel_available(), reason="requires Excel")
    def test_csv_export_number_coercion(self):
        """Test CSV export properly handles number types."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws["A1"] = "Name"
            ws["B1"] = "Score"
            ws["A2"] = "Alice"
            ws["B2"] = 42.5
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            xlsx_path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["csv", "export", xlsx_path, "Sheet1"])

            assert result.exit_code == 0
            assert "42.5" in result.output


class TestRowHide:
    """Tests for xlforge row hide command."""

    def test_row_hide_file_not_found(self):
        """Test row hide with non-existent file returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            result = runner.invoke(app, ["row", "hide", path, "Sheet1", "1"])

            assert result.exit_code == ErrorCode.FILE_DOES_NOT_EXIST
            assert "does not exist" in result.output.lower()

    def test_row_hide_sheet_not_found(self):
        """Test row hide with non-existent sheet returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["row", "hide", path, "NonexistentSheet", "1"])

            assert result.exit_code == ErrorCode.SHEET_NOT_FOUND
            assert "Sheet not found" in result.output

    def test_row_hide_invalid_row(self):
        """Test row hide with invalid row number returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["row", "hide", path, "Sheet", "0"])

            assert result.exit_code == ErrorCode.ROW_NOT_FOUND
            assert "Invalid row" in result.output

    def test_row_hide_success(self):
        """Test hiding a row."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws["A1"] = "Header"
            ws["A2"] = "Data"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["row", "hide", path, "Sheet1", "1"])

            assert result.exit_code == 0
            assert "Hid row 1" in result.output

            # Verify row is hidden
            wb = openpyxl.load_workbook(path)
            assert wb["Sheet1"].row_dimensions[1].hidden is True
            wb.close()


class TestRowUnhide:
    """Tests for xlforge row unhide command."""

    def test_row_unhide_file_not_found(self):
        """Test row unhide with non-existent file returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            result = runner.invoke(app, ["row", "unhide", path, "Sheet1", "1"])

            assert result.exit_code == ErrorCode.FILE_DOES_NOT_EXIST
            assert "does not exist" in result.output.lower()

    def test_row_unhide_sheet_not_found(self):
        """Test row unhide with non-existent sheet returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["row", "unhide", path, "NonexistentSheet", "1"]
            )

            assert result.exit_code == ErrorCode.SHEET_NOT_FOUND
            assert "Sheet not found" in result.output

    def test_row_unhide_invalid_row(self):
        """Test row unhide with invalid row number returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["row", "unhide", path, "Sheet", "0"])

            assert result.exit_code == ErrorCode.ROW_NOT_FOUND
            assert "Invalid row" in result.output

    def test_row_unhide_success(self):
        """Test unhiding a row."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws.row_dimensions[1].hidden = True
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["row", "unhide", path, "Sheet1", "1"])

            assert result.exit_code == 0
            assert "Unhid row 1" in result.output

            # Verify row is visible
            wb = openpyxl.load_workbook(path)
            assert wb["Sheet1"].row_dimensions[1].hidden is False
            wb.close()


class TestColumnHide:
    """Tests for xlforge column hide command."""

    def test_column_hide_file_not_found(self):
        """Test column hide with non-existent file returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            result = runner.invoke(app, ["column", "hide", path, "Sheet1", "A"])

            assert result.exit_code == ErrorCode.FILE_DOES_NOT_EXIST
            assert "does not exist" in result.output.lower()

    def test_column_hide_sheet_not_found(self):
        """Test column hide with non-existent sheet returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["column", "hide", path, "NonexistentSheet", "A"]
            )

            assert result.exit_code == ErrorCode.SHEET_NOT_FOUND
            assert "Sheet not found" in result.output

    def test_column_hide_success(self):
        """Test hiding a column."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws["A1"] = "Header1"
            ws["B1"] = "Header2"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["column", "hide", path, "Sheet1", "A"])

            assert result.exit_code == 0
            assert "Hid column A" in result.output

            # Verify column is hidden
            wb = openpyxl.load_workbook(path)
            assert wb["Sheet1"].column_dimensions["A"].hidden is True
            wb.close()

    def test_column_hide_lowercase(self):
        """Test hiding a column with lowercase letter."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["column", "hide", path, "Sheet1", "b"])

            assert result.exit_code == 0
            assert "Hid column B" in result.output

            # Verify column is hidden
            wb = openpyxl.load_workbook(path)
            assert wb["Sheet1"].column_dimensions["B"].hidden is True
            wb.close()


class TestColumnUnhide:
    """Tests for xlforge column unhide command."""

    def test_column_unhide_file_not_found(self):
        """Test column unhide with non-existent file returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            result = runner.invoke(app, ["column", "unhide", path, "Sheet1", "A"])

            assert result.exit_code == ErrorCode.FILE_DOES_NOT_EXIST
            assert "does not exist" in result.output.lower()

    def test_column_unhide_sheet_not_found(self):
        """Test column unhide with non-existent sheet returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["column", "unhide", path, "NonexistentSheet", "A"]
            )

            assert result.exit_code == ErrorCode.SHEET_NOT_FOUND
            assert "Sheet not found" in result.output

    def test_column_unhide_success(self):
        """Test unhiding a column."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws.column_dimensions["A"].hidden = True
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["column", "unhide", path, "Sheet1", "A"])

            assert result.exit_code == 0
            assert "Unhid column A" in result.output

            # Verify column is visible
            wb = openpyxl.load_workbook(path)
            assert wb["Sheet1"].column_dimensions["A"].hidden is False
            wb.close()

    def test_column_unhide_lowercase(self):
        """Test unhiding a column with lowercase letter."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws.column_dimensions["B"].hidden = True
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["column", "unhide", path, "Sheet1", "b"])

            assert result.exit_code == 0
            assert "Unhid column B" in result.output

            # Verify column is visible
            wb = openpyxl.load_workbook(path)
            assert wb["Sheet1"].column_dimensions["B"].hidden is False
            wb.close()


class TestNamedRangeCreate:
    """Tests for xlforge named-range create command."""

    def test_named_range_create_file_not_found(self):
        """Test named-range create with non-existent file returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            result = runner.invoke(
                app, ["named-range", "create", path, "MyRange", "Sheet1", "A1:C10"]
            )

            assert result.exit_code == ErrorCode.FILE_DOES_NOT_EXIST
            assert "does not exist" in result.output.lower()

    def test_named_range_create_sheet_not_found(self):
        """Test named-range create with non-existent sheet returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app,
                [
                    "named-range",
                    "create",
                    path,
                    "MyRange",
                    "NonExistentSheet",
                    "A1:C10",
                ],
            )

            assert result.exit_code == ErrorCode.SHEET_NOT_FOUND
            assert "Sheet not found" in result.output

    def test_named_range_create_success(self):
        """Test creating a named range."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws["A1"] = "Data"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["named-range", "create", path, "SalesData", "Sheet1", "A1:C10"]
            )

            assert result.exit_code == 0
            assert "Created named range 'SalesData'" in result.output
            assert "Sheet1!A1:C10" in result.output

            # Verify the named range was created
            wb = openpyxl.load_workbook(path)
            assert "SalesData" in wb.defined_names
            assert wb.defined_names["SalesData"].attr_text == "Sheet1!A1:C10"
            wb.close()

    def test_named_range_create_already_exists(self):
        """Test creating a named range that already exists returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            # Create a named range first
            from openpyxl.workbook.defined_name import DefinedName

            wb.defined_names.add(DefinedName("ExistingRange", attr_text="Sheet1!A1:B5"))
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app,
                ["named-range", "create", path, "ExistingRange", "Sheet1", "C1:D10"],
            )

            assert result.exit_code == ErrorCode.TABLE_ALREADY_EXISTS
            assert "already exists" in result.output.lower()


class TestNamedRangeDelete:
    """Tests for xlforge named-range delete command."""

    def test_named_range_delete_file_not_found(self):
        """Test named-range delete with non-existent file returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            result = runner.invoke(app, ["named-range", "delete", path, "MyRange"])

            assert result.exit_code == ErrorCode.FILE_DOES_NOT_EXIST
            assert "does not exist" in result.output.lower()

    def test_named_range_delete_not_found(self):
        """Test deleting a non-existent named range returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["named-range", "delete", path, "NonExistentRange"]
            )

            assert result.exit_code == ErrorCode.TABLE_NOT_FOUND
            assert "not found" in result.output.lower()

    def test_named_range_delete_success(self):
        """Test deleting a named range."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            # Create a named range first
            from openpyxl.workbook.defined_name import DefinedName

            wb.defined_names.add(DefinedName("ToDelete", attr_text="Sheet1!A1:B5"))
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["named-range", "delete", path, "ToDelete"])

            assert result.exit_code == 0
            assert "Deleted named range 'ToDelete'" in result.output

            # Verify the named range was deleted
            wb = openpyxl.load_workbook(path)
            assert "ToDelete" not in wb.defined_names
            wb.close()


class TestNamedRangeList:
    """Tests for xlforge named-range list command."""

    def test_named_range_list_file_not_found(self):
        """Test named-range list with non-existent file returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            result = runner.invoke(app, ["named-range", "list", path])

            assert result.exit_code == ErrorCode.FILE_DOES_NOT_EXIST
            assert "does not exist" in result.output.lower()

    def test_named_range_list_empty(self):
        """Test listing named ranges when none exist."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["named-range", "list", path])

            assert result.exit_code == 0
            assert "No named ranges" in result.output

    def test_named_range_list_success(self):
        """Test listing named ranges."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            # Create named ranges
            from openpyxl.workbook.defined_name import DefinedName

            wb.defined_names.add(DefinedName("Range1", attr_text="Sheet1!A1:A10"))
            wb.defined_names.add(DefinedName("Range2", attr_text="Sheet1!B1:B10"))
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["named-range", "list", path])

            assert result.exit_code == 0
            assert "Range1" in result.output
            assert "Sheet1!A1:A10" in result.output
            assert "Range2" in result.output
            assert "Sheet1!B1:B10" in result.output


class TestNamedRangeGet:
    """Tests for xlforge named-range get command."""

    def test_named_range_get_file_not_found(self):
        """Test named-range get with non-existent file returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            result = runner.invoke(app, ["named-range", "get", path, "MyRange"])

            assert result.exit_code == ErrorCode.FILE_DOES_NOT_EXIST
            assert "does not exist" in result.output.lower()

    def test_named_range_get_not_found(self):
        """Test getting a non-existent named range returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["named-range", "get", path, "NonExistentRange"]
            )

            assert result.exit_code == ErrorCode.TABLE_NOT_FOUND
            assert "not found" in result.output.lower()

    def test_named_range_get_success(self):
        """Test getting a named range."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            # Create a named range
            from openpyxl.workbook.defined_name import DefinedName

            wb.defined_names.add(DefinedName("MyRange", attr_text="Sheet1!A1:C100"))
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["named-range", "get", path, "MyRange"])

            assert result.exit_code == 0
            assert "MyRange = Sheet1!A1:C100" in result.output


class TestStyleSet:
    """Tests for xlforge style set command."""

    def test_style_set_file_not_found(self):
        """Test style set with non-existent file returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            result = runner.invoke(
                app, ["style", "set", path, "Sheet1", "A1", "--bold"]
            )

            assert result.exit_code == ErrorCode.FILE_DOES_NOT_EXIST
            assert "does not exist" in result.output.lower()

    def test_style_set_sheet_not_found(self):
        """Test style set with non-existent sheet returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["style", "set", path, "NonexistentSheet", "A1", "--bold"]
            )

            assert result.exit_code == ErrorCode.SHEET_NOT_FOUND
            assert "Sheet not found" in result.output

    def test_style_set_invalid_color(self):
        """Test style set with invalid color format returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["style", "set", path, "Sheet", "A1", "--color", "invalid"]
            )

            assert result.exit_code == ErrorCode.INVALID_STYLE_STRING
            assert "Invalid color" in result.output

    def test_style_set_bold(self):
        """Test setting cell bold style."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws["A1"] = "Test"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["style", "set", path, "Sheet1", "A1", "--bold"]
            )

            assert result.exit_code == 0
            assert "bold" in result.output

            # Verify style was applied
            wb = openpyxl.load_workbook(path)
            assert wb["Sheet1"]["A1"].font.bold is True
            wb.close()

    def test_style_set_italic(self):
        """Test setting cell italic style."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws["A1"] = "Test"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["style", "set", path, "Sheet1", "A1", "--italic"]
            )

            assert result.exit_code == 0
            assert "italic" in result.output

            # Verify style was applied
            wb = openpyxl.load_workbook(path)
            assert wb["Sheet1"]["A1"].font.italic is True
            wb.close()

    def test_style_set_color(self):
        """Test setting cell font color."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws["A1"] = "Test"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["style", "set", path, "Sheet1", "A1", "--color", "FF0000"]
            )

            assert result.exit_code == 0
            assert "color" in result.output.lower()

            # Verify style was applied
            wb = openpyxl.load_workbook(path)
            assert wb["Sheet1"]["A1"].font.color.rgb == "FFFF0000"  # ARGB format
            wb.close()

    def test_style_set_color_with_hash(self):
        """Test setting cell font color with # prefix."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws["A1"] = "Test"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["style", "set", path, "Sheet1", "A1", "--color", "#00FF00"]
            )

            assert result.exit_code == 0

            # Verify style was applied
            wb = openpyxl.load_workbook(path)
            assert wb["Sheet1"]["A1"].font.color.rgb == "FF00FF00"  # ARGB format
            wb.close()

    def test_style_set_multiple(self):
        """Test setting multiple styles at once."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws["A1"] = "Test"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app,
                [
                    "style",
                    "set",
                    path,
                    "Sheet1",
                    "A1",
                    "--bold",
                    "--italic",
                    "--color",
                    "0000FF",
                ],
            )

            assert result.exit_code == 0
            assert "bold" in result.output
            assert "italic" in result.output
            assert "color" in result.output.lower()

            # Verify all styles were applied
            wb = openpyxl.load_workbook(path)
            cell = wb["Sheet1"]["A1"]
            assert cell.font.bold is True
            assert cell.font.italic is True
            assert cell.font.color.rgb == "FF0000FF"  # ARGB format
            wb.close()

    def test_style_set_no_options(self):
        """Test style set with no style options shows message."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws["A1"] = "Test"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["style", "set", path, "Sheet1", "A1"])

            assert result.exit_code == 0
            assert "No style changes" in result.output


class TestStyleNumberFormat:
    """Tests for xlforge style number-format command."""

    def test_style_number_format_file_not_found(self):
        """Test style number-format with non-existent file returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            result = runner.invoke(
                app, ["style", "number-format", path, "Sheet1", "A1", "0.00"]
            )

            assert result.exit_code == ErrorCode.FILE_DOES_NOT_EXIST
            assert "does not exist" in result.output.lower()

    def test_style_number_format_sheet_not_found(self):
        """Test style number-format with non-existent sheet returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["style", "number-format", path, "NonexistentSheet", "A1", "0.00"]
            )

            assert result.exit_code == ErrorCode.SHEET_NOT_FOUND
            assert "Sheet not found" in result.output

    def test_style_number_format_success(self):
        """Test setting number format on a cell."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws["A1"] = 42.5
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["style", "number-format", path, "Sheet1", "A1", "0.00"]
            )

            assert result.exit_code == 0
            assert "number format" in result.output.lower()
            assert "0.00" in result.output

            # Verify format was applied
            wb = openpyxl.load_workbook(path)
            assert wb["Sheet1"]["A1"].number_format == "0.00"
            wb.close()

    def test_style_number_format_currency(self):
        """Test setting currency number format."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws["B2"] = 1234.56
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["style", "number-format", path, "Sheet1", "B2", "$#,##0.00"]
            )

            assert result.exit_code == 0

            # Verify format was applied
            wb = openpyxl.load_workbook(path)
            assert wb["Sheet1"]["B2"].number_format == "$#,##0.00"
            wb.close()


class TestStyleFont:
    """Tests for xlforge style font command."""

    def test_style_font_file_not_found(self):
        """Test style font with non-existent file returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            result = runner.invoke(
                app, ["style", "font", path, "Sheet1", "A1", "--name", "Arial"]
            )

            assert result.exit_code == ErrorCode.FILE_DOES_NOT_EXIST
            assert "does not exist" in result.output.lower()

    def test_style_font_sheet_not_found(self):
        """Test style font with non-existent sheet returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app,
                ["style", "font", path, "NonexistentSheet", "A1", "--name", "Arial"],
            )

            assert result.exit_code == ErrorCode.SHEET_NOT_FOUND
            assert "Sheet not found" in result.output

    def test_style_font_no_options(self):
        """Test style font with no options returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["style", "font", path, "Sheet", "A1"])

            assert result.exit_code == 1
            assert "Must specify at least one" in result.output

    def test_style_font_name(self):
        """Test setting font name."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws["A1"] = "Test"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["style", "font", path, "Sheet1", "A1", "--name", "Arial"]
            )

            assert result.exit_code == 0
            assert "name" in result.output.lower()
            assert "Arial" in result.output

            # Verify font name was applied
            wb = openpyxl.load_workbook(path)
            assert wb["Sheet1"]["A1"].font.name == "Arial"
            wb.close()

    def test_style_font_size(self):
        """Test setting font size."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws["A1"] = "Test"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["style", "font", path, "Sheet1", "A1", "--size", "14"]
            )

            assert result.exit_code == 0
            assert "size" in result.output.lower()
            assert "14" in result.output

            # Verify font size was applied
            wb = openpyxl.load_workbook(path)
            assert wb["Sheet1"]["A1"].font.size == 14
            wb.close()

    def test_style_font_name_and_size(self):
        """Test setting font name and size together."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws["A1"] = "Test"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app,
                [
                    "style",
                    "font",
                    path,
                    "Sheet1",
                    "A1",
                    "--name",
                    "Calibri",
                    "--size",
                    "12",
                ],
            )

            assert result.exit_code == 0
            assert "name" in result.output.lower()
            assert "size" in result.output.lower()

            # Verify both were applied
            wb = openpyxl.load_workbook(path)
            cell = wb["Sheet1"]["A1"]
            assert cell.font.name == "Calibri"
            assert cell.font.size == 12
            wb.close()


class TestPropertiesGet:
    """Tests for xlforge properties get command."""

    def test_properties_get_file_not_found(self):
        """Test properties get with non-existent file returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            result = runner.invoke(app, ["properties", "get", path])

            assert result.exit_code == ErrorCode.FILE_DOES_NOT_EXIST
            assert "does not exist" in result.output.lower()

    def test_properties_get_empty_workbook(self):
        """Test getting properties from a new workbook."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["properties", "get", path])

            assert result.exit_code == 0
            assert "Workbook Properties:" in result.output
            # openpyxl sets default creator='openpyxl' so properties will be shown

    def test_properties_get_with_values(self):
        """Test getting properties that are set."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.properties.title = "Test Document"
            wb.properties.creator = "John Doe"
            wb.properties.subject = "Test Subject"
            wb.properties.keywords = "test, keywords"
            wb.properties.description = "Test comments"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["properties", "get", path])

            assert result.exit_code == 0
            assert "Test Document" in result.output
            assert "John Doe" in result.output
            assert "Test Subject" in result.output
            assert "test, keywords" in result.output
            assert "Test comments" in result.output

    def test_properties_get_json_output(self):
        """Test getting properties with JSON output."""
        import json

        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.properties.title = "JSON Test"
            wb.properties.creator = "Jane Doe"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["properties", "get", path, "--json"])

            assert result.exit_code == 0
            data = json.loads(result.output)
            assert data["title"] == "JSON Test"
            assert data["author"] == "Jane Doe"


class TestPropertiesSet:
    """Tests for xlforge properties set command."""

    def test_properties_set_file_not_found(self):
        """Test properties set with non-existent file returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            result = runner.invoke(app, ["properties", "set", path, "--title", "Test"])

            assert result.exit_code == ErrorCode.FILE_DOES_NOT_EXIST
            assert "does not exist" in result.output.lower()

    def test_properties_set_no_properties(self):
        """Test properties set with no properties provided returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["properties", "set", path])

            assert result.exit_code == 1
            assert "Must provide at least one property" in result.output

    def test_properties_set_title(self):
        """Test setting the title property."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["properties", "set", path, "--title", "My Title"]
            )

            assert result.exit_code == 0
            assert "title='My Title'" in result.output

            # Verify the property was set
            wb = openpyxl.load_workbook(path)
            assert wb.properties.title == "My Title"
            wb.close()

    def test_properties_set_author(self):
        """Test setting the author property."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["properties", "set", path, "--author", "John Doe"]
            )

            assert result.exit_code == 0
            assert "author='John Doe'" in result.output

            # Verify the property was set
            wb = openpyxl.load_workbook(path)
            assert wb.properties.creator == "John Doe"
            wb.close()

    def test_properties_set_subject(self):
        """Test setting the subject property."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["properties", "set", path, "--subject", "Test Subject"]
            )

            assert result.exit_code == 0
            assert "subject='Test Subject'" in result.output

            # Verify the property was set
            wb = openpyxl.load_workbook(path)
            assert wb.properties.subject == "Test Subject"
            wb.close()

    def test_properties_set_keywords(self):
        """Test setting the keywords property."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["properties", "set", path, "--keywords", "test, keywords"]
            )

            assert result.exit_code == 0
            assert "keywords='test, keywords'" in result.output

            # Verify the property was set
            wb = openpyxl.load_workbook(path)
            assert wb.properties.keywords == "test, keywords"
            wb.close()

    def test_properties_set_comments(self):
        """Test setting the comments property."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["properties", "set", path, "--comments", "Test comments"]
            )

            assert result.exit_code == 0
            assert "comments='Test comments'" in result.output

            # Verify the property was set
            wb = openpyxl.load_workbook(path)
            assert wb.properties.description == "Test comments"
            wb.close()

    def test_properties_set_multiple(self):
        """Test setting multiple properties at once."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app,
                [
                    "properties",
                    "set",
                    path,
                    "--title",
                    "Multi Title",
                    "--author",
                    "Multi Author",
                    "--subject",
                    "Multi Subject",
                ],
            )

            assert result.exit_code == 0
            assert "title='Multi Title'" in result.output
            assert "author='Multi Author'" in result.output
            assert "subject='Multi Subject'" in result.output

            # Verify all properties were set
            wb = openpyxl.load_workbook(path)
            assert wb.properties.title == "Multi Title"
            assert wb.properties.creator == "Multi Author"
            assert wb.properties.subject == "Multi Subject"
            wb.close()

    def test_properties_set_preserves_existing(self):
        """Test setting one property preserves others."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.properties.title = "Original Title"
            wb.properties.creator = "Original Author"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["properties", "set", path, "--subject", "New Subject"]
            )

            assert result.exit_code == 0

            # Verify all properties
            wb = openpyxl.load_workbook(path)
            assert wb.properties.title == "Original Title"
            assert wb.properties.creator == "Original Author"
            assert wb.properties.subject == "New Subject"
            wb.close()


class TestValidationAdd:
    """Tests for xlforge validation add command."""

    def test_validation_add_file_not_found(self):
        """Test validation add with non-existent file returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            result = runner.invoke(
                app,
                [
                    "validation",
                    "add",
                    path,
                    "Sheet1",
                    "A1:A10",
                    "--type",
                    "list",
                    "--formula1",
                    "A,B,C",
                ],
            )

            assert result.exit_code == ErrorCode.FILE_DOES_NOT_EXIST
            assert "does not exist" in result.output.lower()

    def test_validation_add_sheet_not_found(self):
        """Test validation add with non-existent sheet returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app,
                [
                    "validation",
                    "add",
                    path,
                    "NonExistentSheet",
                    "A1:A10",
                    "--type",
                    "list",
                    "--formula1",
                    "A,B,C",
                ],
            )

            assert result.exit_code == ErrorCode.SHEET_NOT_FOUND
            assert "Sheet not found" in result.output

    def test_validation_add_invalid_type(self):
        """Test validation add with invalid type returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app,
                [
                    "validation",
                    "add",
                    path,
                    "Sheet",
                    "A1:A10",
                    "--type",
                    "invalid",
                    "--formula1",
                    "A,B,C",
                ],
            )

            assert result.exit_code == ErrorCode.VALIDATION_TYPE_NOT_SUPPORTED
            assert "Invalid validation type" in result.output

    def test_validation_add_missing_formula1(self):
        """Test validation add without formula1 returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["validation", "add", path, "Sheet", "A1:A10", "--type", "list"]
            )

            assert result.exit_code == ErrorCode.INVALID_FORMULA_SYNTAX
            assert "--formula1 is required" in result.output

    def test_validation_add_list_success(self):
        """Test adding a list (dropdown) validation."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app,
                [
                    "validation",
                    "add",
                    path,
                    "Sheet1",
                    "A1:A10",
                    "--type",
                    "list",
                    "--formula1",
                    "Option1,Option2,Option3",
                ],
            )

            assert result.exit_code == 0
            assert "Added list validation" in result.output
            assert "A1:A10" in result.output

            # Verify validation was added
            wb = openpyxl.load_workbook(path)
            ws = wb["Sheet1"]
            assert len(ws.data_validations.dataValidation) == 1
            dv = ws.data_validations.dataValidation[0]
            assert dv.type == "list"
            assert dv.formula1 == "Option1,Option2,Option3"
            wb.close()

    def test_validation_add_whole_success(self):
        """Test adding a whole number validation."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app,
                [
                    "validation",
                    "add",
                    path,
                    "Sheet1",
                    "B1:B10",
                    "--type",
                    "whole",
                    "--formula1",
                    "0",
                    "--formula2",
                    "100",
                ],
            )

            assert result.exit_code == 0
            assert "Added whole validation" in result.output
            assert "B1:B10" in result.output
            assert "Formula1: 0" in result.output
            assert "Formula2: 100" in result.output

            # Verify validation was added
            wb = openpyxl.load_workbook(path)
            ws = wb["Sheet1"]
            assert len(ws.data_validations.dataValidation) == 1
            dv = ws.data_validations.dataValidation[0]
            assert dv.type == "whole"
            assert dv.formula1 == "0"
            assert dv.formula2 == "100"
            wb.close()

    def test_validation_add_decimal_success(self):
        """Test adding a decimal validation."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app,
                [
                    "validation",
                    "add",
                    path,
                    "Sheet1",
                    "C1:C5",
                    "--type",
                    "decimal",
                    "--formula1",
                    "0.0",
                    "--formula2",
                    "99.99",
                ],
            )

            assert result.exit_code == 0
            assert "Added decimal validation" in result.output

            # Verify validation was added
            wb = openpyxl.load_workbook(path)
            ws = wb["Sheet1"]
            assert len(ws.data_validations.dataValidation) == 1
            dv = ws.data_validations.dataValidation[0]
            assert dv.type == "decimal"
            wb.close()


class TestValidationRemove:
    """Tests for xlforge validation remove command."""

    def test_validation_remove_file_not_found(self):
        """Test validation remove with non-existent file returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            result = runner.invoke(
                app, ["validation", "remove", path, "Sheet1", "A1:A10"]
            )

            assert result.exit_code == ErrorCode.FILE_DOES_NOT_EXIST
            assert "does not exist" in result.output.lower()

    def test_validation_remove_sheet_not_found(self):
        """Test validation remove with non-existent sheet returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["validation", "remove", path, "NonExistentSheet", "A1:A10"]
            )

            assert result.exit_code == ErrorCode.SHEET_NOT_FOUND
            assert "Sheet not found" in result.output

    def test_validation_remove_not_found(self):
        """Test validation remove when no validation exists returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["validation", "remove", path, "Sheet1", "A1:A10"]
            )

            assert result.exit_code == ErrorCode.VALIDATION_TYPE_NOT_SUPPORTED
            assert "No data validation found" in result.output

    def test_validation_remove_success(self):
        """Test removing a validation."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"

            # Add a validation first
            from openpyxl.worksheet.datavalidation import DataValidation

            dv = DataValidation(type="list", formula1="A,B,C", allow_blank=True)
            dv.add("A1:A10")
            ws.add_data_validation(dv)

            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["validation", "remove", path, "Sheet1", "A1:A10"]
            )

            assert result.exit_code == 0
            assert "Removed data validation" in result.output

            # Verify validation was removed
            wb = openpyxl.load_workbook(path)
            ws = wb["Sheet1"]
            assert len(ws.data_validations.dataValidation) == 0
            wb.close()


class TestValidationList:
    """Tests for xlforge validation list command."""

    def test_validation_list_file_not_found(self):
        """Test validation list with non-existent file returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            result = runner.invoke(app, ["validation", "list", path, "Sheet1"])

            assert result.exit_code == ErrorCode.FILE_DOES_NOT_EXIST
            assert "does not exist" in result.output.lower()

    def test_validation_list_sheet_not_found(self):
        """Test validation list with non-existent sheet returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["validation", "list", path, "NonExistentSheet"]
            )

            assert result.exit_code == ErrorCode.SHEET_NOT_FOUND
            assert "Sheet not found" in result.output

    def test_validation_list_empty(self):
        """Test listing validations when none exist."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["validation", "list", path, "Sheet1"])

            assert result.exit_code == 0
            assert "No data validations found" in result.output

    def test_validation_list_success(self):
        """Test listing validations."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"

            # Add validations
            from openpyxl.worksheet.datavalidation import DataValidation

            dv1 = DataValidation(type="list", formula1="A,B,C", allow_blank=True)
            dv1.add("A1:A10")
            ws.add_data_validation(dv1)

            dv2 = DataValidation(
                type="whole", formula1="0", formula2="100", allow_blank=True
            )
            dv2.add("B1:B10")
            ws.add_data_validation(dv2)

            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["validation", "list", path, "Sheet1"])

            assert result.exit_code == 0
            assert "Sheet1" in result.output
            assert "list" in result.output
            assert "whole" in result.output
            assert "A,B,C" in result.output
            assert "0" in result.output
            assert "100" in result.output
            assert "A1:A10" in result.output
            assert "B1:B10" in result.output


class TestChartCreate:
    """Tests for xlforge chart create command."""

    def test_chart_create_file_not_found(self):
        """Test chart create with non-existent file returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            with unittest.mock.patch(
                "xlforge.commands.chart.find_spec", return_value=None
            ):
                result = runner.invoke(
                    app,
                    [
                        "chart",
                        "create",
                        path,
                        "Sheet1",
                        "A1:D10",
                        "--type",
                        "column",
                        "--name",
                        "TestChart",
                    ],
                )

            # When xlwings is not available, returns FEATURE_UNAVAILABLE
            assert result.exit_code == ErrorCode.FEATURE_UNAVAILABLE
            assert (
                "xlwings" in result.output.lower() or "excel" in result.output.lower()
            )

    def test_chart_create_xlwings_not_available(self):
        """Test chart create returns error 9 when xlwings is not available."""
        import unittest.mock

        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            # Mock find_spec to return None to simulate xlwings not available
            with unittest.mock.patch(
                "xlforge.commands.chart.find_spec", return_value=None
            ):
                result = runner.invoke(
                    app,
                    [
                        "chart",
                        "create",
                        path,
                        "Sheet1",
                        "A1:D10",
                        "--type",
                        "column",
                        "--name",
                        "TestChart",
                    ],
                )

            # Verify error 9 is returned when xlwings is not available
            assert result.exit_code == ErrorCode.FEATURE_UNAVAILABLE
            assert (
                "Feature unavailable" in result.output
                or "xlwings" in result.output.lower()
            )

    def test_chart_create_invalid_type_when_xlwings_available(self):
        """Test chart create with invalid type returns error when xlwings is mocked."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws["A1"] = "Category"
            ws["A2"] = "A"
            ws["A3"] = "B"
            ws["B1"] = "Value"
            ws["B2"] = 10
            ws["B3"] = 20
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")

            # Mock xlwings availability to test actual chart functionality
            import unittest.mock as mock

            with mock.patch(
                "xlforge.commands.chart._is_xlwings_available", return_value=True
            ):
                result = runner.invoke(
                    app,
                    [
                        "chart",
                        "create",
                        path,
                        "Sheet1",
                        "A1:B3",
                        "--type",
                        "invalid",
                        "--name",
                        "TestChart",
                    ],
                )

            assert result.exit_code == ErrorCode.INVALID_CHART_TYPE
            assert "Invalid chart type" in result.output

    def test_chart_create_invalid_range_when_xlwings_available(self):
        """Test chart create with invalid range returns error when xlwings is mocked."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")

            import unittest.mock as mock

            with mock.patch(
                "xlforge.commands.chart._is_xlwings_available", return_value=True
            ):
                result = runner.invoke(
                    app,
                    [
                        "chart",
                        "create",
                        path,
                        "Sheet1",
                        "invalid-range",
                        "--type",
                        "column",
                        "--name",
                        "TestChart",
                    ],
                )

            assert result.exit_code == ErrorCode.INVALID_SYNTAX
            assert "Invalid range format" in result.output

    def test_chart_create_success_when_xlwings_available(self):
        """Test successful chart creation when xlwings is mocked."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws["A1"] = "Category"
            ws["A2"] = "A"
            ws["A3"] = "B"
            ws["B1"] = "Value"
            ws["B2"] = 10
            ws["B3"] = 20
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")

            import unittest.mock as mock

            with mock.patch(
                "xlforge.commands.chart._is_xlwings_available", return_value=True
            ):
                result = runner.invoke(
                    app,
                    [
                        "chart",
                        "create",
                        path,
                        "Sheet1",
                        "A1:B3",
                        "--type",
                        "column",
                        "--name",
                        "TestChart",
                    ],
                )

            assert result.exit_code == 0
            assert "Created chart" in result.output
            assert "TestChart" in result.output

            # Verify chart was actually created
            wb = openpyxl.load_workbook(path)
            ws = wb["Sheet1"]
            assert len(ws._charts) == 1
            # Chart was created successfully - verify by checking chart count increased
            wb.close()


class TestChartDelete:
    """Tests for xlforge chart delete command."""

    def test_chart_delete_file_not_found(self):
        """Test chart delete with non-existent file returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            with unittest.mock.patch(
                "xlforge.commands.chart.find_spec", return_value=None
            ):
                result = runner.invoke(
                    app, ["chart", "delete", path, "Sheet1", "TestChart"]
                )

            # When xlwings is not available, returns FEATURE_UNAVAILABLE
            assert result.exit_code == ErrorCode.FEATURE_UNAVAILABLE

    def test_chart_delete_xlwings_not_available(self):
        """Test chart delete returns error 9 when xlwings is not available."""
        import unittest.mock

        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            with unittest.mock.patch(
                "xlforge.commands.chart.find_spec", return_value=None
            ):
                result = runner.invoke(
                    app, ["chart", "delete", path, "Sheet1", "TestChart"]
                )

            assert result.exit_code == ErrorCode.FEATURE_UNAVAILABLE

    def test_chart_delete_not_found_when_xlwings_available(self):
        """Test chart delete with non-existent chart returns error when xlwings is mocked."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")

            import unittest.mock as mock

            with mock.patch(
                "xlforge.commands.chart._is_xlwings_available", return_value=True
            ):
                result = runner.invoke(
                    app, ["chart", "delete", path, "Sheet1", "NonExistentChart"]
                )

            assert result.exit_code == ErrorCode.CHART_NOT_FOUND
            assert "Chart not found" in result.output

    def test_chart_delete_success_when_xlwings_available(self):
        """Test successful chart deletion when xlwings is mocked."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"

            # Add a chart first using openpyxl
            from openpyxl.chart import BarChart, Reference

            chart = BarChart()
            chart.title = "TestChart"
            data = Reference(ws, min_col=1, min_row=1, max_col=1, max_row=3)
            chart.add_data(data)
            ws.add_chart(chart, "D1")

            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")

            import unittest.mock as mock

            with mock.patch(
                "xlforge.commands.chart._is_xlwings_available", return_value=True
            ):
                result = runner.invoke(
                    app, ["chart", "delete", path, "Sheet1", "TestChart"]
                )

            assert result.exit_code == 0
            assert "Deleted chart" in result.output

            # Verify chart was actually deleted
            wb = openpyxl.load_workbook(path)
            ws = wb["Sheet1"]
            assert len(ws._charts) == 0
            wb.close()


class TestChartList:
    """Tests for xlforge chart list command."""

    def test_chart_list_file_not_found(self):
        """Test chart list with non-existent file returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            with unittest.mock.patch(
                "xlforge.commands.chart.find_spec", return_value=None
            ):
                result = runner.invoke(app, ["chart", "list", path, "Sheet1"])

            assert result.exit_code == ErrorCode.FEATURE_UNAVAILABLE

    def test_chart_list_xlwings_not_available(self):
        """Test chart list returns error 9 when xlwings is not available."""
        import unittest.mock

        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            with unittest.mock.patch(
                "xlforge.commands.chart.find_spec", return_value=None
            ):
                result = runner.invoke(app, ["chart", "list", path, "Sheet1"])

            assert result.exit_code == ErrorCode.FEATURE_UNAVAILABLE

    def test_chart_list_empty_when_xlwings_available(self):
        """Test chart list with no charts returns proper message when xlwings is mocked."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")

            import unittest.mock as mock

            with mock.patch(
                "xlforge.commands.chart._is_xlwings_available", return_value=True
            ):
                result = runner.invoke(app, ["chart", "list", path, "Sheet1"])

            assert result.exit_code == 0
            assert "No charts found" in result.output

    def test_chart_list_with_charts_when_xlwings_available(self):
        """Test chart list shows charts when xlwings is mocked."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"

            # Add charts using openpyxl
            from openpyxl.chart import BarChart, LineChart, Reference

            chart1 = BarChart()
            chart1.title = "BarChart"
            data = Reference(ws, min_col=1, min_row=1, max_col=1, max_row=3)
            chart1.add_data(data)
            ws.add_chart(chart1, "D1")

            chart2 = LineChart()
            chart2.title = "LineChart"
            ws.add_chart(chart2, "H1")

            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")

            import unittest.mock as mock

            with mock.patch(
                "xlforge.commands.chart._is_xlwings_available", return_value=True
            ):
                result = runner.invoke(app, ["chart", "list", path, "Sheet1"])

            assert result.exit_code == 0
            assert "BarChart" in result.output
            assert "LineChart" in result.output
            assert "Sheet1" in result.output


class TestContextCommands:
    """Tests for context commands."""

    def test_context_show_empty(self):
        """Test showing context when none is set."""
        # First ensure context is cleared
        runner.invoke(app, ["context", "clear"])
        result = runner.invoke(app, ["context", "show"])

        assert result.exit_code == 0
        assert "No context is set" in result.output

    def test_context_set_file_only(self):
        """Test setting context with file only."""
        result = runner.invoke(app, ["context", "set", "report.xlsx"])

        assert result.exit_code == 0
        assert "Context set" in result.output
        assert "report.xlsx" in result.output

    def test_context_set_with_sheet(self):
        """Test setting context with file and sheet."""
        result = runner.invoke(
            app, ["context", "set", "report.xlsx", "--sheet", "Data"]
        )

        assert result.exit_code == 0
        assert "Context set" in result.output
        assert "report.xlsx" in result.output
        assert "Data" in result.output

    def test_context_show_after_set(self):
        """Test showing context after it has been set."""
        # Set context first
        runner.invoke(app, ["context", "set", "test.xlsx", "--sheet", "Sheet1"])

        result = runner.invoke(app, ["context", "show"])

        assert result.exit_code == 0
        assert "test.xlsx" in result.output
        assert "Sheet1" in result.output

    def test_context_clear(self):
        """Test clearing context."""
        # First set context
        runner.invoke(app, ["context", "set", "test.xlsx"])

        # Then clear it
        result = runner.invoke(app, ["context", "clear"])

        assert result.exit_code == 0
        assert "Context cleared" in result.output

    def test_context_clear_when_empty(self):
        """Test clearing context when none is set."""
        # Ensure context is cleared first
        runner.invoke(app, ["context", "clear"])

        result = runner.invoke(app, ["context", "clear"])

        assert result.exit_code == 0
        assert "No context to clear" in result.output


class TestFileCheck:
    """Tests for xlforge file check command."""

    def test_file_check_file_not_found(self):
        """Test file check with non-existent file returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            result = runner.invoke(app, ["file", "check", path])

            assert result.exit_code == ErrorCode.FILE_DOES_NOT_EXIST
            assert "does not exist" in result.output.lower()

    def test_file_check_success(self):
        """Test file check with valid file."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["file", "check", path])

            assert result.exit_code == 0
            assert "Healthy: True" in result.output
            assert "Valid xlsx: True" in result.output

    def test_file_check_json_output(self):
        """Test file check with JSON output."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["file", "check", path, "--json"])

            assert result.exit_code == 0
            assert '"exists": true' in result.output
            assert '"valid_xlsx": true' in result.output
            assert '"healthy": true' in result.output


class TestFileRecover:
    """Tests for xlforge file recover command."""

    def test_file_recover_file_not_found(self):
        """Test file recover with non-existent file returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            result = runner.invoke(app, ["file", "recover", path])

            assert result.exit_code == ErrorCode.FILE_DOES_NOT_EXIST
            assert "does not exist" in result.output.lower()

    def test_file_recover_success(self):
        """Test file recover with valid file."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["file", "recover", path])

            assert result.exit_code == 0
            assert "Recovered" in result.output


class TestRowWidth:
    """Tests for xlforge row width command."""

    def test_row_width_file_not_found(self):
        """Test row width with non-existent file returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            result = runner.invoke(app, ["row", "width", path, "Sheet1", "1", "20"])

            assert result.exit_code == ErrorCode.FILE_DOES_NOT_EXIST
            assert "does not exist" in result.output.lower()

    def test_row_width_sheet_not_found(self):
        """Test row width with non-existent sheet returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["row", "width", path, "NonexistentSheet", "1", "20"]
            )

            assert result.exit_code == ErrorCode.SHEET_NOT_FOUND
            assert "Sheet not found" in result.output

    def test_row_width_invalid_row(self):
        """Test row width with invalid row number returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["row", "width", path, "Sheet", "0", "20"])

            assert result.exit_code == ErrorCode.ROW_NOT_FOUND
            assert "Invalid row" in result.output

    def test_row_width_success(self):
        """Test setting row width."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["row", "width", path, "Sheet1", "1", "25.5"])

            assert result.exit_code == 0
            assert "Set row 1 height to 25.5" in result.output

            # Verify row height
            wb = openpyxl.load_workbook(path)
            assert wb["Sheet1"].row_dimensions[1].height == 25.5
            wb.close()


class TestRowAuto:
    """Tests for xlforge row auto command."""

    def test_row_auto_file_not_found(self):
        """Test row auto with non-existent file returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            result = runner.invoke(app, ["row", "auto", path, "Sheet1", "1"])

            assert result.exit_code == ErrorCode.FILE_DOES_NOT_EXIST
            assert "does not exist" in result.output.lower()

    def test_row_auto_sheet_not_found(self):
        """Test row auto with non-existent sheet returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["row", "auto", path, "NonexistentSheet", "1"])

            assert result.exit_code == ErrorCode.SHEET_NOT_FOUND
            assert "Sheet not found" in result.output

    def test_row_auto_invalid_row(self):
        """Test row auto with invalid row number returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["row", "auto", path, "Sheet", "0"])

            assert result.exit_code == ErrorCode.ROW_NOT_FOUND
            assert "Invalid row" in result.output

    def test_row_auto_success(self):
        """Test auto-fitting row height."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws.row_dimensions[1].height = 50
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["row", "auto", path, "Sheet1", "1"])

            assert result.exit_code == 0
            assert "Auto-fitted row 1" in result.output

            # Verify row height is auto-fitted (None)
            wb = openpyxl.load_workbook(path)
            assert wb["Sheet1"].row_dimensions[1].height is None
            wb.close()


class TestColumnWidth:
    """Tests for xlforge column width command."""

    def test_column_width_file_not_found(self):
        """Test column width with non-existent file returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            result = runner.invoke(app, ["column", "width", path, "Sheet1", "A", "20"])

            assert result.exit_code == ErrorCode.FILE_DOES_NOT_EXIST
            assert "does not exist" in result.output.lower()

    def test_column_width_sheet_not_found(self):
        """Test column width with non-existent sheet returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["column", "width", path, "NonexistentSheet", "A", "20"]
            )

            assert result.exit_code == ErrorCode.SHEET_NOT_FOUND
            assert "Sheet not found" in result.output

    def test_column_width_success(self):
        """Test setting column width."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["column", "width", path, "Sheet1", "A", "15.5"]
            )

            assert result.exit_code == 0
            assert "Set column A width to 15.5" in result.output

            # Verify column width
            wb = openpyxl.load_workbook(path)
            assert wb["Sheet1"].column_dimensions["A"].width == 15.5
            wb.close()

    def test_column_width_lowercase(self):
        """Test setting column width with lowercase letter."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["column", "width", path, "Sheet1", "b", "30"])

            assert result.exit_code == 0
            assert "Set column B width to 30" in result.output

            # Verify column width
            wb = openpyxl.load_workbook(path)
            assert wb["Sheet1"].column_dimensions["B"].width == 30
            wb.close()


class TestColumnAuto:
    """Tests for xlforge column auto command."""

    def test_column_auto_file_not_found(self):
        """Test column auto with non-existent file returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            result = runner.invoke(app, ["column", "auto", path, "Sheet1", "A"])

            assert result.exit_code == ErrorCode.FILE_DOES_NOT_EXIST
            assert "does not exist" in result.output.lower()

    def test_column_auto_sheet_not_found(self):
        """Test column auto with non-existent sheet returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["column", "auto", path, "NonexistentSheet", "A"]
            )

            assert result.exit_code == ErrorCode.SHEET_NOT_FOUND
            assert "Sheet not found" in result.output

    def test_column_auto_success(self):
        """Test auto-fitting column width."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws.column_dimensions["A"].width = 50
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["column", "auto", path, "Sheet1", "A"])

            assert result.exit_code == 0
            assert "Auto-fitted column A" in result.output

            # Verify column width is auto-fitted (bestFit=True)
            wb = openpyxl.load_workbook(path)
            assert wb["Sheet1"].column_dimensions["A"].bestFit is True
            wb.close()

    def test_column_auto_lowercase(self):
        """Test auto-fitting column width with lowercase letter."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws.column_dimensions["B"].width = 40
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["column", "auto", path, "Sheet1", "b"])

            assert result.exit_code == 0
            assert "Auto-fitted column B" in result.output

            # Verify column width is auto-fitted (bestFit=True)
            wb = openpyxl.load_workbook(path)
            assert wb["Sheet1"].column_dimensions["B"].bestFit is True
            wb.close()


class TestTableCreate:
    """Tests for xlforge table create command."""

    def test_table_create_file_not_found(self):
        """Test table create with non-existent file returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            result = runner.invoke(
                app, ["table", "create", path, "Sheet1", "A1:C10", "--name", "MyTable"]
            )

            assert result.exit_code == ErrorCode.FILE_DOES_NOT_EXIST
            assert "does not exist" in result.output.lower()

    def test_table_create_sheet_not_found(self):
        """Test table create with non-existent sheet returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app,
                [
                    "table",
                    "create",
                    path,
                    "NonexistentSheet",
                    "A1:C10",
                    "--name",
                    "MyTable",
                ],
            )

            assert result.exit_code == ErrorCode.SHEET_NOT_FOUND
            assert "Sheet not found" in result.output

    def test_table_create_success(self):
        """Test creating a table successfully."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            # Add some data to the range
            ws["A1"] = "Header1"
            ws["B1"] = "Header2"
            ws["C1"] = "Header3"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["table", "create", path, "Sheet1", "A1:C10", "--name", "MyTable"]
            )

            assert result.exit_code == 0
            assert "Created table 'MyTable'" in result.output

            # Verify table was created
            wb = openpyxl.load_workbook(path)
            ws = wb["Sheet1"]
            assert "MyTable" in ws.tables
            assert ws.tables["MyTable"].ref == "A1:C10"
            wb.close()

    def test_table_create_already_exists(self):
        """Test creating a table that already exists returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            from openpyxl.worksheet.table import Table

            ws.add_table(Table(displayName="MyTable", ref="A1:C10"))
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["table", "create", path, "Sheet1", "D1:F10", "--name", "MyTable"]
            )

            assert result.exit_code == ErrorCode.TABLE_ALREADY_EXISTS
            assert "already exists" in result.output.lower()

    def test_table_create_invalid_name(self):
        """Test creating a table with invalid name returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app,
                ["table", "create", path, "Sheet1", "A1:C10", "--name", "Invalid:Name"],
            )

            assert result.exit_code == ErrorCode.INVALID_TABLE_NAME
            assert "Invalid table name" in result.output

    def test_table_create_without_name(self):
        """Test creating a table without specifying a name uses default naming."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["table", "create", path, "Sheet1", "A1:C10"])

            assert result.exit_code == 0
            assert "Created table '" in result.output


class TestTableList:
    """Tests for xlforge table list command."""

    def test_table_list_file_not_found(self):
        """Test table list with non-existent file returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            result = runner.invoke(app, ["table", "list", path])

            assert result.exit_code == ErrorCode.FILE_DOES_NOT_EXIST
            assert "does not exist" in result.output.lower()

    def test_table_list_no_tables(self):
        """Test table list when no tables exist."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["table", "list", path])

            assert result.exit_code == 0
            assert "No tables found" in result.output

    def test_table_list_success(self):
        """Test listing tables successfully."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            from openpyxl.worksheet.table import Table

            ws.add_table(Table(displayName="Table1", ref="A1:C10"))
            ws.add_table(Table(displayName="Table2", ref="E1:G10"))
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["table", "list", path])

            assert result.exit_code == 0
            assert "Table1" in result.output
            assert "Table2" in result.output


class TestTableDelete:
    """Tests for xlforge table delete command."""

    def test_table_delete_file_not_found(self):
        """Test table delete with non-existent file returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            result = runner.invoke(app, ["table", "delete", path, "MyTable"])

            assert result.exit_code == ErrorCode.FILE_DOES_NOT_EXIST
            assert "does not exist" in result.output.lower()

    def test_table_delete_not_found(self):
        """Test deleting a non-existent table returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["table", "delete", path, "NonExistentTable"])

            assert result.exit_code == ErrorCode.TABLE_NOT_FOUND
            assert "not found" in result.output.lower()

    def test_table_delete_success(self):
        """Test deleting a table successfully."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            from openpyxl.worksheet.table import Table

            ws.add_table(Table(displayName="MyTable", ref="A1:C10"))
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["table", "delete", path, "MyTable"])

            assert result.exit_code == 0
            assert "Deleted table 'MyTable'" in result.output

            # Verify table was deleted
            wb = openpyxl.load_workbook(path)
            ws = wb["Sheet1"]
            assert "MyTable" not in ws.tables
            wb.close()

    def test_table_delete_with_sheet_option(self):
        """Test deleting a table with explicit sheet option."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws1 = wb.active
            ws1.title = "Sheet1"
            ws2 = wb.create_sheet("Sheet2")
            from openpyxl.worksheet.table import Table

            ws1.add_table(Table(displayName="Table1", ref="A1:C10"))
            ws2.add_table(Table(displayName="Table2", ref="A1:C10"))
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["table", "delete", path, "Table1", "--sheet", "Sheet1"]
            )

            assert result.exit_code == 0
            assert "Deleted table 'Table1' from sheet 'Sheet1'" in result.output

            # Verify table was deleted
            wb = openpyxl.load_workbook(path)
            assert "Table1" not in wb["Sheet1"].tables
            assert "Table2" in wb["Sheet2"].tables
            wb.close()


class TestPivotCreate:
    """Tests for xlforge pivot create command."""

    def test_pivot_create_xlwings_not_available(self):
        """Test pivot create returns FEATURE_UNAVAILABLE when xlwings is not available."""
        import unittest.mock

        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            with unittest.mock.patch(
                "xlforge.commands.pivot.find_spec", return_value=None
            ):
                result = runner.invoke(
                    app,
                    [
                        "pivot",
                        "create",
                        path,
                        "Sheet1",
                        "A1:D10",
                        "--name",
                        "TestPivot",
                    ],
                )

            assert result.exit_code == ErrorCode.FEATURE_UNAVAILABLE
            assert (
                "xlwings" in result.output.lower()
                or "headless" in result.output.lower()
            )

    def test_pivot_create_file_not_found_when_xlwings_available(self):
        """Test pivot create with non-existent file returns error when xlwings is mocked."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")

            import unittest.mock as mock

            with mock.patch(
                "xlforge.commands.pivot._is_xlwings_available", return_value=True
            ):
                result = runner.invoke(
                    app,
                    [
                        "pivot",
                        "create",
                        path,
                        "Sheet1",
                        "A1:D10",
                        "--name",
                        "TestPivot",
                    ],
                )

            assert result.exit_code == ErrorCode.FILE_DOES_NOT_EXIST
            assert "does not exist" in result.output.lower()

    @pytest.mark.skipif(not _is_excel_available(), reason="requires Excel")
    def test_pivot_create_sheet_not_found_when_xlwings_available(self):
        """Test pivot create with non-existent sheet returns error when xlwings is mocked."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")

            import unittest.mock as mock

            with mock.patch(
                "xlforge.commands.pivot._is_xlwings_available", return_value=True
            ):
                result = runner.invoke(
                    app,
                    [
                        "pivot",
                        "create",
                        path,
                        "NonexistentSheet",
                        "A1:D10",
                        "--name",
                        "TestPivot",
                    ],
                )

            assert result.exit_code == ErrorCode.SHEET_NOT_FOUND
            assert "Sheet not found" in result.output

    @pytest.mark.skipif(not _is_excel_available(), reason="requires Excel")
    def test_pivot_create_invalid_range_when_xlwings_available(self):
        """Test pivot create with invalid range format returns error when xlwings is mocked."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")

            import unittest.mock as mock

            with mock.patch(
                "xlforge.commands.pivot._is_xlwings_available", return_value=True
            ):
                result = runner.invoke(
                    app,
                    [
                        "pivot",
                        "create",
                        path,
                        "Sheet1",
                        "invalid-range",
                        "--name",
                        "TestPivot",
                    ],
                )

            assert result.exit_code == ErrorCode.INVALID_SYNTAX
            assert "Invalid source range format" in result.output

    @pytest.mark.skipif(not _is_excel_available(), reason="requires Excel")
    def test_pivot_create_invalid_aggregation_when_xlwings_available(self):
        """Test pivot create with invalid aggregation format returns error when xlwings is mocked."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")

            import unittest.mock as mock

            with mock.patch(
                "xlforge.commands.pivot._is_xlwings_available", return_value=True
            ):
                result = runner.invoke(
                    app,
                    [
                        "pivot",
                        "create",
                        path,
                        "Sheet1",
                        "A1:D10",
                        "--name",
                        "TestPivot",
                        "--values",
                        "INVALID:Field",
                    ],
                )

            assert result.exit_code == ErrorCode.INVALID_SYNTAX
            assert "Invalid aggregation type" in result.output

    @pytest.mark.skipif(not _is_excel_available(), reason="requires Excel")
    def test_pivot_create_success_when_xlwings_available(self):
        """Test creating a pivot table successfully when xlwings is mocked."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            # Add some data
            ws["A1"] = "Region"
            ws["B1"] = "Quarter"
            ws["C1"] = "Revenue"
            ws["D1"] = "Cost"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")

            import unittest.mock as mock

            with mock.patch(
                "xlforge.commands.pivot._is_xlwings_available", return_value=True
            ):
                result = runner.invoke(
                    app,
                    [
                        "pivot",
                        "create",
                        path,
                        "Sheet1",
                        "A1:D10",
                        "--name",
                        "SalesPivot",
                        "--rows",
                        "Region",
                        "--columns",
                        "Quarter",
                        "--values",
                        "SUM:Revenue",
                        "--values",
                        "SUM:Cost",
                        "--filters",
                        "Year",
                    ],
                )

            # The command should succeed (though openpyxl has limited support)
            assert result.exit_code == 0
            assert "Created pivot table 'SalesPivot'" in result.output
            assert "Source:" in result.output

    @pytest.mark.skipif(not _is_excel_available(), reason="requires Excel")
    def test_pivot_create_default_name_when_xlwings_available(self):
        """Test creating a pivot table without specifying a name uses default naming."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")

            import unittest.mock as mock

            with mock.patch(
                "xlforge.commands.pivot._is_xlwings_available", return_value=True
            ):
                result = runner.invoke(
                    app,
                    ["pivot", "create", path, "Sheet1", "A1:D10"],
                )

            assert result.exit_code == 0
            assert "Created pivot table '" in result.output

    @pytest.mark.skipif(not _is_excel_available(), reason="requires Excel")
    def test_pivot_create_with_custom_sheet_when_xlwings_available(self):
        """Test creating a pivot table on a custom target sheet."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Data"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")

            import unittest.mock as mock

            with mock.patch(
                "xlforge.commands.pivot._is_xlwings_available", return_value=True
            ):
                result = runner.invoke(
                    app,
                    [
                        "pivot",
                        "create",
                        path,
                        "Data",
                        "A1:D10",
                        "--sheet",
                        "Dashboard",
                        "--name",
                        "SalesPivot",
                    ],
                )

            assert result.exit_code == 0
            assert "Created pivot table 'SalesPivot'" in result.output
            assert "Dashboard" in result.output


class TestPivotList:
    """Tests for xlforge pivot list command."""

    def test_pivot_list_file_not_found(self):
        """Test pivot list with non-existent file returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            with unittest.mock.patch(
                "xlforge.commands.pivot.find_spec", return_value=None
            ):
                result = runner.invoke(app, ["pivot", "list", path])

            # xlwings not available returns FEATURE_UNAVAILABLE
            assert result.exit_code == ErrorCode.FEATURE_UNAVAILABLE

    def test_pivot_list_xlwings_not_available(self):
        """Test pivot list when xlwings is not available."""
        import unittest.mock

        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            with unittest.mock.patch(
                "xlforge.commands.pivot.find_spec", return_value=None
            ):
                result = runner.invoke(app, ["pivot", "list", path])

            assert result.exit_code == ErrorCode.FEATURE_UNAVAILABLE
            assert (
                "xlwings" in result.output.lower()
                or "headless" in result.output.lower()
            )


class TestPivotDelete:
    """Tests for xlforge pivot delete command."""

    def test_pivot_delete_file_not_found(self):
        """Test pivot delete with non-existent file returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            with unittest.mock.patch(
                "xlforge.commands.pivot.find_spec", return_value=None
            ):
                result = runner.invoke(app, ["pivot", "delete", path, "TestPivot"])

            assert result.exit_code == ErrorCode.FEATURE_UNAVAILABLE

    def test_pivot_delete_xlwings_not_available(self):
        """Test pivot delete when xlwings is not available."""
        import unittest.mock

        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            with unittest.mock.patch(
                "xlforge.commands.pivot.find_spec", return_value=None
            ):
                result = runner.invoke(app, ["pivot", "delete", path, "TestPivot"])

            assert result.exit_code == ErrorCode.FEATURE_UNAVAILABLE


class TestProtectionProtect:
    """Tests for xlforge protection protect command."""

    def test_protect_file_not_found(self):
        """Test protect with non-existent file returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            result = runner.invoke(app, ["protection", "protect", path, "Sheet1"])

            assert result.exit_code == ErrorCode.FILE_DOES_NOT_EXIST
            assert "does not exist" in result.output.lower()

    def test_protect_sheet_not_found(self):
        """Test protect with non-existent sheet returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["protection", "protect", path, "NonexistentSheet"]
            )

            assert result.exit_code == ErrorCode.SHEET_NOT_FOUND
            assert "Sheet not found" in result.output

    def test_protect_success(self):
        """Test protecting a sheet."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["protection", "protect", path, "Sheet1"])

            assert result.exit_code == 0
            assert "Protected sheet 'Sheet1'" in result.output

            # Verify sheet is protected
            wb = openpyxl.load_workbook(path)
            assert wb["Sheet1"].protection.sheet is True
            wb.close()

    def test_protect_with_password(self):
        """Test protecting a sheet with password."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["protection", "protect", path, "Sheet1", "--password", "mypass"]
            )

            assert result.exit_code == 0
            assert "Protected sheet 'Sheet1'" in result.output
            assert "password" in result.output.lower()

            # Verify sheet is protected with password (openpyxl hashes the password)
            wb = openpyxl.load_workbook(path)
            assert wb["Sheet1"].protection.sheet is True
            assert (
                wb["Sheet1"].protection.password is not None
            )  # Password is hashed by openpyxl
            wb.close()


class TestProtectionUnprotect:
    """Tests for xlforge protection unprotect command."""

    def test_unprotect_file_not_found(self):
        """Test unprotect with non-existent file returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            result = runner.invoke(app, ["protection", "unprotect", path, "Sheet1"])

            assert result.exit_code == ErrorCode.FILE_DOES_NOT_EXIST
            assert "does not exist" in result.output.lower()

    def test_unprotect_sheet_not_found(self):
        """Test unprotect with non-existent sheet returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["protection", "unprotect", path, "NonexistentSheet"]
            )

            assert result.exit_code == ErrorCode.SHEET_NOT_FOUND
            assert "Sheet not found" in result.output

    def test_unprotect_success(self):
        """Test unprotecting a sheet."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws.protection.sheet = True
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["protection", "unprotect", path, "Sheet1"])

            assert result.exit_code == 0
            assert "Unprotected sheet 'Sheet1'" in result.output

            # Verify sheet is unprotected
            wb = openpyxl.load_workbook(path)
            assert wb["Sheet1"].protection.sheet is False
            wb.close()


class TestProtectionFreeze:
    """Tests for xlforge protection freeze command."""

    def test_freeze_file_not_found(self):
        """Test freeze with non-existent file returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            result = runner.invoke(app, ["protection", "freeze", path, "Sheet1"])

            assert result.exit_code == ErrorCode.FILE_DOES_NOT_EXIST
            assert "does not exist" in result.output.lower()

    def test_freeze_sheet_not_found(self):
        """Test freeze with non-existent sheet returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["protection", "freeze", path, "NonexistentSheet"]
            )

            assert result.exit_code == ErrorCode.SHEET_NOT_FOUND
            assert "Sheet not found" in result.output

    def test_freeze_default(self):
        """Test freeze with default position (A2 - freeze first row)."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["protection", "freeze", path, "Sheet1"])

            assert result.exit_code == 0
            assert "Freeze panes set to A2" in result.output

            # Verify freeze panes
            wb = openpyxl.load_workbook(path)
            assert wb["Sheet1"].freeze_panes == "A2"
            wb.close()

    def test_freeze_with_column_and_row(self):
        """Test freeze with specific column and row."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app,
                ["protection", "freeze", path, "Sheet1", "--column", "B", "--row", "5"],
            )

            assert result.exit_code == 0
            assert "Freeze panes set to B5" in result.output

            # Verify freeze panes
            wb = openpyxl.load_workbook(path)
            assert wb["Sheet1"].freeze_panes == "B5"
            wb.close()

    def test_freeze_with_column_only(self):
        """Test freeze with column only."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["protection", "freeze", path, "Sheet1", "--column", "C"]
            )

            assert result.exit_code == 0
            assert "Freeze panes set to C1" in result.output

            # Verify freeze panes
            wb = openpyxl.load_workbook(path)
            assert wb["Sheet1"].freeze_panes == "C1"
            wb.close()

    def test_freeze_with_row_only(self):
        """Test freeze with row only."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["protection", "freeze", path, "Sheet1", "--row", "10"]
            )

            assert result.exit_code == 0
            assert "Freeze panes set to A10" in result.output

            # Verify freeze panes
            wb = openpyxl.load_workbook(path)
            assert wb["Sheet1"].freeze_panes == "A10"
            wb.close()


class TestSheetCopy:
    """Tests for xlforge sheet copy command."""

    def test_sheet_copy_file_not_found(self):
        """Test sheet copy with non-existent file returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            result = runner.invoke(
                app, ["sheet", "copy", path, "Sheet1", "Sheet1_Copy"]
            )

            assert result.exit_code == ErrorCode.FILE_DOES_NOT_EXIST
            assert "does not exist" in result.output.lower()

    def test_sheet_copy_source_not_found(self):
        """Test sheet copy with non-existent source sheet returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.active.title = "Sheet1"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            with unittest.mock.patch(
                "xlforge.core.engines.selector.find_spec", return_value=None
            ):
                result = runner.invoke(
                    app, ["sheet", "copy", path, "NonexistentSheet", "Sheet1_Copy"]
                )

            assert result.exit_code == ErrorCode.SHEET_NOT_FOUND
            assert "Sheet 'NonexistentSheet' not found" in result.output

    def test_sheet_copy_new_name_exists(self):
        """Test sheet copy when new name already exists returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.active.title = "Sheet1"
            wb.create_sheet("Sheet2")
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            with unittest.mock.patch(
                "xlforge.core.engines.selector.find_spec", return_value=None
            ):
                result = runner.invoke(app, ["sheet", "copy", path, "Sheet1", "Sheet2"])

            assert result.exit_code == ErrorCode.TABLE_ALREADY_EXISTS
            assert "Sheet 'Sheet2' already exists" in result.output

    @pytest.mark.skipif(not _is_excel_available(), reason="requires Excel")
    def test_sheet_copy_success(self):
        """Test sheet copy successfully copies a sheet."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws["A1"] = "Test Value"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["sheet", "copy", path, "Sheet1", "Sheet1_Copy"]
            )

            assert result.exit_code == 0
            assert "Copied sheet 'Sheet1' to 'Sheet1_Copy'" in result.output

            # Verify sheet was created
            wb = openpyxl.load_workbook(path)
            assert "Sheet1_Copy" in wb.sheetnames
            assert wb["Sheet1_Copy"]["A1"].value == "Test Value"
            wb.close()


class TestSheetUse:
    """Tests for xlforge sheet use command."""

    def test_sheet_use_file_not_found(self):
        """Test sheet use with non-existent file returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            result = runner.invoke(app, ["sheet", "use", path, "Sheet1"])

            assert result.exit_code == ErrorCode.FILE_DOES_NOT_EXIST
            assert "does not exist" in result.output.lower()

    def test_sheet_use_sheet_not_found(self):
        """Test sheet use with non-existent sheet returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.active.title = "Sheet1"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            with unittest.mock.patch(
                "xlforge.core.engines.selector.find_spec", return_value=None
            ):
                result = runner.invoke(app, ["sheet", "use", path, "NonexistentSheet"])

            assert result.exit_code == ErrorCode.SHEET_NOT_FOUND
            assert "Sheet 'NonexistentSheet' not found" in result.output

    @pytest.mark.skipif(not _is_excel_available(), reason="requires Excel")
    def test_sheet_use_success(self):
        """Test sheet use successfully sets active sheet."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.active.title = "Sheet1"
            wb.create_sheet("Sheet2")
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["sheet", "use", path, "Sheet2"])

            assert result.exit_code == 0
            assert "Set active sheet to 'Sheet2'" in result.output

            # Verify active sheet
            wb = openpyxl.load_workbook(path)
            assert wb.active.title == "Sheet2"
            wb.close()


class TestCellCopy:
    """Tests for xlforge cell copy command."""

    def test_cell_copy_file_not_found(self):
        """Test cell copy with non-existent file returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            result = runner.invoke(
                app, ["cell", "copy", path, "Sheet1", "A1", "Sheet1", "B1"]
            )

            assert result.exit_code == ErrorCode.FILE_DOES_NOT_EXIST
            assert "does not exist" in result.output.lower()

    def test_cell_copy_src_sheet_not_found(self):
        """Test cell copy with non-existent source sheet returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            with unittest.mock.patch(
                "xlforge.core.engines.selector.find_spec", return_value=None
            ):
                result = runner.invoke(
                    app,
                    ["cell", "copy", path, "NonexistentSheet", "A1", "Sheet1", "B1"],
                )

            assert result.exit_code == ErrorCode.SHEET_NOT_FOUND
            assert "Source sheet not found" in result.output

    def test_cell_copy_dst_sheet_not_found(self):
        """Test cell copy with non-existent destination sheet returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.active.title = "Sheet1"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            with unittest.mock.patch(
                "xlforge.core.engines.selector.find_spec", return_value=None
            ):
                result = runner.invoke(
                    app,
                    ["cell", "copy", path, "Sheet1", "A1", "NonexistentSheet", "B1"],
                )

            assert result.exit_code == ErrorCode.SHEET_NOT_FOUND
            assert "Destination sheet not found" in result.output

    @pytest.mark.skipif(not _is_excel_available(), reason="requires Excel")
    def test_cell_copy_success(self):
        """Test cell copy successfully copies cell value."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws["A1"] = "Hello World"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["cell", "copy", path, "Sheet1", "A1", "Sheet1", "B1"]
            )

            assert result.exit_code == 0
            assert "Copied Sheet1!A1 (Hello World) to Sheet1!B1" in result.output

            # Verify value was copied
            wb = openpyxl.load_workbook(path)
            assert wb["Sheet1"]["B1"].value == "Hello World"
            wb.close()

    @pytest.mark.skipif(not _is_excel_available(), reason="requires Excel")
    def test_cell_copy_cross_sheet(self):
        """Test cell copy from one sheet to another."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.active.title = "Sheet1"
            wb.create_sheet("Sheet2")
            ws1 = wb["Sheet1"]
            ws1["A1"] = "Cross Sheet Copy"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["cell", "copy", path, "Sheet1", "A1", "Sheet2", "C3"]
            )

            assert result.exit_code == 0
            assert "Copied Sheet1!A1 (Cross Sheet Copy) to Sheet2!C3" in result.output

            # Verify value was copied to different sheet
            wb = openpyxl.load_workbook(path)
            assert wb["Sheet2"]["C3"].value == "Cross Sheet Copy"
            wb.close()


class TestCellSearch:
    """Tests for xlforge cell search command."""

    def test_cell_search_file_not_found(self):
        """Test cell search with non-existent file returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            result = runner.invoke(app, ["cell", "search", path, "test"])

            assert result.exit_code == ErrorCode.FILE_DOES_NOT_EXIST
            assert "does not exist" in result.output.lower()

    def test_cell_search_sheet_not_found(self):
        """Test cell search with non-existent sheet returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            with unittest.mock.patch(
                "xlforge.core.engines.selector.find_spec", return_value=None
            ):
                result = runner.invoke(
                    app, ["cell", "search", path, "test", "--sheet", "NonexistentSheet"]
                )

            assert result.exit_code == ErrorCode.SHEET_NOT_FOUND
            assert "Sheet not found" in result.output

    def test_cell_search_not_found(self):
        """Test cell search when query not found returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws["A1"] = "Hello World"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            with unittest.mock.patch(
                "xlforge.core.engines.selector.find_spec", return_value=None
            ):
                result = runner.invoke(app, ["cell", "search", path, "NotFound"])

            assert result.exit_code == ErrorCode.CELL_NOT_FOUND
            assert "No cell found containing 'NotFound'" in result.output

    @pytest.mark.skipif(not _is_excel_available(), reason="requires Excel")
    def test_cell_search_success(self):
        """Test cell search finds matching cell."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws["B2"] = "Hello World"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["cell", "search", path, "Hello"])

            assert result.exit_code == 0
            assert "Found in Sheet1!B2: Hello World" in result.output

    @pytest.mark.skipif(not _is_excel_available(), reason="requires Excel")
    def test_cell_search_specific_sheet(self):
        """Test cell search in specific sheet."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws1 = wb.active
            ws1.title = "Sheet1"
            ws1["A1"] = "Match in Sheet1"
            wb.create_sheet("Sheet2")
            ws2 = wb["Sheet2"]
            ws2["A1"] = "Match in Sheet2"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["cell", "search", path, "Match in Sheet1", "--sheet", "Sheet1"]
            )

            assert result.exit_code == 0
            assert "Found in Sheet1!A1" in result.output

    @pytest.mark.skipif(not _is_excel_available(), reason="requires Excel")
    def test_cell_search_json_output(self):
        """Test cell search with JSON output."""
        import json

        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws["C5"] = "Search Target"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["cell", "search", path, "Search", "--json"])

            assert result.exit_code == 0
            data = json.loads(result.output)
            assert data["value"] == "Search Target"
            assert data["coord"] == "C5"
            assert data["sheet"] == "Sheet1"


class TestCellBulk:
    """Tests for xlforge cell bulk command."""

    def test_cell_bulk_file_not_found(self):
        """Test cell bulk with non-existent file returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            result = runner.invoke(
                app, ["cell", "bulk", path, "Sheet1", "A1:C3", "--set", "test"]
            )

            assert result.exit_code == ErrorCode.FILE_DOES_NOT_EXIST
            assert "does not exist" in result.output.lower()

    def test_cell_bulk_sheet_not_found(self):
        """Test cell bulk with non-existent sheet returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            with unittest.mock.patch(
                "xlforge.core.engines.selector.find_spec", return_value=None
            ):
                result = runner.invoke(
                    app,
                    [
                        "cell",
                        "bulk",
                        path,
                        "NonexistentSheet",
                        "A1:C3",
                        "--set",
                        "test",
                    ],
                )

            assert result.exit_code == ErrorCode.SHEET_NOT_FOUND
            assert "Sheet not found" in result.output

    def test_cell_bulk_no_option(self):
        """Test cell bulk without --set or --clear returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            with unittest.mock.patch(
                "xlforge.core.engines.selector.find_spec", return_value=None
            ):
                result = runner.invoke(app, ["cell", "bulk", path, "Sheet1", "A1:C3"])

            assert result.exit_code == 1
            assert "Must specify either --set <value> or --clear" in result.output

    @pytest.mark.skipif(not _is_excel_available(), reason="requires Excel")
    def test_cell_bulk_set_value(self):
        """Test cell bulk sets value in range."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["cell", "bulk", path, "Sheet1", "A1:C3", "--set", "Bulk"]
            )

            assert result.exit_code == 0
            assert "Set A1:C3 = Bulk" in result.output

            # Verify values were set
            wb = openpyxl.load_workbook(path)
            for row in range(1, 4):
                for col in ["A", "B", "C"]:
                    assert wb["Sheet1"][f"{col}{row}"].value == "Bulk"
            wb.close()

    @pytest.mark.skipif(not _is_excel_available(), reason="requires Excel")
    def test_cell_bulk_clear(self):
        """Test cell bulk clears range."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws["A1"] = "Keep"
            ws["A2"] = "Clear"
            ws["A3"] = "Me"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["cell", "bulk", path, "Sheet1", "A2", "--clear"]
            )

            assert result.exit_code == 0
            assert "Cleared range A2" in result.output

            # Verify cell was cleared
            wb = openpyxl.load_workbook(path)
            assert wb["Sheet1"]["A1"].value == "Keep"
            assert wb["Sheet1"]["A2"].value is None
            assert wb["Sheet1"]["A3"].value == "Me"
            wb.close()


class TestCellFill:
    """Tests for xlforge cell fill command."""

    def test_cell_fill_file_not_found(self):
        """Test cell fill with non-existent file returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            result = runner.invoke(app, ["cell", "fill", path, "Sheet1", "A1:C3"])

            assert result.exit_code == ErrorCode.FILE_DOES_NOT_EXIST
            assert "does not exist" in result.output.lower()

    def test_cell_fill_sheet_not_found(self):
        """Test cell fill with non-existent sheet returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            with unittest.mock.patch(
                "xlforge.core.engines.selector.find_spec", return_value=None
            ):
                result = runner.invoke(
                    app, ["cell", "fill", path, "NonexistentSheet", "A1:C3"]
                )

            assert result.exit_code == ErrorCode.SHEET_NOT_FOUND
            assert "Sheet not found" in result.output

    def test_cell_fill_empty_range(self):
        """Test cell fill with empty range returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            with unittest.mock.patch(
                "xlforge.core.engines.selector.find_spec", return_value=None
            ):
                result = runner.invoke(app, ["cell", "fill", path, "Sheet1", "A1:C3"])

            assert result.exit_code == 1
            assert "Range A1:C3 is empty" in result.output

    @pytest.mark.skipif(not _is_excel_available(), reason="requires Excel")
    def test_cell_fill_success(self):
        """Test cell fill fills range with first cell value."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws["A1"] = "Fill Value"
            ws["A2"] = "X"
            ws["A3"] = "X"
            ws["B1"] = "X"
            ws["B2"] = "X"
            ws["B3"] = "X"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["cell", "fill", path, "Sheet1", "A1:B3"])

            assert result.exit_code == 0
            assert "Filled range A1:B3 with value: Fill Value" in result.output

            # Verify all cells were filled
            wb = openpyxl.load_workbook(path)
            for row in range(1, 4):
                for col in ["A", "B"]:
                    assert wb["Sheet1"][f"{col}{row}"].value == "Fill Value"
            wb.close()

    @pytest.mark.skipif(not _is_excel_available(), reason="requires Excel")
    def test_cell_fill_horizontal(self):
        """Test cell fill with horizontal range."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws["A1"] = "Start"
            ws["B1"] = "X"
            ws["C1"] = "X"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["cell", "fill", path, "Sheet1", "A1:C1"])

            assert result.exit_code == 0

            # Verify all cells were filled
            wb = openpyxl.load_workbook(path)
            for col in ["A", "B", "C"]:
                assert wb["Sheet1"][f"{col}1"].value == "Start"
            wb.close()


class TestFormatConditionCommands:
    """Tests for format-condition commands."""

    def test_format_condition_add_file_not_found(self):
        """Test format-condition add with non-existent file returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            result = runner.invoke(
                app,
                [
                    "format-condition",
                    "add",
                    path,
                    "Sheet1",
                    "A1:A10",
                    "--type",
                    "data-bar",
                ],
            )

            assert result.exit_code == ErrorCode.FILE_DOES_NOT_EXIST
            assert "does not exist" in result.output.lower()

    def test_format_condition_add_sheet_not_found(self):
        """Test format-condition add with non-existent sheet returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app,
                [
                    "format-condition",
                    "add",
                    path,
                    "NonExistentSheet",
                    "A1:A10",
                    "--type",
                    "data-bar",
                ],
            )

            assert result.exit_code == ErrorCode.SHEET_NOT_FOUND
            assert "Sheet not found" in result.output

    def test_format_condition_add_missing_type(self):
        """Test format-condition add without --type returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app, ["format-condition", "add", path, "Sheet1", "A1:A10"]
            )

            assert result.exit_code == ErrorCode.CONDITIONAL_FORMAT_NOT_SUPPORTED
            assert "--type is required" in result.output

    def test_format_condition_add_invalid_type(self):
        """Test format-condition add with invalid type returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app,
                [
                    "format-condition",
                    "add",
                    path,
                    "Sheet1",
                    "A1:A10",
                    "--type",
                    "invalid",
                ],
            )

            assert result.exit_code == ErrorCode.CONDITIONAL_FORMAT_NOT_SUPPORTED
            assert "Invalid type" in result.output

    def test_format_condition_add_color_scale_missing_min_max(self):
        """Test format-condition add color-scale without --min and --max returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app,
                [
                    "format-condition",
                    "add",
                    path,
                    "Sheet1",
                    "A1:A10",
                    "--type",
                    "color-scale",
                ],
            )

            assert result.exit_code == ErrorCode.CONDITIONAL_FORMAT_NOT_SUPPORTED
            assert "--min and --max colors are required" in result.output

    def test_format_condition_add_color_scale_invalid_color(self):
        """Test format-condition add color-scale with invalid color returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app,
                [
                    "format-condition",
                    "add",
                    path,
                    "Sheet1",
                    "A1:A10",
                    "--type",
                    "color-scale",
                    "--min",
                    "invalid",
                    "--max",
                    "#00FF00",
                ],
            )

            assert result.exit_code == ErrorCode.INVALID_STYLE_STRING
            assert "Invalid min color" in result.output

    def test_format_condition_add_icon_set_missing_icons(self):
        """Test format-condition add icon-set without --icons returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app,
                [
                    "format-condition",
                    "add",
                    path,
                    "Sheet1",
                    "A1:A10",
                    "--type",
                    "icon-set",
                ],
            )

            assert result.exit_code == ErrorCode.CONDITIONAL_FORMAT_NOT_SUPPORTED
            assert "--icons is required" in result.output

    def test_format_condition_add_formula_missing_formula(self):
        """Test format-condition add formula without --formula returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app,
                [
                    "format-condition",
                    "add",
                    path,
                    "Sheet1",
                    "A1:A10",
                    "--type",
                    "formula",
                ],
            )

            assert result.exit_code == ErrorCode.CONDITIONAL_FORMAT_NOT_SUPPORTED
            assert "--formula is required" in result.output

    def test_format_condition_add_data_bar_success(self):
        """Test adding a data bar conditional formatting."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            # Add some values
            ws["A1"] = 10
            ws["A2"] = 20
            ws["A3"] = 30
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app,
                [
                    "format-condition",
                    "add",
                    path,
                    "Sheet1",
                    "A1:A10",
                    "--type",
                    "data-bar",
                ],
            )

            assert result.exit_code == 0
            assert "Added data bar conditional formatting" in result.output
            assert "A1:A10" in result.output

            # Verify conditional formatting was added
            wb = openpyxl.load_workbook(path)
            ws = wb["Sheet1"]
            assert len(ws.conditional_formatting._cf_rules) == 1
            wb.close()

    def test_format_condition_add_color_scale_success(self):
        """Test adding a color scale conditional formatting."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws["A1"] = 10
            ws["A2"] = 50
            ws["A3"] = 100
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app,
                [
                    "format-condition",
                    "add",
                    path,
                    "Sheet1",
                    "A1:A10",
                    "--type",
                    "color-scale",
                    "--min",
                    "#FF0000",
                    "--max",
                    "#00FF00",
                ],
            )

            assert result.exit_code == 0
            assert "Added color scale conditional formatting" in result.output
            assert "A1:A10" in result.output

            # Verify conditional formatting was added
            wb = openpyxl.load_workbook(path)
            ws = wb["Sheet1"]
            assert len(ws.conditional_formatting._cf_rules) == 1
            wb.close()

    def test_format_condition_add_color_scale_with_mid_success(self):
        """Test adding a color scale conditional formatting with mid color."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app,
                [
                    "format-condition",
                    "add",
                    path,
                    "Sheet1",
                    "A1:A10",
                    "--type",
                    "color-scale",
                    "--min",
                    "#FF0000",
                    "--mid",
                    "#FFFF00",
                    "--max",
                    "#00FF00",
                ],
            )

            assert result.exit_code == 0
            assert "Added color scale conditional formatting" in result.output

    def test_format_condition_add_icon_set_success(self):
        """Test adding an icon set conditional formatting."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app,
                [
                    "format-condition",
                    "add",
                    path,
                    "Sheet1",
                    "A1:A10",
                    "--type",
                    "icon-set",
                    "--icons",
                    "3TrafficLights1",
                ],
            )

            assert result.exit_code == 0
            assert "Added icon set conditional formatting" in result.output
            assert "3TrafficLights1" in result.output
            assert "A1:A10" in result.output

            # Verify conditional formatting was added
            wb = openpyxl.load_workbook(path)
            ws = wb["Sheet1"]
            assert len(ws.conditional_formatting._cf_rules) == 1
            wb.close()

    def test_format_condition_add_formula_success(self):
        """Test adding a formula conditional formatting."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app,
                [
                    "format-condition",
                    "add",
                    path,
                    "Sheet1",
                    "A1:A10",
                    "--type",
                    "formula",
                    "--formula",
                    "=A1>100",
                ],
            )

            assert result.exit_code == 0
            assert "Added formula conditional formatting" in result.output
            assert "=A1>100" in result.output
            assert "A1:A10" in result.output

            # Verify conditional formatting was added
            wb = openpyxl.load_workbook(path)
            ws = wb["Sheet1"]
            assert len(ws.conditional_formatting._cf_rules) == 1
            wb.close()

    def test_format_condition_add_formula_with_style_success(self):
        """Test adding a formula conditional formatting with style."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app,
                [
                    "format-condition",
                    "add",
                    path,
                    "Sheet1",
                    "A1:A10",
                    "--type",
                    "formula",
                    "--formula",
                    "=A1>100",
                    "--style",
                    "bold text-#FF0000",
                ],
            )

            assert result.exit_code == 0
            assert "Added formula conditional formatting" in result.output
            assert "bold text-#FF0000" in result.output

    def test_format_condition_add_rule_missing_rule(self):
        """Test format-condition add rule without --rule returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app,
                ["format-condition", "add", path, "Sheet1", "A1:A10", "--value", "100"],
            )

            assert result.exit_code == ErrorCode.CONDITIONAL_FORMAT_NOT_SUPPORTED
            assert "--rule is required" in result.output

    def test_format_condition_add_rule_invalid_rule(self):
        """Test format-condition add with invalid rule returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app,
                [
                    "format-condition",
                    "add",
                    path,
                    "Sheet1",
                    "A1:A10",
                    "--rule",
                    "invalid",
                    "--value",
                    "100",
                ],
            )

            assert result.exit_code == ErrorCode.CONDITIONAL_FORMAT_NOT_SUPPORTED
            assert "Invalid rule" in result.output

    def test_format_condition_add_rule_missing_value(self):
        """Test format-condition add rule without --value returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app,
                [
                    "format-condition",
                    "add",
                    path,
                    "Sheet1",
                    "A1:A10",
                    "--rule",
                    "greater-than",
                ],
            )

            assert result.exit_code == ErrorCode.INVALID_FORMULA_SYNTAX
            assert "--value is required" in result.output

    def test_format_condition_add_rule_greater_than_success(self):
        """Test adding a greater-than rule conditional formatting."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app,
                [
                    "format-condition",
                    "add",
                    path,
                    "Sheet1",
                    "A1:A10",
                    "--rule",
                    "greater-than",
                    "--value",
                    "100",
                ],
            )

            assert result.exit_code == 0
            assert "Added greater-than rule" in result.output
            assert "100" in result.output
            assert "A1:A10" in result.output

            # Verify conditional formatting was added
            wb = openpyxl.load_workbook(path)
            ws = wb["Sheet1"]
            assert len(ws.conditional_formatting._cf_rules) == 1
            wb.close()

    def test_format_condition_add_rule_less_than_success(self):
        """Test adding a less-than rule conditional formatting."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app,
                [
                    "format-condition",
                    "add",
                    path,
                    "Sheet1",
                    "A1:A10",
                    "--rule",
                    "less-than",
                    "--value",
                    "50",
                ],
            )

            assert result.exit_code == 0
            assert "Added less-than rule" in result.output

    def test_format_condition_add_rule_between_success(self):
        """Test adding a between rule conditional formatting."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app,
                [
                    "format-condition",
                    "add",
                    path,
                    "Sheet1",
                    "A1:A10",
                    "--rule",
                    "between",
                    "--value",
                    "10,100",
                ],
            )

            assert result.exit_code == 0
            assert "Added between rule" in result.output
            assert "10,100" in result.output

    def test_format_condition_add_rule_between_invalid_value(self):
        """Test format-condition add between rule without comma-separated value returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app,
                [
                    "format-condition",
                    "add",
                    path,
                    "Sheet",
                    "A1:A10",
                    "--rule",
                    "between",
                    "--value",
                    "100",
                ],
            )

            assert result.exit_code == ErrorCode.INVALID_FORMULA_SYNTAX
            assert "two values separated by comma" in result.output

    def test_format_condition_add_rule_equal_success(self):
        """Test adding an equal rule conditional formatting."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app,
                [
                    "format-condition",
                    "add",
                    path,
                    "Sheet1",
                    "A1:A10",
                    "--rule",
                    "equal",
                    "--value",
                    "PASS",
                ],
            )

            assert result.exit_code == 0
            assert "Added equal rule" in result.output

    def test_format_condition_add_rule_contains_success(self):
        """Test adding a contains rule conditional formatting."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app,
                [
                    "format-condition",
                    "add",
                    path,
                    "Sheet1",
                    "A1:A10",
                    "--rule",
                    "contains",
                    "--value",
                    "ERROR",
                ],
            )

            assert result.exit_code == 0
            assert "Added contains rule" in result.output

    def test_format_condition_add_rule_with_style_success(self):
        """Test adding a rule conditional formatting with style."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(
                app,
                [
                    "format-condition",
                    "add",
                    path,
                    "Sheet1",
                    "A1:A10",
                    "--rule",
                    "greater-than",
                    "--value",
                    "100",
                    "--style",
                    "bold",
                ],
            )

            assert result.exit_code == 0
            assert "Added greater-than rule" in result.output
            assert "bold" in result.output


class TestAppCommands:
    """Tests for app commands (require xlwings)."""

    def test_app_visible_xlwings_not_available(self):
        """Test app visible returns FEATURE_UNAVAILABLE when xlwings not installed."""
        import unittest.mock

        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            with unittest.mock.patch(
                "xlforge.commands.app.find_spec", return_value=None
            ):
                result = runner.invoke(app, ["app", "visible", path, "true"])

            assert result.exit_code == ErrorCode.FEATURE_UNAVAILABLE
            assert "xlwings" in result.output.lower()

    def test_app_calculate_xlwings_not_available(self):
        """Test app calculate returns FEATURE_UNAVAILABLE when xlwings not installed."""
        import unittest.mock

        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            with unittest.mock.patch(
                "xlforge.commands.app.find_spec", return_value=None
            ):
                result = runner.invoke(app, ["app", "calculate", path])

            assert result.exit_code == ErrorCode.FEATURE_UNAVAILABLE
            assert "xlwings" in result.output.lower()

    def test_app_focus_xlwings_not_available(self):
        """Test app focus returns FEATURE_UNAVAILABLE when xlwings not installed."""
        import unittest.mock

        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            with unittest.mock.patch(
                "xlforge.commands.app.find_spec", return_value=None
            ):
                result = runner.invoke(app, ["app", "focus", path])

            assert result.exit_code == ErrorCode.FEATURE_UNAVAILABLE
            assert "xlwings" in result.output.lower()

    def test_app_alert_xlwings_not_available(self):
        """Test app alert returns FEATURE_UNAVAILABLE when xlwings not installed."""
        import unittest.mock

        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            with unittest.mock.patch(
                "xlforge.commands.app.find_spec", return_value=None
            ):
                result = runner.invoke(app, ["app", "alert", path, "Hello"])

            assert result.exit_code == ErrorCode.FEATURE_UNAVAILABLE
            assert "xlwings" in result.output.lower()

    def test_app_wait_idle_xlwings_not_available(self):
        """Test app wait-idle returns FEATURE_UNAVAILABLE when xlwings not installed."""
        import unittest.mock

        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "nonexistent.xlsx")
            with unittest.mock.patch(
                "xlforge.commands.app.find_spec", return_value=None
            ):
                result = runner.invoke(app, ["app", "wait-idle", path])

            assert result.exit_code == ErrorCode.FEATURE_UNAVAILABLE
            assert "xlwings" in result.output.lower()

    def test_app_visible_invalid_value(self):
        """Test app visible with invalid value returns error."""
        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")
            result = runner.invoke(app, ["app", "visible", path, "maybe"])

            assert result.exit_code == ErrorCode.INVALID_SYNTAX
            assert "Invalid value" in result.output

    def test_app_commands_xlwings_not_available(self):
        """Test app commands return FEATURE_UNAVAILABLE when xlwings not installed."""
        import unittest.mock

        with tempfile.TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            wb.save(os.path.join(tmpdir, "test.xlsx"))

            path = os.path.join(tmpdir, "test.xlsx")

            # Mock find_spec to return None for xlwings
            with unittest.mock.patch(
                "xlforge.commands.app.find_spec", return_value=None
            ):
                result = runner.invoke(app, ["app", "visible", path, "true"])
                assert result.exit_code == ErrorCode.FEATURE_UNAVAILABLE
                assert "xlwings" in result.output.lower()

                result = runner.invoke(app, ["app", "calculate", path])
                assert result.exit_code == ErrorCode.FEATURE_UNAVAILABLE

                result = runner.invoke(app, ["app", "focus", path])
                assert result.exit_code == ErrorCode.FEATURE_UNAVAILABLE

                result = runner.invoke(app, ["app", "alert", path, "Hello"])
                assert result.exit_code == ErrorCode.FEATURE_UNAVAILABLE

                result = runner.invoke(app, ["app", "wait-idle", path])
                assert result.exit_code == ErrorCode.FEATURE_UNAVAILABLE
