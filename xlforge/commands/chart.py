"""Chart operations CLI commands."""

from __future__ import annotations

from importlib.util import find_spec
from pathlib import Path
from typing import Annotated, Optional

import openpyxl
import typer
from openpyxl.chart import (
    AreaChart,
    BarChart,
    DoughnutChart,
    LineChart,
    PieChart,
    RadarChart,
    Reference,
    ScatterChart,
)

from xlforge.core.errors import ErrorCode, XlforgeError

chart_app = typer.Typer(help="Chart operations for Excel workbooks.")

# Supported chart types mapping to openpyxl chart classes and their type values
CHART_TYPES: dict[str, tuple[type, str]] = {
    "column": (BarChart, "col"),
    "bar": (BarChart, "bar"),
    "line": (LineChart, "line"),
    "pie": (PieChart, "pie"),
    "scatter": (ScatterChart, "scatter"),
    "area": (AreaChart, "area"),
    "radar": (RadarChart, "radar"),
    "doughnut": (DoughnutChart, "doughnut"),
}


def _is_xlwings_available() -> bool:
    """Check if xlwings is available (Excel integration possible)."""
    return find_spec("xlwings") is not None


def _get_chart_title(chart) -> str:
    """Extract the title string from an openpyxl chart.

    Args:
        chart: An openpyxl chart object.

    Returns:
        The chart title as a string, or empty string if no title.
    """
    title = chart.title
    if title is None:
        return ""
    # If title is already a string, return it
    if isinstance(title, str):
        return title
    # Otherwise, it's a Title object - extract text from nested structure
    try:
        if hasattr(title, "tx") and hasattr(title.tx, "rich"):
            # Navigate: tx.rich.p[0].r[0].t
            return title.tx.rich.p[0].r[0].t
    except AttributeError, IndexError:
        pass
    # Fallback: convert to string
    return str(title)


def _parse_range(range_str: str) -> tuple[int, int, int, int]:
    """Parse a range string like 'A1:D10' into (min_col, min_row, max_col, max_row).

    Returns:
        Tuple of (min_col, min_row, max_col, max_row) as 1-indexed integers.
    """
    from openpyxl.utils import column_index_from_string

    parts = range_str.split(":")
    if len(parts) != 2:
        raise ValueError(f"Invalid range format: {range_str}")

    start_cell, end_cell = parts

    # Parse start cell
    start_col_str = "".join(c for c in start_cell if c.isalpha())
    start_row_str = "".join(c for c in start_cell if c.isdigit())
    if not start_col_str or not start_row_str:
        raise ValueError(f"Invalid range format: {range_str}")

    min_col = column_index_from_string(start_col_str)
    min_row = int(start_row_str)

    # Parse end cell
    end_col_str = "".join(c for c in end_cell if c.isalpha())
    end_row_str = "".join(c for c in end_cell if c.isdigit())
    if not end_col_str or not end_row_str:
        raise ValueError(f"Invalid range format: {range_str}")

    max_col = column_index_from_string(end_col_str)
    max_row = int(end_row_str)

    return min_col, min_row, max_col, max_row


def _create_regular_chart(
    path: Path,
    sheet: str,
    range: str,
    chart_type: str,
    name: Optional[str],
) -> None:
    """Create a regular chart using openpyxl.

    Args:
        path: Path to the workbook file.
        sheet: Sheet name containing the data.
        range: Data range for the chart (e.g., A1:D10).
        chart_type: Type of chart to create.
        name: Optional name for the chart.
    """
    chart_type_lower = chart_type.lower()

    try:
        # Load workbook directly with openpyxl for chart operations
        wb = openpyxl.load_workbook(path)

        # Check if sheet exists
        if sheet not in wb.sheetnames:
            typer.secho(
                f"Error: Sheet not found: {sheet}",
                fg=typer.colors.RED,
                err=True,
            )
            wb.close()
            raise typer.Exit(code=int(ErrorCode.SHEET_NOT_FOUND))

        ws = wb[sheet]

        # Parse the data range
        try:
            min_col, min_row, max_col, max_row = _parse_range(range)
        except ValueError as e:
            typer.secho(
                f"Error: Invalid range format: {range}",
                fg=typer.colors.RED,
                err=True,
            )
            wb.close()
            raise typer.Exit(code=int(ErrorCode.INVALID_SYNTAX)) from e

        # Check if chart name already exists
        if name:
            for existing_chart in ws._charts:  # type: ignore[attr-defined]
                if _get_chart_title(existing_chart) == name:
                    typer.secho(
                        f"Error: Chart with name '{name}' already exists in {sheet}",
                        fg=typer.colors.RED,
                        err=True,
                    )
                    typer.secho(
                        "Use --replace to replace existing chart",
                        fg=typer.colors.YELLOW,
                        err=True,
                    )
                    wb.close()
                    raise typer.Exit(code=int(ErrorCode.CHART_EXISTS))

        # Create the chart
        chart_class, chart_style = CHART_TYPES[chart_type_lower]
        chart = chart_class()

        # Set chart type style (col, bar, line, etc.)
        if hasattr(chart, "type"):
            chart.type = chart_style

        # Create data reference - series from columns, rows from rows
        # For most charts, first row is category (x-axis), remaining rows are data series
        data = Reference(
            ws, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row
        )

        # Add data to chart - titles_from_data=True means first row contains series names
        chart.add_data(data, titles_from_data=True)

        # Set categories (x-axis labels) - typically from first column
        cats = Reference(
            ws, min_col=min_col, min_row=min_row + 1, max_col=max_col, max_row=max_row
        )
        chart.set_categories(cats)

        # Set chart title
        if name:
            chart.title = name
        else:
            chart.title = f"{chart_type_lower.capitalize()} Chart"

        # Determine where to place the chart (default to right of data range)
        anchor_cell = f"{openpyxl.utils.get_column_letter(max_col + 2)}{min_row}"
        ws.add_chart(chart, anchor_cell)  # type: ignore[call-arg]

        # Save the workbook
        wb.save(path)
        wb.close()

        chart_name = name or _get_chart_title(chart)
        typer.echo(
            f"Created chart '{chart_name}' of type '{chart_type_lower}' in {path} ({sheet})"
        )

    except XlforgeError:
        raise
    except typer.Exit:
        raise
    except Exception as e:
        typer.secho(f"Error: {e}", fg=typer.colors.RED, err=True)
        raise typer.Exit(code=1)


def _create_pivot_chart(
    path: Path,
    sheet: str,
    range: str,
    chart_type: str,
    name: Optional[str],
    pivot_name: str,
) -> None:
    """Create a PivotChart linked to a pivot table using win32com.

    When a chart is created on the same sheet as a pivot table with its
    SetSourceData pointing to the pivot's TableRange1, Excel automatically
    links it via the PivotLayout property for dynamic updates.

    Args:
        path: Path to the workbook file.
        sheet: Sheet name containing the pivot table.
        range: Not used for pivot charts (data range comes from pivot).
        chart_type: Type of chart to create.
        name: Optional name for the chart.
        pivot_name: Name of the pivot table to link to.
    """
    chart_type_lower = chart_type.lower()

    # Map chart types to Excel enum values
    # Excel XlChartType enum values
    xl_chart_types = {
        "column": 51,  # xlColumnClustered
        "bar": 57,  # xlBarClustered
        "line": 4,  # xlLine
        "pie": 5,  # xlPie
        "scatter": -4169,  # xlXYScatter
        "area": 1,  # xlArea
        "radar": -4151,  # xlRadar
        "doughnut": 20,  # xlDoughnut
    }

    if chart_type_lower not in xl_chart_types:
        typer.secho(
            f"Error: Invalid chart type for pivot chart: {chart_type}",
            fg=typer.colors.RED,
            err=True,
        )
        typer.secho(
            f"Valid types: {', '.join(xl_chart_types.keys())}",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.INVALID_CHART_TYPE))

    xl_chart_type = xl_chart_types[chart_type_lower]

    # Note: We don't check for file open in Excel here because win32com's
    # Dispatch("Excel.Application") can connect to an already-open instance

    excel = None
    try:
        import win32com.client  # type: ignore[import-untyped]

        # Use win32com directly for pivot chart creation
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb_com = excel.Workbooks.Open(str(path.absolute()))

        # Get the sheet
        try:
            ws_com = wb_com.Sheets(sheet)
        except Exception:
            typer.secho(
                f"Error: Sheet not found: {sheet}",
                fg=typer.colors.RED,
                err=True,
            )
            raise typer.Exit(code=int(ErrorCode.SHEET_NOT_FOUND))

        # Find the pivot table by name
        pivot_table = None
        for pt in ws_com.PivotTables():
            if pt.Name == pivot_name:
                pivot_table = pt
                break

        if pivot_table is None:
            typer.secho(
                f"Error: Pivot table '{pivot_name}' not found in sheet '{sheet}'",
                fg=typer.colors.RED,
                err=True,
            )
            raise typer.Exit(code=int(ErrorCode.PIVOT_NOT_FOUND))

        # Get the pivot table's data range
        pivot_range = pivot_table.TableRange1

        # Create a chart on the sheet using the pivot table's data range
        # This automatically links it to the pivot via PivotLayout
        chart_object = ws_com.ChartObjects().Add(
            Left=pivot_table.TableRange1.Left + pivot_table.TableRange1.Width + 20,
            Top=pivot_table.TableRange1.Top,
            Width=300,
            Height=200,
        )
        chart = chart_object.Chart

        # Set chart type
        chart.ChartType = xl_chart_type

        # Set source data to the pivot table's range - this links it to the pivot
        chart.SetSourceData(pivot_range)

        # Set chart title
        if name:
            chart.HasTitle = True
            chart.ChartTitle.Text = name

        # Save the workbook
        wb_com.Save()
        wb_com.Close()

        chart_label = name or f"PivotChart ({pivot_name})"
        typer.echo(
            f"Created PivotChart '{chart_label}' linked to pivot table '{pivot_name}' in {path} ({sheet})"
        )

    except typer.Exit:
        raise
    except Exception as e:
        typer.secho(f"Error: {e}", fg=typer.colors.RED, err=True)
        raise typer.Exit(code=1)
    finally:
        # Clean up COM objects
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass


@chart_app.command()
def create(
    path: Annotated[Path, typer.Argument(help="Path to the workbook file.")],
    sheet: Annotated[str, typer.Argument(help="Sheet name containing the data.")],
    range: Annotated[
        str, typer.Argument(help="Data range for the chart (e.g., A1:D10).")
    ],
    type: Annotated[
        str,
        typer.Option(
            "--type", "-t", help=f"Chart type: {', '.join(CHART_TYPES.keys())}."
        ),
    ],
    name: Annotated[
        Optional[str],
        typer.Option("--name", "-n", help="Name for the chart."),
    ] = None,
    pivot: Annotated[
        Optional[str],
        typer.Option(
            "--pivot",
            "-p",
            help="Name of the pivot table to link the chart to (creates a PivotChart).",
        ),
    ] = None,
) -> None:
    """Create a chart in a sheet.

    When --pivot is specified, creates a PivotChart linked to the specified
    pivot table. The chart will automatically update when the pivot is refreshed.
    This requires Excel via win32com.
    """
    # Check if xlwings is available (chart operations require Excel)
    if not _is_xlwings_available():
        typer.secho(
            "Error: Chart operations require Excel via xlwings engine.",
            fg=typer.colors.RED,
            err=True,
        )
        typer.secho(
            "Feature unavailable in headless mode (openpyxl only).",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.FEATURE_UNAVAILABLE))

    if not path.exists():
        typer.secho(
            f"Error: File does not exist: {path}",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.FILE_DOES_NOT_EXIST))

    # Validate chart type
    chart_type_lower = type.lower()
    if chart_type_lower not in CHART_TYPES:
        typer.secho(
            f"Error: Invalid chart type: {type}",
            fg=typer.colors.RED,
            err=True,
        )
        typer.secho(
            f"Valid types: {', '.join(CHART_TYPES.keys())}",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.INVALID_CHART_TYPE))

    # If pivot option is specified, use win32com to create a true PivotChart
    if pivot:
        _create_pivot_chart(path, sheet, range, type, name, pivot)
    else:
        _create_regular_chart(path, sheet, range, type, name)


@chart_app.command()
def delete(
    path: Annotated[Path, typer.Argument(help="Path to the workbook file.")],
    sheet: Annotated[str, typer.Argument(help="Sheet name.")],
    name: Annotated[str, typer.Argument(help="Name of the chart to delete.")],
) -> None:
    """Delete a chart from a sheet."""
    # Check if xlwings is available (chart operations require Excel)
    if not _is_xlwings_available():
        typer.secho(
            "Error: Chart operations require Excel via xlwings engine.",
            fg=typer.colors.RED,
            err=True,
        )
        typer.secho(
            "Feature unavailable in headless mode (openpyxl only).",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.FEATURE_UNAVAILABLE))

    if not path.exists():
        typer.secho(
            f"Error: File does not exist: {path}",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.FILE_DOES_NOT_EXIST))

    try:
        # Load workbook directly with openpyxl for chart operations
        wb = openpyxl.load_workbook(path)

        # Check if sheet exists
        if sheet not in wb.sheetnames:
            typer.secho(
                f"Error: Sheet not found: {sheet}",
                fg=typer.colors.RED,
                err=True,
            )
            wb.close()
            raise typer.Exit(code=int(ErrorCode.SHEET_NOT_FOUND))

        ws = wb[sheet]

        # Find the chart by name
        chart_to_delete = None
        for chart in ws._charts:  # type: ignore[attr-defined]
            if _get_chart_title(chart) == name:
                chart_to_delete = chart
                break

        if chart_to_delete is None:
            typer.secho(
                f"Error: Chart not found: {name}",
                fg=typer.colors.RED,
                err=True,
            )
            wb.close()
            raise typer.Exit(code=int(ErrorCode.CHART_NOT_FOUND))

        # Remove the chart
        ws._charts.remove(chart_to_delete)  # type: ignore[attr-defined]

        # Save the workbook
        wb.save(path)
        wb.close()

        typer.echo(f"Deleted chart '{name}' from {path} ({sheet})")

    except XlforgeError:
        raise
    except typer.Exit:
        raise
    except Exception as e:
        typer.secho(f"Error: {e}", fg=typer.colors.RED, err=True)
        raise typer.Exit(code=1)


@chart_app.command()
def list(
    path: Annotated[Path, typer.Argument(help="Path to the workbook file.")],
    sheet: Annotated[str, typer.Argument(help="Sheet name.")],
) -> None:
    """List all charts in a sheet."""
    # Check if xlwings is available (chart operations require Excel)
    if not _is_xlwings_available():
        typer.secho(
            "Error: Chart operations require Excel via xlwings engine.",
            fg=typer.colors.RED,
            err=True,
        )
        typer.secho(
            "Feature unavailable in headless mode (openpyxl only).",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.FEATURE_UNAVAILABLE))

    if not path.exists():
        typer.secho(
            f"Error: File does not exist: {path}",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.FILE_DOES_NOT_EXIST))

    try:
        # Load workbook directly with openpyxl for chart operations
        wb = openpyxl.load_workbook(path)

        # Check if sheet exists
        if sheet not in wb.sheetnames:
            typer.secho(
                f"Error: Sheet not found: {sheet}",
                fg=typer.colors.RED,
                err=True,
            )
            wb.close()
            raise typer.Exit(code=int(ErrorCode.SHEET_NOT_FOUND))

        ws = wb[sheet]

        # List all charts
        charts = ws._charts  # type: ignore[attr-defined]

        if not charts:
            typer.echo(f"No charts found in {path} ({sheet})")
            wb.close()
            return

        typer.echo(f"Charts in {path} ({sheet}):")
        typer.echo("")

        for i, chart in enumerate(charts, 1):
            chart_type = type(chart).__name__.replace("Chart", "").lower()
            anchor = getattr(chart, "anchor", "unknown")
            if hasattr(chart, "coordinates"):
                anchor = chart.coordinates

            typer.echo(f"  {i}. Name: {_get_chart_title(chart)}")
            typer.echo(f"     Type: {chart_type}")
            typer.echo(f"     Position: {anchor}")
            typer.echo("")

        wb.close()

    except XlforgeError:
        raise
    except typer.Exit:
        raise
    except Exception as e:
        typer.secho(f"Error: {e}", fg=typer.colors.RED, err=True)
        raise typer.Exit(code=1)
