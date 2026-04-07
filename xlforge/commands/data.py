"""Data operations CLI commands."""

from __future__ import annotations

from collections import Counter
from pathlib import Path
from typing import Annotated

import openpyxl  # type: ignore[import]
import typer

from xlforge.core.errors import ErrorCode

data_app = typer.Typer(help="Data operations for Excel workbooks.")


def _column_letter_to_index(column_letter: str) -> int:
    """Convert column letter to 1-based column index.

    Args:
        column_letter: Column letter (e.g., 'A', 'B', 'AA')

    Returns:
        1-based column index
    """
    result = 0
    for char in column_letter.upper():
        result = result * 26 + (ord(char) - ord("A") + 1)
    return result


@data_app.command(name="unique")
def unique(
    path: Annotated[Path, typer.Argument(help="Path to the workbook file.")],
    sheet: Annotated[str, typer.Argument(help="Sheet name.")],
    column: Annotated[str, typer.Argument(help="Column name (header) or column letter (e.g., 'A', 'B').")],
    has_header: Annotated[
        bool,
        typer.Option("--has-header", "-H", help="Treat the first row as headers."),
    ] = True,
    json_output: Annotated[
        bool, typer.Option("--json", "-j", help="Output as JSON.")
    ] = False,
    top_n: Annotated[
        int | None,
        typer.Option("--top", "-n", help="Show only top N values by frequency."),
    ] = None,
) -> None:
    """List unique values in a column with their frequency counts.

    Shows frequency distribution of values in the specified column,
    sorted by frequency in descending order.

    Example:
        xlforge data unique movements.xlsx "data export" "Type"
        # Output:
        # Piger: 4521
        # Charger: 8938
    """
    # Check if file exists
    if not path.exists():
        typer.secho(
            f"Error: File does not exist: {path}",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.FILE_DOES_NOT_EXIST))

    try:
        # Load workbook with openpyxl directly for efficient column iteration
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)  # type: ignore[arg-type]

        try:
            # Check if sheet exists
            if sheet not in wb.sheetnames:
                typer.secho(
                    f"Error: Sheet not found: {sheet}",
                    fg=typer.colors.RED,
                    err=True,
                )
                raise typer.Exit(code=int(ErrorCode.SHEET_NOT_FOUND))

            ws = wb[sheet]

            # Determine max row with data
            max_row = ws.max_row or 0
            if max_row == 0:
                typer.echo("Sheet is empty.")
                return

            # Determine if column is a letter (e.g., 'A', 'AA') or a header name
            is_column_letter = (
                column.isalpha() and column.isupper()
            ) or (
                len(column) >= 2
                and column[0].isupper()
                and all(c.isupper() for c in column)
            )

            col_index: int
            if is_column_letter:
                col_index = _column_letter_to_index(column)
            else:
                # Treat as header name - find column index
                if has_header and max_row >= 1:
                    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
                    headers = list(header_row) if header_row else []
                    try:
                        col_index = headers.index(column) + 1  # 1-based
                    except ValueError:
                        typer.secho(
                            f"Error: Column header '{column}' not found in sheet.",
                            fg=typer.colors.RED,
                            err=True,
                        )
                        raise typer.Exit(code=int(ErrorCode.COLUMN_NOT_FOUND))
                elif not has_header:
                    # If no header, treat column as 1-based index
                    try:
                        col_index = int(column)
                    except ValueError:
                        typer.secho(
                            f"Error: Invalid column specification: {column}",
                            fg=typer.colors.RED,
                            err=True,
                        )
                        raise typer.Exit(code=1)
                else:
                    typer.secho(
                        "Error: Sheet has no header row and --has-header was not specified.",
                        fg=typer.colors.RED,
                        err=True,
                    )
                    raise typer.Exit(code=1)

            # Collect all values from the column
            values: list[str] = []
            start_row = 2 if has_header else 1
            for row in ws.iter_rows(
                min_row=start_row,
                max_row=max_row,
                min_col=col_index,
                max_col=col_index,
                values_only=True,
            ):
                if row and row[0] is not None:
                    value = str(row[0])
                    values.append(value)

            # Count frequencies
            freq = Counter(values)

            # Sort by frequency descending
            sorted_items = sorted(freq.items(), key=lambda x: (-x[1], x[0]))

            # Apply top_n limit if specified
            if top_n is not None and top_n > 0:
                sorted_items = sorted_items[:top_n]

            # Calculate total for percentage display
            total = sum(freq.values())
            distinct_count = len(freq)

            # Output results
            if json_output:
                import json

                data = {
                    "path": str(path),
                    "sheet": sheet,
                    "column": column,
                    "distinct_count": distinct_count,
                    "total_count": total,
                    "frequencies": [
                        {"value": value, "count": count, "percentage": round(count / total * 100, 2)}
                        for value, count in sorted_items
                    ],
                }
                typer.echo(json.dumps(data, indent=2))
            else:
                typer.echo(f"Column: {column}")
                typer.echo(f"Distinct values: {distinct_count}")
                typer.echo(f"Total non-empty cells: {total}")
                typer.echo("")
                typer.echo("Frequency distribution:")
                for value, count in sorted_items:
                    percentage = count / total * 100
                    typer.echo(f"  {value}: {count} ({percentage:.1f}%)")

        finally:
            wb.close()

    except XlforgeError:
        raise
    except typer.Exit:
        raise
    except Exception as e:
        typer.secho(f"Error: {e}", fg=typer.colors.RED, err=True)
        raise typer.Exit(code=int(ErrorCode.GENERAL_ERROR))


@data_app.command()
def stats(
    path: Annotated[Path, typer.Argument(help="Path to the workbook file.")],
    sheet: Annotated[str, typer.Argument(help="Sheet name.")],
    column_name: Annotated[str, typer.Argument(help="Column name (header) to analyze.")],
) -> None:
    """Generate basic statistics for a numeric column."""
    # Check if file exists
    if not path.exists():
        typer.secho(
            f"Error: File does not exist: {path}",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.FILE_DOES_NOT_EXIST))

    try:
        wb = openpyxl.load_workbook(path, data_only=True)  # type: ignore[no-redef]

        # Check if sheet exists
        if sheet not in wb.sheetnames:
            wb.close()
            typer.secho(
                f"Error: Sheet not found: {sheet}",
                fg=typer.colors.RED,
                err=True,
            )
            raise typer.Exit(code=int(ErrorCode.SHEET_NOT_FOUND))

        ws = wb[sheet]

        # Find the column index by looking at the first row (header row)
        if ws.max_row < 1:
            wb.close()
            typer.echo("Column: " + column_name)
            typer.echo("Count: 0")
            typer.echo("Empty: 0")
            typer.echo("Sum: 0")
            typer.echo("Avg: 0")
            typer.echo("Min: 0")
            typer.echo("Max: 0")
            return

        # Find header row and column index
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        col_index: int | None = None
        for idx, header in enumerate(header_row):
            if header is not None and str(header) == column_name:
                col_index = idx + 1  # Convert to 1-based
                break

        if col_index is None:
            wb.close()
            typer.secho(
                f"Error: Column '{column_name}' not found in sheet '{sheet}'",
                fg=typer.colors.RED,
                err=True,
            )
            raise typer.Exit(code=int(ErrorCode.COLUMN_NOT_FOUND))

        # Read all values in that column (excluding header)
        values: list[float] = []
        empty_count = 0
        total_count = 0

        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_index)
            total_count += 1

            if cell.value is None:
                empty_count += 1
                continue

            # Try to convert to numeric
            try:
                numeric_value = float(cell.value)
                values.append(numeric_value)
            except (ValueError, TypeError):
                # Skip non-numeric values for stats calculations
                continue

        # Calculate statistics
        count = len(values)
        if count > 0:
            total_sum = sum(values)
            avg = total_sum / count
            min_val = min(values)
            max_val = max(values)
        else:
            total_sum = 0
            avg = 0
            min_val = 0
            max_val = 0

        wb.close()

        # Output formatted results
        typer.echo(f"Column: {column_name}")
        typer.echo(f"Count: {count}")
        typer.echo(f"Empty: {empty_count}")
        typer.echo(f"Sum: {total_sum}")
        typer.echo(f"Avg: {avg:.1f}")
        typer.echo(f"Min: {min_val}")
        typer.echo(f"Max: {max_val}")

    except typer.Exit:
        raise
    except Exception as e:
        typer.secho(f"Error: {e}", fg=typer.colors.RED, err=True)
        raise typer.Exit(code=1)
