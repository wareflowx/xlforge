"""File operations CLI commands."""

from __future__ import annotations

from pathlib import Path
from typing import Annotated, Optional

import openpyxl
import typer

from xlforge.core.entities.workbook import Workbook
from xlforge.core.engines.selector import EngineSelector
from xlforge.core.errors import ErrorCode, XlforgeError

file_app = typer.Typer(help="File operations for Excel workbooks.")


@file_app.command()
def open(
    path: Annotated[Path, typer.Argument(help="Path to the workbook file.")],
    engine: Annotated[
        Optional[str],
        typer.Option("--engine", "-e", help="Engine to use: xlwings or openpyxl."),
    ] = None,
    read_only: Annotated[
        bool, typer.Option("--read-only", "-r", help="Open in read-only mode.")
    ] = False,
    auto_create: Annotated[
        bool,
        typer.Option(
            "--auto-create",
            "-c",
            help="Create a new workbook if the file doesn't exist.",
        ),
    ] = False,
) -> None:
    """Open a workbook file."""
    # Check if file exists
    if not path.exists():
        if auto_create:
            # Create a new workbook with default sheet
            wb = openpyxl.Workbook()
            wb.save(path)
            typer.echo(f"Created new workbook: {path}")
        else:
            typer.secho(
                f"Error: File does not exist: {path}",
                fg=typer.colors.RED,
                err=True,
            )
            raise typer.Exit(code=int(ErrorCode.FILE_DOES_NOT_EXIST))

    try:
        if engine is not None:
            engine_obj = EngineSelector.for_engine_name(engine)
        else:
            engine_obj = EngineSelector.for_path(path)

        workbook = Workbook(path=path, engine=engine_obj, read_only=read_only)
        workbook.open()

        engine_name = engine_obj.__class__.__name__.replace("Engine", "").lower()
        typer.echo(f"Opened: {path}")
        typer.echo(f"Engine: {engine_name}")
    except XlforgeError:
        raise
    except Exception as e:
        typer.secho(f"Error: {e}", fg=typer.colors.RED, err=True)
        raise typer.Exit(code=1)


@file_app.command()
def save(
    path: Annotated[Path, typer.Argument(help="Path to the workbook file.")],
    output: Annotated[
        Optional[Path],
        typer.Option("--output", "-o", help="Output path for saving."),
    ] = None,
) -> None:
    """Save a workbook file."""
    if not path.exists():
        typer.secho(
            f"Error: File does not exist: {path}",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.FILE_DOES_NOT_EXIST))

    try:
        engine = EngineSelector.for_path(path)
        workbook = Workbook(path=path, engine=engine, read_only=False)
        workbook.open()

        try:
            if output is not None:
                # For now, just save to the original path
                # The output path would require more complex logic for copy/save-as
                typer.echo(f"Output path specified: {output}")
                typer.echo("Saving to original path...")

            workbook.save()
            typer.echo(f"Saved: {path}")
        finally:
            workbook.close()
    except XlforgeError:
        raise
    except Exception as e:
        typer.secho(f"Error: {e}", fg=typer.colors.RED, err=True)
        raise typer.Exit(code=1)


@file_app.command()
def info(
    path: Annotated[Path, typer.Argument(help="Path to the workbook file.")],
    json_output: Annotated[
        bool, typer.Option("--json", "-j", help="Output as JSON.")
    ] = False,
) -> None:
    """Show information about a workbook file."""
    if not path.exists():
        typer.secho(
            f"Error: File does not exist: {path}",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.FILE_DOES_NOT_EXIST))

    try:
        engine = EngineSelector.for_path(path)
        workbook = Workbook(path=path, engine=engine, read_only=True)
        workbook.open()

        try:
            sheet_names = workbook.sheets()
            engine_name = engine.__class__.__name__.replace("Engine", "").lower()

            if json_output:
                import json

                data = {
                    "path": str(path),
                    "engine": engine_name,
                    "sheets": [s.name for s in sheet_names],
                }
                typer.echo(json.dumps(data, indent=2))
            else:
                typer.echo(f"Path: {path}")
                typer.echo(f"Engine: {engine_name}")
                typer.echo(f"Sheets: {', '.join(s.name for s in sheet_names)}")
        finally:
            workbook.close()
    except XlforgeError:
        raise
    except Exception as e:
        typer.secho(f"Error: {e}", fg=typer.colors.RED, err=True)
        raise typer.Exit(code=1)


@file_app.command()
def close(
    path: Annotated[Path, typer.Argument(help="Path to the workbook file.")],
) -> None:
    """Close a workbook file."""
    if not path.exists():
        typer.secho(
            f"Error: File does not exist: {path}",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.FILE_DOES_NOT_EXIST))

    try:
        engine = EngineSelector.for_path(path)
        workbook = Workbook(path=path, engine=engine, read_only=True)
        workbook.open()

        try:
            workbook.close()
            typer.echo(f"Closed: {path}")
        except Exception as e:
            typer.secho(f"Error: {e}", fg=typer.colors.RED, err=True)
            raise typer.Exit(code=1)
    except XlforgeError:
        raise
    except Exception as e:
        typer.secho(f"Error: {e}", fg=typer.colors.RED, err=True)
        raise typer.Exit(code=1)


@file_app.command()
def check(
    path: Annotated[Path, typer.Argument(help="Path to the workbook file.")],
    json_output: Annotated[
        bool, typer.Option("--json", "-j", help="Output as JSON.")
    ] = False,
) -> None:
    """Analyze file health and check for issues."""
    import json

    # Check if file exists
    if not path.exists():
        error_msg = f"Error: File does not exist: {path}"
        if json_output:
            typer.echo(json.dumps({
                "path": str(path),
                "exists": False,
                "valid_xlsx": False,
                "healthy": False,
                "errors": [error_msg],
            }, indent=2))
        else:
            typer.secho(error_msg, fg=typer.colors.RED, err=True)
        raise typer.Exit(code=int(ErrorCode.FILE_DOES_NOT_EXIST))

    errors = []
    is_valid_xlsx = False
    is_healthy = False

    # Check if file is valid xlsx
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        is_valid_xlsx = True
        wb.close()
    except Exception as e:
        errors.append(f"Invalid xlsx: {e}")

    # Overall health check
    if path.exists() and is_valid_xlsx:
        is_healthy = True

    if json_output:
        output = {
            "path": str(path),
            "exists": True,
            "valid_xlsx": is_valid_xlsx,
            "healthy": is_healthy,
            "errors": errors if errors else None,
        }
        typer.echo(json.dumps(output, indent=2))
    else:
        typer.echo(f"Path: {path}")
        typer.echo(f"Exists: {True}")
        typer.echo(f"Valid xlsx: {is_valid_xlsx}")
        typer.echo(f"Healthy: {is_healthy}")
        if errors:
            for error in errors:
                typer.secho(f"  Error: {error}", fg=typer.colors.RED)


@file_app.command()
def recover(
    path: Annotated[Path, typer.Argument(help="Path to the workbook file.")],
) -> None:
    """Attempt to recover a file from corruption by re-opening and saving."""
    if not path.exists():
        typer.secho(
            f"Error: File does not exist: {path}",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.FILE_DOES_NOT_EXIST))

    try:
        # Open and save to recover
        wb = openpyxl.load_workbook(path)
        wb.save(path)
        wb.close()
        typer.echo(f"Recovered: {path}")
    except Exception as e:
        typer.secho(f"Error: Recovery failed: {e}", fg=typer.colors.RED, err=True)
        raise typer.Exit(code=int(ErrorCode.RECOVERY_FAILED))
