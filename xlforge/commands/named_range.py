"""Named range operations CLI commands."""

# Note: This command uses openpyxl directly and bypasses the Engine abstraction.
# It works with OpenpyxlEngine but not with XlwingsEngine.

from __future__ import annotations

from pathlib import Path
from typing import Annotated

import openpyxl
import typer

from xlforge.core.errors import ErrorCode

named_range_app = typer.Typer(help="Named range operations for Excel workbooks.")


@named_range_app.command()
def create(
    path: Annotated[Path, typer.Argument(help="Path to the workbook file.")],
    name: Annotated[str, typer.Argument(help="Name for the named range.")],
    sheet: Annotated[str, typer.Argument(help="Sheet name containing the range.")],
    range_ref: Annotated[str, typer.Argument(help="Range reference (e.g., A1:C10).")],
) -> None:
    """Create a named range in the workbook."""
    # Check if file exists
    if not path.exists():
        typer.secho(
            f"Error: File does not exist: {path}",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.FILE_DOES_NOT_EXIST))

    try:
        wb = openpyxl.load_workbook(path)

        # Check if sheet exists
        if sheet not in wb.sheetnames:
            typer.secho(
                f"Error: Sheet not found: {sheet}",
                fg=typer.colors.RED,
                err=True,
            )
            wb.close()
            raise typer.Exit(code=int(ErrorCode.SHEET_NOT_FOUND))

        # Check if named range already exists
        if name in wb.defined_names:
            typer.secho(
                f"Error: Named range already exists: {name}",
                fg=typer.colors.RED,
                err=True,
            )
            wb.close()
            raise typer.Exit(code=int(ErrorCode.TABLE_ALREADY_EXISTS))

        # Create the named range (workbook-scoped)
        # The attr_text should be "SheetName!Range" format
        defined_name = openpyxl.workbook.defined_name.DefinedName(
            name, attr_text=f"{sheet}!{range_ref}"
        )
        wb.defined_names.add(defined_name)

        wb.save(path)
        wb.close()

        typer.echo(f"Created named range '{name}' = {sheet}!{range_ref}")

    except typer.Exit:
        raise
    except Exception as e:
        typer.secho(f"Error: {e}", fg=typer.colors.RED, err=True)
        raise typer.Exit(code=1)


@named_range_app.command()
def delete(
    path: Annotated[Path, typer.Argument(help="Path to the workbook file.")],
    name: Annotated[str, typer.Argument(help="Name of the named range to delete.")],
) -> None:
    """Delete a named range from the workbook."""
    # Check if file exists
    if not path.exists():
        typer.secho(
            f"Error: File does not exist: {path}",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.FILE_DOES_NOT_EXIST))

    try:
        wb = openpyxl.load_workbook(path)

        # Check if named range exists
        if name not in wb.defined_names:
            typer.secho(
                f"Error: Named range not found: {name}",
                fg=typer.colors.RED,
                err=True,
            )
            wb.close()
            raise typer.Exit(code=int(ErrorCode.TABLE_NOT_FOUND))

        # Delete the named range
        del wb.defined_names[name]

        wb.save(path)
        wb.close()

        typer.echo(f"Deleted named range '{name}'")

    except typer.Exit:
        raise
    except Exception as e:
        typer.secho(f"Error: {e}", fg=typer.colors.RED, err=True)
        raise typer.Exit(code=1)


@named_range_app.command("list")
def list_ranges(
    path: Annotated[Path, typer.Argument(help="Path to the workbook file.")],
) -> None:
    """List all named ranges in the workbook."""
    # Check if file exists
    if not path.exists():
        typer.secho(
            f"Error: File does not exist: {path}",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.FILE_DOES_NOT_EXIST))

    try:
        wb = openpyxl.load_workbook(path)

        # List all defined names
        named_ranges = list(wb.defined_names.values())

        if not named_ranges:
            typer.echo("No named ranges found in workbook.")
        else:
            for nr in named_ranges:
                typer.echo(f"{nr.name} = {nr.attr_text}")

        wb.close()

    except typer.Exit:
        raise
    except Exception as e:
        typer.secho(f"Error: {e}", fg=typer.colors.RED, err=True)
        raise typer.Exit(code=1)


@named_range_app.command()
def get(
    path: Annotated[Path, typer.Argument(help="Path to the workbook file.")],
    name: Annotated[str, typer.Argument(help="Name of the named range.")],
) -> None:
    """Get the range reference for a named range."""
    # Check if file exists
    if not path.exists():
        typer.secho(
            f"Error: File does not exist: {path}",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.FILE_DOES_NOT_EXIST))

    try:
        wb = openpyxl.load_workbook(path)

        # Check if named range exists
        if name not in wb.defined_names:
            typer.secho(
                f"Error: Named range not found: {name}",
                fg=typer.colors.RED,
                err=True,
            )
            wb.close()
            raise typer.Exit(code=int(ErrorCode.TABLE_NOT_FOUND))

        # Get the named range
        defined_name = wb.defined_names[name]
        typer.echo(f"{name} = {defined_name.attr_text}")

        wb.close()

    except typer.Exit:
        raise
    except Exception as e:
        typer.secho(f"Error: {e}", fg=typer.colors.RED, err=True)
        raise typer.Exit(code=1)
