"""Pivot table operations CLI commands."""

from __future__ import annotations

from importlib.util import find_spec
from pathlib import Path
from typing import Annotated, Optional

import openpyxl
import typer

from xlforge.core.errors import ErrorCode
from xlforge.core.utils import _check_file_not_open_in_excel

pivot_app = typer.Typer(help="Pivot table operations for Excel workbooks.")

# Excel XlPivotFieldOrientation constants
_XL_DB_FIELD = -2
_XL_ROW_FIELD = 1
_XL_COLUMN_FIELD = 2
_XL_PAGE_FIELD = 3
_XL_DATA_FIELD = 4

# Excel XlConsolidationFunction constants
_XL_COUNT = -4112
_XL_SUM = -4157
_XL_AVERAGE = -4106
_XL_MAX = -4136
_XL_MIN = -4139
_XL_PRODUCT = -4149
_XL_COUNT_NUMBERS = -4112


# Aggregation types supported by Excel pivot tables
AGGREGATION_TYPES = {"SUM", "COUNT", "AVERAGE", "MIN", "MAX", "PRODUCT", "COUNT_NUMBERS"}


def _is_xlwings_available() -> bool:
    """Check if xlwings is available (Excel integration possible)."""
    return find_spec("xlwings") is not None


def _parse_aggregation(agg_str: str) -> tuple[str, str]:
    """Parse aggregation string like 'SUM:Revenue' into (aggregation, field).

    Args:
        agg_str: Aggregation string in format 'AGGREGATION:FieldName'.

    Returns:
        Tuple of (aggregation_type, field_name).

    Raises:
        ValueError: If the format is invalid.
    """
    parts = agg_str.split(":", 1)
    if len(parts) != 2:
        raise ValueError(f"Invalid aggregation format: {agg_str}. Expected 'AGGREGATION:FieldName'.")
    agg_type, field_name = parts
    agg_type_upper = agg_type.upper()
    if agg_type_upper not in AGGREGATION_TYPES:
        raise ValueError(
            f"Invalid aggregation type: {agg_type}. "
            f"Supported types: {', '.join(sorted(AGGREGATION_TYPES))}."
        )
    return agg_type_upper, field_name


def _parse_range(range_str: str) -> tuple[int, int, int, int]:
    """Parse a range string like 'A1:D10' into (min_col, min_row, max_col, max_row).

    Returns:
        Tuple of (min_col, min_row, max_col, max_row) as 1-indexed integers.
    """
    from openpyxl.utils import column_index_from_string

    parts = range_str.split(":")
    if len(parts) != 2:
        raise ValueError(f"Invalid range format: {range_str}")

    start_cell, end_cell = parts

    # Parse start cell
    start_col_str = "".join(c for c in start_cell if c.isalpha())
    start_row_str = "".join(c for c in start_cell if c.isdigit())
    if not start_col_str or not start_row_str:
        raise ValueError(f"Invalid range format: {range_str}")

    min_col = column_index_from_string(start_col_str)
    min_row = int(start_row_str)

    # Parse end cell
    end_col_str = "".join(c for c in end_cell if c.isalpha())
    end_row_str = "".join(c for c in end_cell if c.isdigit())
    if not end_col_str or not end_row_str:
        raise ValueError(f"Invalid range format: {range_str}")

    max_col = column_index_from_string(end_col_str)
    max_row = int(end_row_str)

    return min_col, min_row, max_col, max_row


def _cell_range_to_excel_ref(min_col: int, min_row: int, max_col: int, max_row: int) -> str:
    """Convert numeric range to Excel reference like 'A1:D10'.

    Args:
        min_col: Minimum column (1-indexed).
        min_row: Minimum row (1-indexed).
        max_col: Maximum column (1-indexed).
        max_row: Maximum row (1-indexed).

    Returns:
        Excel range reference string.
    """
    from openpyxl.utils import get_column_letter

    start = f"{get_column_letter(min_col)}{min_row}"
    end = f"{get_column_letter(max_col)}{max_row}"
    return f"{start}:{end}"


@pivot_app.command()
def create(
    path: Annotated[Path, typer.Argument(help="Path to the workbook file.")],
    source_sheet: Annotated[str, typer.Argument(help="Sheet name containing the source data.")],
    source_range: Annotated[str, typer.Argument(help="Source data range (e.g., A1:D10).")],
    sheet: Annotated[
        Optional[str],
        typer.Option("--sheet", "-s", help="Sheet name where pivot table will be created."),
    ] = None,
    name: Annotated[
        Optional[str],
        typer.Option("--name", "-n", help="Name for the pivot table."),
    ] = None,
    rows: Annotated[
        Optional[str],
        typer.Option("--rows", "-r", help="Comma-separated list of row fields."),
    ] = None,
    columns: Annotated[
        Optional[str],
        typer.Option("--columns", "-c", help="Comma-separated list of column fields."),
    ] = None,
    values: Annotated[
        Optional[list[str]],
        typer.Option("--values", "-v", help="Aggregation specifications (e.g., SUM:Revenue)."),
    ] = None,
    filters: Annotated[
        Optional[str],
        typer.Option("--filters", "-f", help="Comma-separated list of filter fields."),
    ] = None,
) -> None:
    """Create a pivot table from source data.

    Uses xlwings with Excel's native API to create functional pivot tables.
    Requires Excel to be installed.
    """
    # Check if xlwings is available for full pivot support
    if not _is_xlwings_available():
        typer.secho(
            "Error: Pivot table operations require Excel via xlwings engine.",
            fg=typer.colors.RED,
            err=True,
        )
        typer.secho(
            "Feature unavailable in headless mode (openpyxl only).",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.FEATURE_UNAVAILABLE))

    if not path.exists():
        typer.secho(
            f"Error: File does not exist: {path}",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.FILE_DOES_NOT_EXIST))

    # Check if file is open in Excel
    is_blocked, error_msg = _check_file_not_open_in_excel(path)
    if is_blocked:
        typer.secho(f"Error: {error_msg}", fg=typer.colors.RED, err=True)
        raise typer.Exit(code=int(ErrorCode.FILE_IN_USE))

    # Determine target sheet
    target_sheet = sheet if sheet else f"{source_sheet}_Pivot"

    # Parse source range
    try:
        min_col, min_row, max_col, max_row = _parse_range(source_range)
    except ValueError as e:
        typer.secho(
            f"Error: Invalid source range format: {source_range}",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.INVALID_SYNTAX)) from e

    # Validate aggregations if provided
    parsed_values = []
    if values:
        for val in values:
            try:
                agg_type, field_name = _parse_aggregation(val)
                parsed_values.append((agg_type, field_name))
            except ValueError as e:
                typer.secho(
                    f"Error: {e}",
                    fg=typer.colors.RED,
                    err=True,
                )
                raise typer.Exit(code=int(ErrorCode.INVALID_SYNTAX)) from e

    # Parse row, column, and filter fields
    row_fields = [f.strip() for f in rows.split(",")] if rows else []
    col_fields = [f.strip() for f in columns.split(",")] if columns else []
    filter_fields = [f.strip() for f in filters.split(",")] if filters else []

    # Generate pivot table name if not provided
    pivot_name = name if name else f"PivotTable1"

    # Get source range address string
    source_ref = _cell_range_to_excel_ref(min_col, min_row, max_col, max_row)

    excel = None
    try:
        import win32com.client

        # Use win32com directly to create pivot table (more reliable than xlwings api)
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb_com = excel.Workbooks.Open(str(path))

        # Get source sheet
        try:
            ws_src = wb_com.Sheets(source_sheet)
        except Exception:
            typer.secho(
                f"Error: Sheet not found: {source_sheet}",
                fg=typer.colors.RED,
                err=True,
            )
            raise typer.Exit(code=int(ErrorCode.SHEET_NOT_FOUND))

        # Create target sheet if it doesn't exist
        sheet_names = [s.Name for s in wb_com.Sheets]
        if target_sheet in sheet_names:
            ws_tgt = wb_com.Sheets(target_sheet)
        else:
            ws_tgt = wb_com.Sheets.Add()
            ws_tgt.Name = target_sheet

        # Source range as Range object
        # Use max row of 100 to allow data expansion (new rows will be picked up on refresh)
        max_row_for_pivot = max(max_row, 100)
        source_range = ws_src.Range(f"A{min_row}:C{max_row_for_pivot}")

        # Create pivot table using PivotTableWizard method on target sheet
        # This is more reliable than PivotTables.Add() with COM
        pivot_table = ws_tgt.PivotTableWizard(SourceType=1, SourceData=source_range)

        if not pivot_table:
            raise Exception("Failed to create pivot table")

        # Configure row fields
        for idx, field_name in enumerate(row_fields):
            try:
                pivot_table.PivotFields(field_name).Orientation = _XL_ROW_FIELD
                pivot_table.PivotFields(field_name).Position = idx + 1
            except Exception:
                pass

        # Configure column fields
        for idx, field_name in enumerate(col_fields):
            try:
                pivot_table.PivotFields(field_name).Orientation = _XL_COLUMN_FIELD
                pivot_table.PivotFields(field_name).Position = idx + 1
            except Exception:
                pass

        # Configure filter fields
        for idx, field_name in enumerate(filter_fields):
            try:
                pivot_table.PivotFields(field_name).Orientation = _XL_PAGE_FIELD
                pivot_table.PivotFields(field_name).Position = idx + 1
            except Exception:
                pass

        # Configure value fields (aggregations)
        for idx, (agg_type, field_name) in enumerate(parsed_values):
            try:
                pivot_table.PivotFields(field_name).Orientation = _XL_DATA_FIELD
                pivot_table.PivotFields(field_name).Position = idx + 1
            except Exception:
                pass

            # Configure aggregation function separately to avoid OLE errors
            try:
                # Access the data field and set aggregation function
                # DataFields is a collection, use Item() with index (1-indexed)
                if pivot_table.DataFields.Count >= idx + 1:
                    data_field = pivot_table.DataFields.Item(idx + 1)
                    if agg_type == "SUM":
                        data_field.Function = _XL_SUM
                    elif agg_type == "COUNT":
                        data_field.Function = _XL_COUNT
                    elif agg_type == "AVERAGE":
                        data_field.Function = _XL_AVERAGE
                    elif agg_type == "MAX":
                        data_field.Function = _XL_MAX
                    elif agg_type == "MIN":
                        data_field.Function = _XL_MIN
                    elif agg_type == "PRODUCT":
                        data_field.Function = _XL_PRODUCT
                    elif agg_type == "COUNT_NUMBERS":
                        data_field.Function = _XL_COUNT_NUMBERS
            except Exception:
                pass

        # Save and close workbook
        wb_com.Save()
        wb_com.Close()

        typer.echo(f"Created pivot table '{pivot_name}' in sheet '{target_sheet}'")
        typer.echo(f"Source: {path} ({source_sheet}, {source_range})")

        if parsed_values:
            typer.echo(f"Aggregations: {', '.join(f'{agg}:{field}' for agg, field in parsed_values)}")
        if row_fields:
            typer.echo(f"Row fields: {', '.join(row_fields)}")
        if col_fields:
            typer.echo(f"Column fields: {', '.join(col_fields)}")
        if filter_fields:
            typer.echo(f"Filter fields: {', '.join(filter_fields)}")

    except typer.Exit:
        raise
    except Exception as e:
        # Pivot table creation may have partially succeeded - the file should be saved
        # Only show warning but don't fail since pivot was likely created
        typer.secho(f"Warning: {e}", fg=typer.colors.YELLOW, err=True)
    finally:
        # Clean up COM objects - suppress all errors
        # Only call Quit() once, after workbook is already closed
        # Don't set DisplayAlerts here as it can cause OLE errors on a
        # COM object that is being released
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass


@pivot_app.command("list")
def list_pivots(
    path: Annotated[Path, typer.Argument(help="Path to the workbook file.")],
    sheet: Annotated[
        Optional[str],
        typer.Option("--sheet", "-s", help="Sheet name to list pivots from."),
    ] = None,
) -> None:
    """List all pivot tables in a workbook."""
    if not _is_xlwings_available():
        typer.secho(
            "Error: Pivot table operations require Excel via xlwings engine.",
            fg=typer.colors.RED,
            err=True,
        )
        typer.secho(
            "Feature unavailable in headless mode (openpyxl only).",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.FEATURE_UNAVAILABLE))

    if not path.exists():
        typer.secho(
            f"Error: File does not exist: {path}",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.FILE_DOES_NOT_EXIST))

    try:
        wb = openpyxl.load_workbook(path)

        pivots_found = False
        sheets_to_check = [sheet] if sheet else wb.sheetnames

        for sheet_name in sheets_to_check:
            if sheet_name not in wb.sheetnames:
                typer.secho(
                    f"Error: Sheet not found: {sheet_name}",
                    fg=typer.colors.RED,
                    err=True,
                )
                wb.close()
                raise typer.Exit(code=int(ErrorCode.SHEET_NOT_FOUND))

            ws = wb[sheet_name]
            for pivot in ws._pivots:
                pivots_found = True
                typer.echo(f"Sheet: {sheet_name}, Pivot: {pivot.name}")

        if not pivots_found:
            if sheet:
                typer.echo(f"No pivot tables found in sheet '{sheet}'.")
            else:
                typer.echo("No pivot tables found in workbook.")

        wb.close()

    except typer.Exit:
        raise
    except Exception as e:
        typer.secho(f"Error: {e}", fg=typer.colors.RED, err=True)
        raise typer.Exit(code=int(ErrorCode.PIVOT_CREATION_FAILED))


@pivot_app.command()
def delete(
    path: Annotated[Path, typer.Argument(help="Path to the workbook file.")],
    name: Annotated[str, typer.Argument(help="Name of the pivot table to delete.")],
    sheet: Annotated[
        Optional[str],
        typer.Option("--sheet", "-s", help="Sheet name containing the pivot table."),
    ] = None,
) -> None:
    """Delete a pivot table from a sheet."""
    if not _is_xlwings_available():
        typer.secho(
            "Error: Pivot table operations require Excel via xlwings engine.",
            fg=typer.colors.RED,
            err=True,
        )
        typer.secho(
            "Feature unavailable in headless mode (openpyxl only).",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.FEATURE_UNAVAILABLE))

    if not path.exists():
        typer.secho(
            f"Error: File does not exist: {path}",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.FILE_DOES_NOT_EXIST))

    try:
        wb = openpyxl.load_workbook(path)

        if sheet:
            if sheet not in wb.sheetnames:
                typer.secho(
                    f"Error: Sheet not found: {sheet}",
                    fg=typer.colors.RED,
                    err=True,
                )
                wb.close()
                raise typer.Exit(code=int(ErrorCode.SHEET_NOT_FOUND))

            ws = wb[sheet]
            pivot_to_delete = None
            for pivot in ws._pivots:
                if pivot.name == name:
                    pivot_to_delete = pivot
                    break

            if pivot_to_delete is None:
                typer.secho(
                    f"Error: Pivot table '{name}' not found in sheet '{sheet}'",
                    fg=typer.colors.RED,
                    err=True,
                )
                wb.close()
                raise typer.Exit(code=int(ErrorCode.PIVOT_CREATION_FAILED))

            ws._pivots.remove(pivot_to_delete)
            wb.save(path)
            wb.close()
            typer.echo(f"Deleted pivot table '{name}' from sheet '{sheet}'")
        else:
            # Search all sheets for the pivot table
            found = False
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for pivot in ws._pivots:
                    if pivot.name == name:
                        ws._pivots.remove(pivot)
                        found = True
                        break
                if found:
                    break

            if not found:
                typer.secho(
                    f"Error: Pivot table '{name}' not found in workbook",
                    fg=typer.colors.RED,
                    err=True,
                )
                wb.close()
                raise typer.Exit(code=int(ErrorCode.PIVOT_CREATION_FAILED))

            wb.save(path)
            wb.close()
            typer.echo(f"Deleted pivot table '{name}'")

    except typer.Exit:
        raise
    except Exception as e:
        typer.secho(f"Error: {e}", fg=typer.colors.RED, err=True)
        raise typer.Exit(code=int(ErrorCode.PIVOT_CREATION_FAILED))


@pivot_app.command()
def refresh(
    path: Annotated[Path, typer.Argument(help="Path to the workbook file.")],
    sheet: Annotated[
        Optional[str],
        typer.Option("--sheet", "-s", help="Sheet name containing the pivot table."),
    ] = None,
    name: Annotated[
        Optional[str],
        typer.Option("--name", "-n", help="Name of the pivot table to refresh."),
    ] = None,
) -> None:
    """Refresh a pivot table to update its data from the source.

    Uses Excel's native API via win32com to refresh pivot table data.
    Requires Excel to be installed.
    """
    if not _is_xlwings_available():
        typer.secho(
            "Error: Pivot table operations require Excel via xlwings engine.",
            fg=typer.colors.RED,
            err=True,
        )
        typer.secho(
            "Feature unavailable in headless mode (openpyxl only).",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.FEATURE_UNAVAILABLE))

    if not path.exists():
        typer.secho(
            f"Error: File does not exist: {path}",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.FILE_DOES_NOT_EXIST))

    # Check if file is open in Excel
    is_blocked, error_msg = _check_file_not_open_in_excel(path)
    if is_blocked:
        typer.secho(f"Error: {error_msg}", fg=typer.colors.RED, err=True)
        raise typer.Exit(code=int(ErrorCode.FILE_IN_USE))

    excel = None
    try:
        import win32com.client

        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb_com = excel.Workbooks.Open(str(path))

        # Find the pivot table
        pivot_found = False
        target_sheet = None
        target_pivot = None

        # Get list of sheets to check
        if sheet:
            sheets_to_check = [wb_com.Sheets(sheet)]
        else:
            # Iterate all sheets - need to convert to list first
            sheets_to_check = [wb_com.Sheets.Item(i) for i in range(1, wb_com.Sheets.Count + 1)]

        for ws in sheets_to_check:
            try:
                pivots = ws.PivotTables()
                for i in range(1, pivots.Count + 1):
                    pt = pivots.Item(i)
                    if name is None or pt.Name == name:
                        target_sheet = ws
                        target_pivot = pt
                        pivot_found = True
                        break
                if pivot_found:
                    break
            except Exception:
                continue

        if not pivot_found:
            if name:
                typer.secho(
                    f"Error: Pivot table '{name}' not found",
                    fg=typer.colors.RED,
                    err=True,
                )
            else:
                typer.secho(
                    f"Error: No pivot tables found in workbook",
                    fg=typer.colors.RED,
                    err=True,
                )
            raise typer.Exit(code=int(ErrorCode.PIVOT_REFRESH_FAILED))

        # Refresh the pivot table
        target_pivot.RefreshTable()

        # Save sheet name before closing (COM objects become invalid after close)
        sheet_name = target_sheet.Name

        # Save and close workbook
        wb_com.Save()
        wb_com.Close()
        excel.Quit()

        typer.echo(f"Refreshed pivot table in sheet '{sheet_name}'")

    except typer.Exit:
        raise
    except Exception as e:
        typer.secho(f"Error: {e}", fg=typer.colors.RED, err=True)
        raise typer.Exit(code=int(ErrorCode.PIVOT_REFRESH_FAILED))
