"""Protection operations CLI commands."""

from __future__ import annotations

from pathlib import Path
from typing import Annotated

import openpyxl
import typer

from xlforge.core.errors import ErrorCode, XlforgeError

protection_app = typer.Typer(help="Sheet protection and freeze pane operations.")


@protection_app.command()
def protect(
    path: Annotated[Path, typer.Argument(help="Path to the workbook file.")],
    sheet_name: Annotated[str, typer.Argument(help="Name of the sheet to protect.")],
    password: Annotated[
        str | None,
        typer.Option("--password", "-p", help="Password for protection."),
    ] = None,
) -> None:
    """Protect a sheet."""
    if not path.exists():
        typer.secho(
            f"Error: File does not exist: {path}",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.FILE_DOES_NOT_EXIST))

    try:
        wb = openpyxl.load_workbook(path)

        # Check if sheet exists
        if sheet_name not in wb.sheetnames:
            typer.secho(
                f"Error: Sheet not found: {sheet_name}",
                fg=typer.colors.RED,
                err=True,
            )
            raise typer.Exit(code=int(ErrorCode.SHEET_NOT_FOUND))

        ws = wb[sheet_name]

        # Protect sheet
        ws.protection.sheet = True
        if password:
            ws.protection.password = password

        wb.save(path)
        wb.close()

        if password:
            typer.echo(f"Protected sheet '{sheet_name}' in {path} with password")
        else:
            typer.echo(f"Protected sheet '{sheet_name}' in {path}")

    except XlforgeError:
        raise
    except typer.Exit:
        raise
    except Exception as e:
        typer.secho(f"Error: {e}", fg=typer.colors.RED, err=True)
        raise typer.Exit(code=int(ErrorCode.GENERAL_ERROR))


@protection_app.command()
def unprotect(
    path: Annotated[Path, typer.Argument(help="Path to the workbook file.")],
    sheet_name: Annotated[str, typer.Argument(help="Name of the sheet to unprotect.")],
    password: Annotated[
        str | None,
        typer.Option("--password", "-p", help="Password for unprotection (if set)."),
    ] = None,
) -> None:
    """Unprotect a sheet."""
    if not path.exists():
        typer.secho(
            f"Error: File does not exist: {path}",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.FILE_DOES_NOT_EXIST))

    try:
        wb = openpyxl.load_workbook(path)

        # Check if sheet exists
        if sheet_name not in wb.sheetnames:
            typer.secho(
                f"Error: Sheet not found: {sheet_name}",
                fg=typer.colors.RED,
                err=True,
            )
            raise typer.Exit(code=int(ErrorCode.SHEET_NOT_FOUND))

        ws = wb[sheet_name]

        # Unprotect sheet
        ws.protection.sheet = False

        wb.save(path)
        wb.close()

        typer.echo(f"Unprotected sheet '{sheet_name}' in {path}")

    except XlforgeError:
        raise
    except typer.Exit:
        raise
    except Exception as e:
        typer.secho(f"Error: {e}", fg=typer.colors.RED, err=True)
        raise typer.Exit(code=int(ErrorCode.GENERAL_ERROR))


@protection_app.command()
def freeze(
    path: Annotated[Path, typer.Argument(help="Path to the workbook file.")],
    sheet_name: Annotated[str, typer.Argument(help="Name of the sheet.")],
    column: Annotated[
        str | None,
        typer.Option("--column", "-c", help="Column letter for freeze pane (e.g., B)."),
    ] = None,
    row: Annotated[
        int | None,
        typer.Option("--row", "-r", help="Row number for freeze pane (e.g., 5)."),
    ] = None,
) -> None:
    """Freeze panes in a sheet."""
    if not path.exists():
        typer.secho(
            f"Error: File does not exist: {path}",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.FILE_DOES_NOT_EXIST))

    try:
        wb = openpyxl.load_workbook(path)

        # Check if sheet exists
        if sheet_name not in wb.sheetnames:
            typer.secho(
                f"Error: Sheet not found: {sheet_name}",
                fg=typer.colors.RED,
                err=True,
            )
            raise typer.Exit(code=int(ErrorCode.SHEET_NOT_FOUND))

        ws = wb[sheet_name]

        # Determine freeze pane position
        if column is None and row is None:
            # Default to A2 (freeze first row)
            ws.freeze_panes = "A2"
            typer.echo(f"Freeze panes set to A2 in '{sheet_name}'")
        elif column is not None and row is not None:
            # Both column and row specified
            cell_ref = f"{column.upper()}{row}"
            ws.freeze_panes = cell_ref
            typer.echo(f"Freeze panes set to {cell_ref} in '{sheet_name}'")
        elif column is not None:
            # Column only - freeze at that column, row 1
            cell_ref = f"{column.upper()}1"
            ws.freeze_panes = cell_ref
            typer.echo(f"Freeze panes set to {cell_ref} in '{sheet_name}'")
        else:
            # Row only - freeze at A and that row
            cell_ref = f"A{row}"
            ws.freeze_panes = cell_ref
            typer.echo(f"Freeze panes set to {cell_ref} in '{sheet_name}'")

        wb.save(path)
        wb.close()

    except XlforgeError:
        raise
    except typer.Exit:
        raise
    except Exception as e:
        typer.secho(f"Error: {e}", fg=typer.colors.RED, err=True)
        raise typer.Exit(code=int(ErrorCode.GENERAL_ERROR))
