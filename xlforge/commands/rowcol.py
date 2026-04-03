"""Row and column operations CLI commands."""

# Note: This command uses openpyxl directly and bypasses the Engine abstraction.
# It works with OpenpyxlEngine but not with XlwingsEngine.

from __future__ import annotations

from pathlib import Path
from typing import Annotated

import openpyxl
import typer

from xlforge.core.errors import ErrorCode

row_app = typer.Typer(help="Row operations for Excel workbooks.")
col_app = typer.Typer(help="Column operations for Excel workbooks.")


@row_app.command()
def hide(
    path: Annotated[Path, typer.Argument(help="Path to the workbook file.")],
    sheet: Annotated[str, typer.Argument(help="Sheet name.")],
    row: Annotated[int, typer.Argument(help="Row number (1-based).")],
) -> None:
    """Hide a row."""
    # Check if file exists
    if not path.exists():
        typer.secho(
            f"Error: File does not exist: {path}",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.FILE_DOES_NOT_EXIST))

    try:
        wb = openpyxl.load_workbook(path)

        # Check if sheet exists
        if sheet not in wb.sheetnames:
            wb.close()
            typer.secho(
                f"Error: Sheet not found: {sheet}",
                fg=typer.colors.RED,
                err=True,
            )
            raise typer.Exit(code=int(ErrorCode.SHEET_NOT_FOUND))

        ws = wb[sheet]

        # Check if row is valid
        if row < 1:
            wb.close()
            typer.secho(
                f"Error: Invalid row number: {row}. Row must be >= 1.",
                fg=typer.colors.RED,
                err=True,
            )
            raise typer.Exit(code=int(ErrorCode.ROW_NOT_FOUND))

        # Hide the row
        ws.row_dimensions[row].hidden = True
        wb.save(path)
        wb.close()

        typer.echo(f"Hid row {row} in sheet '{sheet}'")

    except typer.Exit:
        raise
    except Exception as e:
        typer.secho(f"Error: {e}", fg=typer.colors.RED, err=True)
        raise typer.Exit(code=1)


@row_app.command()
def unhide(
    path: Annotated[Path, typer.Argument(help="Path to the workbook file.")],
    sheet: Annotated[str, typer.Argument(help="Sheet name.")],
    row: Annotated[int, typer.Argument(help="Row number (1-based).")],
) -> None:
    """Unhide a row."""
    # Check if file exists
    if not path.exists():
        typer.secho(
            f"Error: File does not exist: {path}",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.FILE_DOES_NOT_EXIST))

    try:
        wb = openpyxl.load_workbook(path)

        # Check if sheet exists
        if sheet not in wb.sheetnames:
            wb.close()
            typer.secho(
                f"Error: Sheet not found: {sheet}",
                fg=typer.colors.RED,
                err=True,
            )
            raise typer.Exit(code=int(ErrorCode.SHEET_NOT_FOUND))

        ws = wb[sheet]

        # Check if row is valid
        if row < 1:
            wb.close()
            typer.secho(
                f"Error: Invalid row number: {row}. Row must be >= 1.",
                fg=typer.colors.RED,
                err=True,
            )
            raise typer.Exit(code=int(ErrorCode.ROW_NOT_FOUND))

        # Unhide the row
        ws.row_dimensions[row].hidden = False
        wb.save(path)
        wb.close()

        typer.echo(f"Unhid row {row} in sheet '{sheet}'")

    except typer.Exit:
        raise
    except Exception as e:
        typer.secho(f"Error: {e}", fg=typer.colors.RED, err=True)
        raise typer.Exit(code=1)


@col_app.command()
def hide(
    path: Annotated[Path, typer.Argument(help="Path to the workbook file.")],
    sheet: Annotated[str, typer.Argument(help="Sheet name.")],
    column: Annotated[str, typer.Argument(help="Column letter (e.g., A, B, C).")],
) -> None:
    """Hide a column."""
    # Check if file exists
    if not path.exists():
        typer.secho(
            f"Error: File does not exist: {path}",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.FILE_DOES_NOT_EXIST))

    try:
        wb = openpyxl.load_workbook(path)

        # Check if sheet exists
        if sheet not in wb.sheetnames:
            wb.close()
            typer.secho(
                f"Error: Sheet not found: {sheet}",
                fg=typer.colors.RED,
                err=True,
            )
            raise typer.Exit(code=int(ErrorCode.SHEET_NOT_FOUND))

        ws = wb[sheet]

        # Normalize column letter to uppercase
        column = column.upper()

        # Hide the column
        ws.column_dimensions[column].hidden = True
        wb.save(path)
        wb.close()

        typer.echo(f"Hid column {column} in sheet '{sheet}'")

    except typer.Exit:
        raise
    except Exception as e:
        typer.secho(f"Error: {e}", fg=typer.colors.RED, err=True)
        raise typer.Exit(code=1)


@col_app.command()
def unhide(
    path: Annotated[Path, typer.Argument(help="Path to the workbook file.")],
    sheet: Annotated[str, typer.Argument(help="Sheet name.")],
    column: Annotated[str, typer.Argument(help="Column letter (e.g., A, B, C).")],
) -> None:
    """Unhide a column."""
    # Check if file exists
    if not path.exists():
        typer.secho(
            f"Error: File does not exist: {path}",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.FILE_DOES_NOT_EXIST))

    try:
        wb = openpyxl.load_workbook(path)

        # Check if sheet exists
        if sheet not in wb.sheetnames:
            wb.close()
            typer.secho(
                f"Error: Sheet not found: {sheet}",
                fg=typer.colors.RED,
                err=True,
            )
            raise typer.Exit(code=int(ErrorCode.SHEET_NOT_FOUND))

        ws = wb[sheet]

        # Normalize column letter to uppercase
        column = column.upper()

        # Unhide the column
        ws.column_dimensions[column].hidden = False
        wb.save(path)
        wb.close()

        typer.echo(f"Unhid column {column} in sheet '{sheet}'")

    except typer.Exit:
        raise
    except Exception as e:
        typer.secho(f"Error: {e}", fg=typer.colors.RED, err=True)
        raise typer.Exit(code=1)


@row_app.command()
def width(
    path: Annotated[Path, typer.Argument(help="Path to the workbook file.")],
    sheet: Annotated[str, typer.Argument(help="Sheet name.")],
    row: Annotated[int, typer.Argument(help="Row number (1-based).")],
    height: Annotated[float, typer.Argument(help="Height in points.")],
) -> None:
    """Set row height."""
    # Check if file exists
    if not path.exists():
        typer.secho(
            f"Error: File does not exist: {path}",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.FILE_DOES_NOT_EXIST))

    try:
        wb = openpyxl.load_workbook(path)

        # Check if sheet exists
        if sheet not in wb.sheetnames:
            wb.close()
            typer.secho(
                f"Error: Sheet not found: {sheet}",
                fg=typer.colors.RED,
                err=True,
            )
            raise typer.Exit(code=int(ErrorCode.SHEET_NOT_FOUND))

        ws = wb[sheet]

        # Check if row is valid
        if row < 1:
            wb.close()
            typer.secho(
                f"Error: Invalid row number: {row}. Row must be >= 1.",
                fg=typer.colors.RED,
                err=True,
            )
            raise typer.Exit(code=int(ErrorCode.ROW_NOT_FOUND))

        # Set row height
        ws.row_dimensions[row].height = height
        wb.save(path)
        wb.close()

        typer.echo(f"Set row {row} height to {height} in sheet '{sheet}'")

    except typer.Exit:
        raise
    except Exception as e:
        typer.secho(f"Error: {e}", fg=typer.colors.RED, err=True)
        raise typer.Exit(code=1)


@row_app.command()
def auto(
    path: Annotated[Path, typer.Argument(help="Path to the workbook file.")],
    sheet: Annotated[str, typer.Argument(help="Sheet name.")],
    row: Annotated[int, typer.Argument(help="Row number (1-based).")],
) -> None:
    """Auto-fit row height."""
    # Check if file exists
    if not path.exists():
        typer.secho(
            f"Error: File does not exist: {path}",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.FILE_DOES_NOT_EXIST))

    try:
        wb = openpyxl.load_workbook(path)

        # Check if sheet exists
        if sheet not in wb.sheetnames:
            wb.close()
            typer.secho(
                f"Error: Sheet not found: {sheet}",
                fg=typer.colors.RED,
                err=True,
            )
            raise typer.Exit(code=int(ErrorCode.SHEET_NOT_FOUND))

        ws = wb[sheet]

        # Check if row is valid
        if row < 1:
            wb.close()
            typer.secho(
                f"Error: Invalid row number: {row}. Row must be >= 1.",
                fg=typer.colors.RED,
                err=True,
            )
            raise typer.Exit(code=int(ErrorCode.ROW_NOT_FOUND))

        # Auto-fit row height (set to None)
        ws.row_dimensions[row].height = None
        wb.save(path)
        wb.close()

        typer.echo(f"Auto-fitted row {row} height in sheet '{sheet}'")

    except typer.Exit:
        raise
    except Exception as e:
        typer.secho(f"Error: {e}", fg=typer.colors.RED, err=True)
        raise typer.Exit(code=1)


@col_app.command()
def width(
    path: Annotated[Path, typer.Argument(help="Path to the workbook file.")],
    sheet: Annotated[str, typer.Argument(help="Sheet name.")],
    column: Annotated[str, typer.Argument(help="Column letter (e.g., A, B, C).")],
    col_width: Annotated[float, typer.Argument(help="Width in points.")],
) -> None:
    """Set column width."""
    # Check if file exists
    if not path.exists():
        typer.secho(
            f"Error: File does not exist: {path}",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.FILE_DOES_NOT_EXIST))

    try:
        wb = openpyxl.load_workbook(path)

        # Check if sheet exists
        if sheet not in wb.sheetnames:
            wb.close()
            typer.secho(
                f"Error: Sheet not found: {sheet}",
                fg=typer.colors.RED,
                err=True,
            )
            raise typer.Exit(code=int(ErrorCode.SHEET_NOT_FOUND))

        ws = wb[sheet]

        # Normalize column letter to uppercase
        column = column.upper()

        # Set column width
        ws.column_dimensions[column].width = col_width
        wb.save(path)
        wb.close()

        typer.echo(f"Set column {column} width to {col_width} in sheet '{sheet}'")

    except typer.Exit:
        raise
    except Exception as e:
        typer.secho(f"Error: {e}", fg=typer.colors.RED, err=True)
        raise typer.Exit(code=1)


@col_app.command()
def auto(
    path: Annotated[Path, typer.Argument(help="Path to the workbook file.")],
    sheet: Annotated[str, typer.Argument(help="Sheet name.")],
    column: Annotated[str, typer.Argument(help="Column letter (e.g., A, B, C).")],
) -> None:
    """Auto-fit column width."""
    # Check if file exists
    if not path.exists():
        typer.secho(
            f"Error: File does not exist: {path}",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.FILE_DOES_NOT_EXIST))

    try:
        wb = openpyxl.load_workbook(path)

        # Check if sheet exists
        if sheet not in wb.sheetnames:
            wb.close()
            typer.secho(
                f"Error: Sheet not found: {sheet}",
                fg=typer.colors.RED,
                err=True,
            )
            raise typer.Exit(code=int(ErrorCode.SHEET_NOT_FOUND))

        ws = wb[sheet]

        # Normalize column letter to uppercase
        column = column.upper()

        # Auto-fit column width (use bestFit=True with width=0)
        ws.column_dimensions[column].bestFit = True
        ws.column_dimensions[column].width = 0
        wb.save(path)
        wb.close()

        typer.echo(f"Auto-fitted column {column} width in sheet '{sheet}'")

    except typer.Exit:
        raise
    except Exception as e:
        typer.secho(f"Error: {e}", fg=typer.colors.RED, err=True)
        raise typer.Exit(code=1)
