"""Style operations CLI commands for Excel cells."""

# Note: This command uses openpyxl directly and bypasses the Engine abstraction.
# It works with OpenpyxlEngine but not with XlwingsEngine.

from __future__ import annotations

import re
from pathlib import Path
from typing import Annotated, Optional

import openpyxl
import openpyxl.styles
import typer

from xlforge.core.errors import ErrorCode, XlforgeError

style_app = typer.Typer(help="Style operations for Excel cells.")


def _is_valid_hex_color(color: str) -> bool:
    """Check if a color string is a valid hex color (#RRGGBB or RRGGBB)."""
    if color.startswith("#"):
        color = color[1:]
    return bool(re.match(r"^[0-9A-Fa-f]{6}$", color))


@style_app.command()
def set(
    path: Annotated[Path, typer.Argument(help="Path to the workbook file.")],
    sheet: Annotated[str, typer.Argument(help="Sheet name.")],
    coord: Annotated[str, typer.Argument(help="Cell coordinate (e.g., A1).")],
    bold: Annotated[
        bool, typer.Option("--bold", "-b", help="Set cell to bold.")
    ] = False,
    italic: Annotated[
        bool, typer.Option("--italic", "-i", help="Set cell to italic.")
    ] = False,
    color: Annotated[
        Optional[str], typer.Option("--color", "-c", help="Font color as hex (#RRGGBB or RRGGBB).")
    ] = None,
) -> None:
    """Set cell style (bold, italic, color)."""
    # Check if file exists
    if not path.exists():
        typer.secho(
            f"Error: File does not exist: {path}",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.FILE_DOES_NOT_EXIST))

    # Validate color if provided
    if color is not None and not _is_valid_hex_color(color):
        typer.secho(
            f"Error: Invalid color format: {color}. Use #RRGGBB or RRGGBB.",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.INVALID_STYLE_STRING))

    try:
        wb = openpyxl.load_workbook(path)

        # Check if sheet exists
        if sheet not in wb.sheetnames:
            wb.close()
            typer.secho(
                f"Error: Sheet not found: {sheet}",
                fg=typer.colors.RED,
                err=True,
            )
            raise typer.Exit(code=int(ErrorCode.SHEET_NOT_FOUND))

        ws = wb[sheet]

        # Check if cell exists (will auto-create if not)
        cell = ws[coord]

        # Get existing font properties
        existing_font = cell.font
        font_kwargs = {
            "bold": bold if bold else existing_font.bold,
            "italic": italic if italic else existing_font.italic,
            "name": existing_font.name,
            "size": existing_font.size,
        }

        # Handle color
        if color is not None:
            # Remove # if present and prepend FF for full opacity (ARGB format)
            rgb_color = color.lstrip("#")
            font_kwargs["color"] = "FF" + rgb_color
        else:
            font_kwargs["color"] = existing_font.color

        # Create new font with applied styles
        new_font = openpyxl.styles.Font(**font_kwargs)
        cell.font = new_font

        wb.save(path)
        wb.close()

        # Build success message
        changes = []
        if bold:
            changes.append("bold")
        if italic:
            changes.append("italic")
        if color is not None:
            changes.append(f"color {color}")

        if changes:
            typer.echo(f"Applied {', '.join(changes)} to {coord} on sheet '{sheet}'")
        else:
            typer.echo(f"No style changes specified for {coord} on sheet '{sheet}'")

    except XlforgeError:
        raise
    except typer.Exit:
        raise
    except Exception as e:
        typer.secho(f"Error: {e}", fg=typer.colors.RED, err=True)
        raise typer.Exit(code=1)


@style_app.command()
def number_format(
    path: Annotated[Path, typer.Argument(help="Path to the workbook file.")],
    sheet: Annotated[str, typer.Argument(help="Sheet name.")],
    coord: Annotated[str, typer.Argument(help="Cell coordinate (e.g., A1).")],
    format: Annotated[str, typer.Argument(help="Number format (e.g., 0.00, $#,##0.00).")],
) -> None:
    """Set cell number format."""
    # Check if file exists
    if not path.exists():
        typer.secho(
            f"Error: File does not exist: {path}",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.FILE_DOES_NOT_EXIST))

    try:
        wb = openpyxl.load_workbook(path)

        # Check if sheet exists
        if sheet not in wb.sheetnames:
            wb.close()
            typer.secho(
                f"Error: Sheet not found: {sheet}",
                fg=typer.colors.RED,
                err=True,
            )
            raise typer.Exit(code=int(ErrorCode.SHEET_NOT_FOUND))

        ws = wb[sheet]
        cell = ws[coord]

        # Set number format
        cell.number_format = format

        wb.save(path)
        wb.close()

        typer.echo(f"Set number format '{format}' for {coord} on sheet '{sheet}'")

    except XlforgeError:
        raise
    except typer.Exit:
        raise
    except Exception as e:
        typer.secho(f"Error: {e}", fg=typer.colors.RED, err=True)
        raise typer.Exit(code=int(ErrorCode.INVALID_NUMBER_FORMAT))


@style_app.command()
def font(
    path: Annotated[Path, typer.Argument(help="Path to the workbook file.")],
    sheet: Annotated[str, typer.Argument(help="Sheet name.")],
    coord: Annotated[str, typer.Argument(help="Cell coordinate (e.g., A1).")],
    name: Annotated[
        Optional[str], typer.Option("--name", "-n", help="Font name (e.g., Arial, Calibri).")
    ] = None,
    size: Annotated[
        Optional[int], typer.Option("--size", "-s", help="Font size (e.g., 10, 12, 14).")
    ] = None,
) -> None:
    """Set cell font (name and size)."""
    # Check if file exists
    if not path.exists():
        typer.secho(
            f"Error: File does not exist: {path}",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.FILE_DOES_NOT_EXIST))

    # Check that at least one option is provided
    if name is None and size is None:
        typer.secho(
            "Error: Must specify at least one of --name or --size.",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=1)

    try:
        wb = openpyxl.load_workbook(path)

        # Check if sheet exists
        if sheet not in wb.sheetnames:
            wb.close()
            typer.secho(
                f"Error: Sheet not found: {sheet}",
                fg=typer.colors.RED,
                err=True,
            )
            raise typer.Exit(code=int(ErrorCode.SHEET_NOT_FOUND))

        ws = wb[sheet]
        cell = ws[coord]

        # Get existing font properties
        existing_font = cell.font

        # Create new font with applied styles
        font_kwargs = {
            "bold": existing_font.bold,
            "italic": existing_font.italic,
            "name": name if name is not None else existing_font.name,
            "size": size if size is not None else existing_font.size,
            "color": existing_font.color,
        }

        new_font = openpyxl.styles.Font(**font_kwargs)
        cell.font = new_font

        wb.save(path)
        wb.close()

        # Build success message
        changes = []
        if name is not None:
            changes.append(f"name '{name}'")
        if size is not None:
            changes.append(f"size {size}")

        typer.echo(f"Set font {', '.join(changes)} for {coord} on sheet '{sheet}'")

    except XlforgeError:
        raise
    except typer.Exit:
        raise
    except Exception as e:
        typer.secho(f"Error: {e}", fg=typer.colors.RED, err=True)
        raise typer.Exit(code=1)
