"""Table operations CLI commands."""

# Note: This command uses openpyxl directly and bypasses the Engine abstraction.
# It works with OpenpyxlEngine but not with XlwingsEngine.
# For native Excel tables with auto-expand behavior, use the --native flag with win32com.

from __future__ import annotations

from importlib.util import find_spec
from pathlib import Path
from typing import Annotated, Optional

import openpyxl
import typer
from openpyxl.worksheet.table import Table

from xlforge.core.errors import ErrorCode

table_app = typer.Typer(help="Table operations for Excel workbooks.")

# Excel ListObject SourceType constants
_XL_SRC_RANGE = 1


def _is_xlwings_available() -> bool:
    """Check if xlwings is available (Excel integration possible)."""
    return find_spec("xlwings") is not None


def _validate_table_name(name: str) -> bool:
    """Validate table name according to Excel rules."""
    # Table names must not be empty, exceed 255 characters,
    # or contain certain special characters
    if not name or len(name) > 255:
        return False
    # Table names cannot contain: : \ / ? * [ ]
    invalid_chars = {":", "\\", "/", "?", "*", "[", "]"}
    return not any(c in invalid_chars for c in name)


def _create_openpyxl_table(
    path: Path,
    sheet: str,
    range_ref: str,
    name: str | None,
) -> None:
    """Create a table using openpyxl (no auto-expand behavior)."""
    try:
        wb = openpyxl.load_workbook(path)

        # Check if sheet exists
        if sheet not in wb.sheetnames:
            typer.secho(
                f"Error: Sheet not found: {sheet}",
                fg=typer.colors.RED,
                err=True,
            )
            wb.close()
            raise typer.Exit(code=int(ErrorCode.SHEET_NOT_FOUND))

        ws = wb[sheet]

        # If no name provided, use a default name
        if not name:
            # Generate a unique name
            base_name = f"Table{len(ws.tables) + 1}"
            counter = 1
            while base_name in ws.tables:
                counter += 1
                base_name = f"Table{len(ws.tables) + counter}"
            name = base_name

        # Check if table already exists with this name
        if name in ws.tables:
            typer.secho(
                f"Error: Table '{name}' already exists in sheet '{sheet}'",
                fg=typer.colors.RED,
                err=True,
            )
            wb.close()
            raise typer.Exit(code=int(ErrorCode.TABLE_ALREADY_EXISTS))

        # Create the table
        table = Table(displayName=name, ref=range_ref)
        ws.add_table(table)

        wb.save(path)
        wb.close()

        typer.echo(f"Created table '{name}' at {range_ref} in sheet '{sheet}'")

    except typer.Exit:
        raise
    except Exception as e:
        typer.secho(f"Error: {e}", fg=typer.colors.RED, err=True)
        raise typer.Exit(code=int(ErrorCode.GENERAL_ERROR))


def _create_native_table(
    path: Path,
    sheet: str,
    range_ref: str,
    name: str | None,
) -> None:
    """Create a native Excel table using win32com (with auto-expand behavior).

    Native Excel tables (ListObjects) automatically expand when data is added
    below the table.
    """
    if not _is_xlwings_available():
        typer.secho(
            "Error: Native table creation requires Excel via win32com.",
            fg=typer.colors.RED,
            err=True,
        )
        typer.secho(
            "Use openpyxl mode (without --native) or install xlwings.",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.FEATURE_UNAVAILABLE))

    # Note: We don't check for file open in Excel here because win32com's
    # Dispatch("Excel.Application") can connect to an already-open instance

    excel = None
    try:
        import win32com.client  # type: ignore[import-untyped]

        # Use win32com directly to create native Excel table
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb_com = excel.Workbooks.Open(str(path.absolute()))

        # Get the sheet
        try:
            ws = wb_com.Sheets(sheet)
        except Exception:
            typer.secho(
                f"Error: Sheet not found: {sheet}",
                fg=typer.colors.RED,
                err=True,
            )
            raise typer.Exit(code=int(ErrorCode.SHEET_NOT_FOUND))

        # Get the range
        try:
            rng = ws.Range(range_ref)
        except Exception:
            typer.secho(
                f"Error: Invalid range: {range_ref}",
                fg=typer.colors.RED,
                err=True,
            )
            raise typer.Exit(code=int(ErrorCode.INVALID_SYNTAX))

        # Create the table using ListObjects.Add
        # SourceType=1 (xlSrcRange) means the table is based on a range
        list_objects = ws.ListObjects
        tbl = list_objects.Add(1, rng)

        # Set the table name if provided
        if name:
            tbl.Name = name

        # Save and close workbook
        wb_com.Save()
        wb_com.Close()

        table_name = name if name else tbl.Name
        typer.echo(
            f"Created native Excel table '{table_name}' at {range_ref} in sheet '{sheet}'"
        )
        typer.echo("Note: Native Excel tables auto-expand when data is added below.")

    except typer.Exit:
        raise
    except Exception as e:
        typer.secho(f"Error: {e}", fg=typer.colors.RED, err=True)
        raise typer.Exit(code=int(ErrorCode.GENERAL_ERROR))
    finally:
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass


@table_app.command()
def create(
    path: Annotated[Path, typer.Argument(help="Path to the workbook file.")],
    sheet: Annotated[str, typer.Argument(help="Sheet name containing the range.")],
    range_ref: Annotated[str, typer.Argument(help="Range reference (e.g., A1:C10).")],
    name: Annotated[
        str | None, typer.Option("--name", "-n", help="Name for the table.")
    ] = None,
    native: Annotated[
        bool,
        typer.Option(
            "--native",
            help="Use native Excel tables via win32com (requires Excel installed).",
        ),
    ] = False,
) -> None:
    """Create an Excel table from a range.

    Use --native to create native Excel ListObjects with auto-expand behavior.
    Native tables require Excel to be installed.
    """
    # Check if file exists
    if not path.exists():
        typer.secho(
            f"Error: File does not exist: {path}",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.FILE_DOES_NOT_EXIST))

    # Validate table name if provided
    if name and not _validate_table_name(name):
        typer.secho(
            f"Error: Invalid table name: {name}",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.INVALID_TABLE_NAME))

    if native:
        _create_native_table(path, sheet, range_ref, name)
    else:
        _create_openpyxl_table(path, sheet, range_ref, name)


@table_app.command("list")
def list_tables(
    path: Annotated[Path, typer.Argument(help="Path to the workbook file.")],
) -> None:
    """List all tables in the workbook."""
    # Check if file exists
    if not path.exists():
        typer.secho(
            f"Error: File does not exist: {path}",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.FILE_DOES_NOT_EXIST))

    try:
        wb = openpyxl.load_workbook(path)

        tables_found = False
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for tbl in ws.tables.values():
                tables_found = True
                typer.echo(f"Sheet: {sheet_name}, Table: {tbl.name}, Range: {tbl.ref}")

        if not tables_found:
            typer.echo("No tables found in workbook.")

        wb.close()

    except typer.Exit:
        raise
    except Exception as e:
        typer.secho(f"Error: {e}", fg=typer.colors.RED, err=True)
        raise typer.Exit(code=int(ErrorCode.GENERAL_ERROR))


@table_app.command()
def delete(
    path: Annotated[Path, typer.Argument(help="Path to the workbook file.")],
    name: Annotated[str, typer.Argument(help="Name of the table to delete.")],
    sheet: Annotated[
        Optional[str],
        typer.Option("--sheet", "-s", help="Sheet name containing the table."),
    ] = None,
) -> None:
    """Delete a table from the workbook."""
    # Check if file exists
    if not path.exists():
        typer.secho(
            f"Error: File does not exist: {path}",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.FILE_DOES_NOT_EXIST))

    try:
        wb = openpyxl.load_workbook(path)

        # If sheet specified, only look in that sheet
        if sheet:
            if sheet not in wb.sheetnames:
                typer.secho(
                    f"Error: Sheet not found: {sheet}",
                    fg=typer.colors.RED,
                    err=True,
                )
                wb.close()
                raise typer.Exit(code=int(ErrorCode.SHEET_NOT_FOUND))

            ws = wb[sheet]
            if name not in ws.tables:
                typer.secho(
                    f"Error: Table '{name}' not found in sheet '{sheet}'",
                    fg=typer.colors.RED,
                    err=True,
                )
                wb.close()
                raise typer.Exit(code=int(ErrorCode.TABLE_NOT_FOUND))

            del ws.tables[name]
            wb.save(path)
            wb.close()
            typer.echo(f"Deleted table '{name}' from sheet '{sheet}'")
        else:
            # Search all sheets for the table
            found = False
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                if name in ws.tables:
                    del ws.tables[name]
                    found = True
                    break

            if not found:
                typer.secho(
                    f"Error: Table '{name}' not found in workbook",
                    fg=typer.colors.RED,
                    err=True,
                )
                wb.close()
                raise typer.Exit(code=int(ErrorCode.TABLE_NOT_FOUND))

            wb.save(path)
            wb.close()
            typer.echo(f"Deleted table '{name}'")

    except typer.Exit:
        raise
    except Exception as e:
        typer.secho(f"Error: {e}", fg=typer.colors.RED, err=True)
        raise typer.Exit(code=int(ErrorCode.GENERAL_ERROR))
