"""Data validation CLI commands for Excel workbooks."""

# Note: This command uses openpyxl directly and bypasses the Engine abstraction.
# It works with OpenpyxlEngine but not with XlwingsEngine.

from __future__ import annotations

from pathlib import Path
from typing import Annotated, Optional

import openpyxl
import typer

from xlforge.core.errors import ErrorCode, XlforgeError

validation_app = typer.Typer(help="Data validation operations for Excel workbooks.")

VALIDATION_TYPES = ["list", "whole", "decimal", "date", "time", "textLength", "custom"]


@validation_app.command()
def add(
    path: Annotated[Path, typer.Argument(help="Path to the workbook file.")],
    sheet: Annotated[str, typer.Argument(help="Sheet name.")],
    range: Annotated[str, typer.Argument(help="Cell range (e.g., A1:A10).")],
    type: Annotated[str, typer.Option("--type", "-t", help=f"Validation type: {', '.join(VALIDATION_TYPES)}.")],
    formula1: Annotated[
        Optional[str],
        typer.Option("--formula1", "-f1", help="Primary formula or value for validation."),
    ] = None,
    formula2: Annotated[
        Optional[str],
        typer.Option("--formula2", "-f2", help="Secondary formula for 'between' operator (e.g., max value)."),
    ] = None,
) -> None:
    """Add data validation to a range."""
    # Check if file exists
    if not path.exists():
        typer.secho(
            f"Error: File does not exist: {path}",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.FILE_DOES_NOT_EXIST))

    # Validate type
    type_lower = type.lower()
    if type_lower not in VALIDATION_TYPES:
        typer.secho(
            f"Error: Invalid validation type: {type}. Valid types: {', '.join(VALIDATION_TYPES)}",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.VALIDATION_TYPE_NOT_SUPPORTED))

    # Validate formula1 is provided
    if formula1 is None:
        typer.secho(
            f"Error: --formula1 is required for {type} validation.",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.INVALID_FORMULA_SYNTAX))

    try:
        wb = openpyxl.load_workbook(path)

        # Check if sheet exists
        if sheet not in wb.sheetnames:
            wb.close()
            typer.secho(
                f"Error: Sheet not found: {sheet}",
                fg=typer.colors.RED,
                err=True,
            )
            raise typer.Exit(code=int(ErrorCode.SHEET_NOT_FOUND))

        ws = wb[sheet]

        # Create data validation
        dv = openpyxl.worksheet.datavalidation.DataValidation(
            type=type_lower,
            formula1=formula1,
            formula2=formula2,
            allow_blank=True,
        )
        dv.add(range)
        ws.add_data_validation(dv)

        wb.save(path)
        wb.close()

        typer.echo(f"Added {type} validation to range {range} on sheet '{sheet}'")
        if formula2:
            typer.echo(f"  Formula1: {formula1}, Formula2: {formula2}")
        else:
            typer.echo(f"  Formula1: {formula1}")

    except XlforgeError:
        raise
    except typer.Exit:
        raise
    except Exception as e:
        typer.secho(f"Error: {e}", fg=typer.colors.RED, err=True)
        raise typer.Exit(code=1)


@validation_app.command()
def remove(
    path: Annotated[Path, typer.Argument(help="Path to the workbook file.")],
    sheet: Annotated[str, typer.Argument(help="Sheet name.")],
    range: Annotated[str, typer.Argument(help="Cell range (e.g., A1:A10).")],
) -> None:
    """Remove data validation from a range."""
    # Check if file exists
    if not path.exists():
        typer.secho(
            f"Error: File does not exist: {path}",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.FILE_DOES_NOT_EXIST))

    try:
        wb = openpyxl.load_workbook(path)

        # Check if sheet exists
        if sheet not in wb.sheetnames:
            wb.close()
            typer.secho(
                f"Error: Sheet not found: {sheet}",
                fg=typer.colors.RED,
                err=True,
            )
            raise typer.Exit(code=int(ErrorCode.SHEET_NOT_FOUND))

        ws = wb[sheet]

        # Find and remove the validation for this range
        initial_count = len(ws.data_validations.dataValidation)
        removed = False

        # We need to filter out validations that overlap with the given range
        new_validations = []
        for dv in ws.data_validations.dataValidation:
            # Check if the ranges overlap
            if dv.sqref and range.upper() in str(dv.sqref).upper():
                removed = True
            else:
                new_validations.append(dv)

        if removed:
            ws.data_validations.dataValidation = new_validations
            wb.save(path)
            wb.close()
            typer.echo(f"Removed data validation from range {range} on sheet '{sheet}'")
        else:
            wb.close()
            typer.secho(
                f"Error: No data validation found for range {range} on sheet '{sheet}'",
                fg=typer.colors.RED,
                err=True,
            )
            raise typer.Exit(code=int(ErrorCode.VALIDATION_TYPE_NOT_SUPPORTED))

    except XlforgeError:
        raise
    except typer.Exit:
        raise
    except Exception as e:
        typer.secho(f"Error: {e}", fg=typer.colors.RED, err=True)
        raise typer.Exit(code=1)


@validation_app.command()
def list(
    path: Annotated[Path, typer.Argument(help="Path to the workbook file.")],
    sheet: Annotated[str, typer.Argument(help="Sheet name.")],
) -> None:
    """List all data validations in a sheet."""
    # Check if file exists
    if not path.exists():
        typer.secho(
            f"Error: File does not exist: {path}",
            fg=typer.colors.RED,
            err=True,
        )
        raise typer.Exit(code=int(ErrorCode.FILE_DOES_NOT_EXIST))

    try:
        wb = openpyxl.load_workbook(path)

        # Check if sheet exists
        if sheet not in wb.sheetnames:
            wb.close()
            typer.secho(
                f"Error: Sheet not found: {sheet}",
                fg=typer.colors.RED,
                err=True,
            )
            raise typer.Exit(code=int(ErrorCode.SHEET_NOT_FOUND))

        ws = wb[sheet]

        validations = ws.data_validations.dataValidation

        if not validations:
            typer.echo(f"No data validations found on sheet '{sheet}'")
        else:
            typer.echo(f"Data validations on sheet '{sheet}':")
            typer.echo()
            for i, dv in enumerate(validations, 1):
                typer.echo(f"{i}. Type: {dv.type}")
                if dv.formula1:
                    typer.echo(f"   Formula1: {dv.formula1}")
                if dv.formula2:
                    typer.echo(f"   Formula2: {dv.formula2}")
                if dv.operator:
                    typer.echo(f"   Operator: {dv.operator}")
                if dv.sqref:
                    typer.echo(f"   Range: {dv.sqref}")
                typer.echo()

        wb.close()

    except XlforgeError:
        raise
    except typer.Exit:
        raise
    except Exception as e:
        typer.secho(f"Error: {e}", fg=typer.colors.RED, err=True)
        raise typer.Exit(code=1)
