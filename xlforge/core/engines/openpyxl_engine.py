"""Openpyxl-based engine for xlforge."""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl

from xlforge.core.engines.base import Engine
from xlforge.core.types.cell_value import CellValue
from xlforge.core.types.value_type import ValueType


class OpenpyxlEngine(Engine):
    """Openpyxl-based engine for headless environments."""

    def __init__(self) -> None:
        self._workbooks: dict[Path, openpyxl.Workbook] = {}

    def open(
        self, path: Path, *, read_only: bool = False, data_only: bool = True
    ) -> None:
        """Open a workbook with openpyxl."""
        if path in self._workbooks:
            return
        wb = openpyxl.load_workbook(path, read_only=read_only, data_only=data_only)
        self._workbooks[path] = wb

    def close(self, path: Path) -> None:
        """Close a workbook."""
        if path in self._workbooks:
            self._workbooks[path].close()
            del self._workbooks[path]

    def list_sheets(self, path: Path) -> list[str]:
        """List all sheet names."""
        wb = self._workbooks.get(path)
        if wb is None:
            raise FileNotFoundError(f"Workbook not open: {path}")
        return wb.sheetnames

    def get_cell(self, path: Path, sheet: str, coord: str) -> CellValue:
        """Get cell value using openpyxl."""
        wb = self._workbooks.get(path)
        if wb is None:
            raise FileNotFoundError(f"Workbook not open: {path}")
        ws = wb[sheet]
        cell = ws[coord]
        return self._cell_to_value(cell)

    def set_cell(self, path: Path, sheet: str, coord: str, value: CellValue) -> None:
        """Set cell value using openpyxl."""
        wb = self._workbooks.get(path)
        if wb is None:
            raise FileNotFoundError(f"Workbook not open: {path}")
        ws = wb[sheet]
        ws[coord] = self._value_to_cell(value)

    def get_range(self, path: Path, sheet: str, coord: str) -> list[list[CellValue]]:
        """Get range values as 2D array."""
        wb = self._workbooks.get(path)
        if wb is None:
            raise FileNotFoundError(f"Workbook not open: {path}")
        ws = wb[sheet]
        rng = ws[coord]
        return [[self._cell_to_value(cell) for cell in row] for row in rng]

    def set_range(
        self, path: Path, sheet: str, coord: str, values: list[list[CellValue]]
    ) -> None:
        """Set range values from 2D array."""
        wb = self._workbooks.get(path)
        if wb is None:
            raise FileNotFoundError(f"Workbook not open: {path}")
        ws = wb[sheet]
        rng = ws[coord]
        for i, row in enumerate(values):
            for j, value in enumerate(row):
                rng[i][j].value = self._value_to_cell(value)

    def create_sheet(self, path: Path, sheet: str) -> None:
        """Create a new sheet."""
        wb = self._workbooks.get(path)
        if wb is None:
            raise FileNotFoundError(f"Workbook not open: {path}")
        wb.create_sheet(title=sheet)

    def delete_sheet(self, path: Path, sheet: str) -> None:
        """Delete a sheet."""
        wb = self._workbooks.get(path)
        if wb is None:
            raise FileNotFoundError(f"Workbook not open: {path}")
        if sheet in wb.sheetnames:
            del wb[sheet]

    def rename_sheet(self, path: Path, old_name: str, new_name: str) -> None:
        """Rename a sheet."""
        wb = self._workbooks.get(path)
        if wb is None:
            raise FileNotFoundError(f"Workbook not open: {path}")
        ws = wb[old_name]
        ws.title = new_name

    def save(self, path: Path) -> None:
        """Save a workbook."""
        wb = self._workbooks.get(path)
        if wb is None:
            raise FileNotFoundError(f"Workbook not open: {path}")
        wb.save(path)

    def sheet_exists(self, path: Path, sheet: str) -> bool:
        """Check if a sheet exists."""
        wb = self._workbooks.get(path)
        if wb is None:
            return False
        return sheet in wb.sheetnames

    def get_sheet_dimensions(self, path: Path, sheet: str) -> str:
        """Get the used range dimensions for a sheet."""
        wb = self._workbooks.get(path)
        if wb is None:
            raise FileNotFoundError(f"Workbook not open: {path}")
        ws = wb[sheet]
        return ws.dimensions

    def cell_exists(self, path: Path, sheet: str, coord: str) -> bool:
        """Check if a cell exists within the used range."""
        wb = self._workbooks.get(path)
        if wb is None:
            raise FileNotFoundError(f"Workbook not open: {path}")
        ws = wb[sheet]
        # Check if coord is within the sheet's used range
        dimensions = ws.dimensions
        if not dimensions:
            return False
        try:
            # openpyxl can access the cell directly
            ws[coord]
            return True
        except (KeyError, AttributeError):
            return False

    def _cell_to_value(self, cell: openpyxl.cell.cell.Cell) -> CellValue:
        """Convert openpyxl cell to CellValue."""
        if cell.value is None:
            return CellValue(raw=None, type=ValueType.EMPTY)

        if isinstance(cell.value, bool):
            return CellValue(raw=cell.value, type=ValueType.BOOL)

        if isinstance(cell.value, (int, float)):
            return CellValue(raw=cell.value, type=ValueType.NUMBER)

        if isinstance(cell.value, str):
            if cell.value.startswith("="):
                return CellValue(raw=cell.value, type=ValueType.FORMULA)
            return CellValue(raw=cell.value, type=ValueType.STRING)

        # openpyxl handles datetime objects
        if hasattr(cell.value, "year") and hasattr(cell.value, "month"):
            return CellValue(raw=cell.value, type=ValueType.DATE)

        return CellValue(raw=str(cell.value), type=ValueType.STRING)

    def _value_to_cell(self, value: CellValue) -> Any:
        """Convert CellValue to openpyxl-compatible value."""
        if value.type == ValueType.EMPTY:
            return None
        return value.raw
